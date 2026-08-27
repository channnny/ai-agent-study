"""검토필요 시트 — 정제 결과에서 사람 확인이 필요한 셀만 추출.

결정적 파싱 부적합/소스 한계 케이스를 1353행에 묻지 않고 별도 시트로 명시 핸드오프.
- 최저 계열별/파싱: 인문·자연 다조건 → 사람이 직접 정제
- 진로 내부확인: 어디가에 A/B/C 환산표 미노출(소스 한계) → 모집요강 등 타소스 확인
- 5a 단계별: 1단계 배수/요소 매핑 스팟체크 권장

실행: python review.py [<정제.xlsx>]
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 정제 시트 컬럼(1-based): 대학명3 전형명5 모집단위7 / ①9→10 ②11 [원]교과12 ③13 [원]진로14 ④15 [원]비율16 5a17 5b18
COL = dict(대학=3, 전형=5, 모집=7, 원최저=9, 정1=10, 정2=11, 원진로=14, 정4=15, 원비율=16, 정5a=17)

# 항목별 (사유, 해결책)
POLICY = {
    "최저-파싱": ("최저 문구 비정형 — 규칙 A~J 어디에도 안 걸림", "원본 확인 후 N합M 수기 입력"),
    "최저-계열택1": ("인문/자연 계열 택1인데 계열별 값이 다름 — 모집단위 계열이 크롤 데이터에 없음",
                 "모집단위 계열(인문/자연) 판정 후 해당 값 선택. 계열 매핑표 확보 시 자동화 가능"),
    "최저-원본결손(부분)": ("어디가 원문이 도중에 끊겨 등급 컷 숫자가 없음(예: '…우수 2과목 평균등급' 에서 종료)",
                       "어디가 원본 자체가 결손임을 확인함. 모집요강 등 외부 소스 보완 필요"),
    "진로-내부확인": ("각주에 A/B/C 환산표 미노출(어디가 소스 한계)", "대학 모집요강/입학처 확인 또는 내부확인 유지"),
    "5a-단계별": ("단계 배수/요소 매핑 불확실", "원본 대조 스팟체크"),
}

HEAD_FILL = PatternFill("solid", fgColor="C55A11")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
LEFT = Alignment(vertical="top", wrap_text=True)
WIDTHS = [14, 16, 22, 13, 30, 44, 28, 34, 20]   # 항목·대학·전형·모집·사유·원본·현재·해결책·검토결과


def _flagged(r):
    """행 → (항목, 원본, 현재정제) 목록. 검토 필요한 셀만."""
    out = []
    v = lambda k: r[COL[k] - 1]
    s = lambda x: "" if x is None else str(x)
    if "검증필요" in s(v("정1")) or "검증필요" in s(v("정2")):
        raw = s(v("원최저"))
        # 원문에 숫자가 하나도 없거나 '평균등급' 뒤 컷이 비면 정제 규칙이 아니라 수집 문제
        결손 = not re.search(r"\d", raw) or bool(re.search(r"평균\s*등급\s*$", raw.strip()))
        if "계열택1" in s(v("정1")):
            sub = "최저-계열택1"
        else:
            sub = "최저-원본결손(부분)" if 결손 else "최저-파싱"
        out.append((sub, raw, f"①{s(v('정1'))} / ②{s(v('정2'))}"))
    # '내부확인'은 데이터랩스 정의상 유효 최종값(진로 반영O·A/B/C값 소스 미공개) → 검토필요 아님.
    if "검증필요" in s(v("정5a")):
        out.append(("5a-단계별", s(v("원비율")), s(v("정5a"))))
    res = []
    for sub, raw, cur in out:
        사유, 해결책 = POLICY.get(sub, ("", ""))
        res.append((sub, v("대학"), v("전형"), v("모집"), 사유, raw, cur, 해결책))
    return res


def build_review(in_path: Path, out_path: Path = None):
    wb = openpyxl.load_workbook(in_path, read_only=True)
    ws = wb["정제"]
    items = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[0]:
            items.extend(_flagged(r))
    wb.close()

    out = openpyxl.Workbook()
    sh = out.active; sh.title = "검토필요"
    sh.append(["항목", "대학명", "전형명", "모집단위명", "사유(왜 검토필요)", "[원본]", "[현재 정제]", "해결책(권장)", "검토결과(직접입력)"])
    for row in items:
        sh.append(list(row) + [""])
    for cell in sh[1]:
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in sh.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = LEFT
    for j, w in enumerate(WIDTHS, 1):
        sh.column_dimensions[get_column_letter(j)].width = w
    sh.freeze_panes = "A2"

    if out_path is None:
        out_path = Path(__file__).resolve().parents[2] / "output" / "검토필요.xlsx"
    out.save(out_path)
    # 항목별 집계
    from collections import Counter
    cnt = Counter(it[0] for it in items)
    return out_path, len(items), dict(cnt)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    in_path = Path(args[0]) if args else (Path(__file__).resolve().parents[2] / "output" / "전형정보_통합_정제.xlsx")
    out, n, cnt = build_review(in_path)
    print(f"저장: {out}")
    print(f"검토필요 {n}건: {cnt}")


if __name__ == "__main__":
    main()

"""오케스트레이션: 통합본 2시트 RAW → 시트3 '정제' 추가. (참고 양식·스타일 동일)

실행: python build.py [<통합본.xlsx>] [--limit=N]
기본 입력 = week10 test 통합본, 출력 = week11/output/<입력>_정제.xlsx.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from cols import c, S2_반영교과                      # noqa: E402
from jonghap import is_jonghap                        # noqa: E402
from gyogwa import refine_gyogwa                       # noqa: E402
from ratio import refine_5a, refine_5b                 # noqa: E402
from choi import refine_choi                           # noqa: E402
from jinro import refine_jinro                         # noqa: E402
from sources import raw_choi, raw_jinro, raw_ratio     # noqa: E402

# ── 시트3 헤더/스타일 (참고 산출물 그대로) ──
IDENT = ["adiga_selcntnm", "학년도", "대학명", "대학코드", "전형명", "전형코드", "모집단위명", "모집단위코드"]
GROUPS = [("기본정보 (RAW A~H)", 8), ("정제①② 수능최저", 3), ("정제③ 교과반영영역", 2),
          ("정제④ 진로 A/B/C", 2), ("정제 5a/5b 반영비율", 3)]
COLNAMES = IDENT + [
    "[원본] 최저학력기준(세부내용)", "[정제①] N합N", "[정제②] 영역조합",
    "[원본] 반영교과", "[정제③] 교과반영영역",
    "[원본] 진로선택+반영방법각주", "[정제④] 진로 A/B/C",
    "[원본] 반영비율/학년·요소", "[정제5a] 전형요소별비율", "[정제5b] 학년별/요소별비율"]

GRP_FILL = PatternFill("solid", fgColor="2E75B6")
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
RAW_FILL = PatternFill("solid", fgColor="FFF2CC")    # [원본] 연노랑
REF_FILL = PatternFill("solid", fgColor="E2EFDA")    # [정제] 연초록
GRP_FONT = Font(bold=True, color="FFFFFF", size=10)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
DATA_FONT = Font(size=9)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(vertical="top", wrap_text=True)
COL_WIDTHS = [20, 6, 16, 9, 24, 12, 14, 11, 44, 14, 30, 34, 14, 46, 22, 50, 30, 30]
RAW_COLS = {9, 12, 14, 16}
REF_COLS = {10, 11, 13, 15, 17, 18}

DEFAULT_IN = Path("/Users/channy/Documents/workspaces/ai/ai-agent-study/"
                  "week10/output/test_피드백/전형정보_통합.xlsx")


def refine_row(s1, s2):
    """RAW 두 행(values tuple) → (정제 18값, flags)."""
    ident = [s1[i] for i in range(8)]
    jh = is_jonghap(c(s1[4]), s2)
    flags = []
    n합, 영역 = refine_choi(s1, flags)
    row = ident + [
        raw_choi(s1), n합, 영역,
        c(s2[S2_반영교과]), refine_gyogwa(s2[S2_반영교과], jh),
        raw_jinro(s2), refine_jinro(s2, jh, flags),
        raw_ratio(s1, s2), refine_5a(s1, s2, jh, flags), refine_5b(s2, jh, flags)]
    return row, flags


def build(in_path: Path, out_path: Path = None, limit: int = None):
    wb = openpyxl.load_workbook(in_path)
    s1 = list(wb["전형일정및방법"].iter_rows(min_row=4, values_only=True))
    m2 = {r[0]: r for r in wb["전형요소"].iter_rows(min_row=4, values_only=True) if r[0]}
    if "정제" in wb.sheetnames:
        del wb["정제"]
    ws = wb.create_sheet("정제")
    # 헤더 2행 (그룹 병합 + 컬럼명)
    ci = 1
    for label, n in GROUPS:
        ws.cell(1, ci, label)
        if n > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + n - 1)
        for k in range(n):
            g = ws.cell(1, ci + k); g.fill = GRP_FILL; g.font = GRP_FONT; g.alignment = CEN
        ci += n
    for j, name in enumerate(COLNAMES, 1):
        cell = ws.cell(2, j, name); cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = CEN
    # 데이터
    flagcount, n = {}, 0
    rows = s1 if limit is None else s1[:limit]
    for r1 in rows:
        if not r1[0]:
            continue
        r2 = m2.get(r1[0], [None] * 55)
        vals, flags = refine_row(r1, r2)
        for f in flags:
            flagcount[f] = flagcount.get(f, 0) + 1
        for j, v in enumerate(vals, 1):
            cell = ws.cell(3 + n, j, v); cell.alignment = LEFT; cell.font = DATA_FONT
            if j in RAW_COLS:
                cell.fill = RAW_FILL
            elif j in REF_COLS:
                cell.fill = REF_FILL
        n += 1
    for j, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 34
    ws.freeze_panes = "I3"
    if out_path is None:
        out_dir = Path(__file__).resolve().parents[2] / "output"   # week11/output
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / (in_path.stem + "_정제.xlsx")
    wb.save(out_path)
    return out_path, n, flagcount


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    in_path = Path(args[0]) if args else DEFAULT_IN
    limit = next((int(a.split("=")[1]) for a in sys.argv[1:] if a.startswith("--limit=")), None)
    out, n, fc = build(in_path, limit=limit)
    print(f"저장: {out}")
    print(f"정제 {n}행")
    print("검증필요/플래그:", {k: v for k, v in sorted(fc.items(), key=lambda x: -x[1])})


if __name__ == "__main__":
    main()

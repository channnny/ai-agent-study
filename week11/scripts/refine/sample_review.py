"""정제 검수 샘플 시트 — 랜덤 N건을 [원본]↔[정제] 나란히 배치. 사람 눈검수용.

대학·전형유형(교과/종합/수능/논술 등) stratified 샘플. 시드 고정(재현 가능).
각 유닛 = 6종 정제항목별 한 행: 항목 / 원본 / 정제 / 판정(O·X) / 메모

실행: python sample_review.py [--n=30] [--seed=11]
"""
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parents[2] / "output"
SRC = OUT_DIR / "전형정보_통합_정제.xlsx"

# 정제 시트 컬럼(1-based) → 0-based idx
I = dict(sel=0, 학년도=1, 대학=2, 대학코드=3, 전형=4, 전형코드=5, 모집=6, 모집코드=7,
         원최저=8, 정1=9, 정2=10, 원교과=11, 정3=12, 원진로=13, 정4=14,
         원비율=15, 정5a=16, 정5b=17)

ITEMS = [
    ("① N합N", "원최저", "정1"),
    ("② 영역조합", "원최저", "정2"),
    ("③ 교과반영영역", "원교과", "정3"),
    ("④ 진로 A/B/C", "원진로", "정4"),
    ("5a 전형요소별비율", "원비율", "정5a"),
    ("5b 학년/요소별비율", "원비율", "정5b"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
UNIT_FILL = PatternFill("solid", fgColor="DDEBF7")
RAW_FILL = PatternFill("solid", fgColor="FFF2CC")
REF_FILL = PatternFill("solid", fgColor="E2EFDA")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
UNIT_FONT = Font(bold=True, size=10, color="1F4E78")
LEFT = Alignment(vertical="top", wrap_text=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = [18, 52, 30, 10, 26]     # 항목·[원본]·[정제]·판정·메모


def _type(전형명: str) -> str:
    t = 전형명 or ""
    for k in ("종합", "교과", "수능", "논술", "실기"):
        if k in t:
            return k
    return "기타"


def build(n: int = 30, seed: int = 11):
    wb = openpyxl.load_workbook(SRC, read_only=True)
    rows = [r for r in wb["정제"].iter_rows(min_row=3, values_only=True) if r and r[0]]
    wb.close()

    # 대학 × 전형유형 stratified: 그룹별로 순회하며 하나씩 뽑기
    rnd = random.Random(seed)
    groups = {}
    for r in rows:
        groups.setdefault((r[I["대학"]], _type(r[I["전형"]])), []).append(r)
    keys = sorted(groups)
    rnd.shuffle(keys)
    picked = []
    for k in keys:
        if len(picked) >= n:
            break
        picked.append(rnd.choice(groups[k]))

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "정제검수"
    ws.append(["항목", "[원본] 어디가 크롤값", "[정제] 결과", "판정(O/X)", "메모"])
    for c in ws[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CEN; c.border = BORDER

    row = 2
    for i, r in enumerate(picked, 1):
        title = (f"[{i:02d}] {r[I['대학']]} | {r[I['전형']]} | {r[I['모집']]}"
                 f"   (검수: 원본 보고 정제값 맞는지 O/X)")
        ws.cell(row, 1, title).font = UNIT_FONT
        for c in range(1, 6):
            ws.cell(row, c).fill = UNIT_FILL; ws.cell(row, c).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 20
        row += 1
        for label, raw_k, ref_k in ITEMS:
            ws.cell(row, 1, label).alignment = LEFT
            ws.cell(row, 2, r[I[raw_k]]).fill = RAW_FILL
            ws.cell(row, 3, r[I[ref_k]]).fill = REF_FILL
            for c in range(1, 6):
                cell = ws.cell(row, c)
                cell.alignment = LEFT; cell.border = BORDER
                if cell.font.size != 9:
                    cell.font = Font(size=9)
            row += 1
        row += 1                                  # 유닛 사이 빈 행

    for j, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    path = OUT_DIR / "정제검수_샘플.xlsx"
    out.save(path)
    return path, len(picked)


def main():
    n = next((int(a.split("=")[1]) for a in sys.argv[1:] if a.startswith("--n=")), 30)
    seed = next((int(a.split("=")[1]) for a in sys.argv[1:] if a.startswith("--seed=")), 11)
    path, cnt = build(n, seed)
    print(f"저장: {path}  | 표본 {cnt}유닛 × 6항목 = {cnt*6}셀 검수")


if __name__ == "__main__":
    main()

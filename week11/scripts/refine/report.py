"""정제 검토필요 + 골든셋 정합성 검증 통합 리포트.

한 파일 3시트:
  ① 검토필요       — 사람 확인이 필요한 건(정제 플래그 + 골든셋 대조로 드러난 어디가 원본 결손)
  ② 정합성_불일치   — 원본이 정상인데 골든셋과 값이 다른 건 (원본 결손 건은 ①로만 관리)
  ③ 특이사항       — 골든셋 오류 의심, 모집단위 개편 등 판단 메모

대학·전형·모집단위 표기는 크롤 원문이 기준. 골든셋에만 있는 모집단위(개편·통합모집)는 다루지 않는다.

실행: python report.py [<정제.xlsx>] [<골든셋.xlsx>] [--out=<경로>]
"""
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from review import COL, POLICY, _flagged            # noqa: E402

DEF_IN = Path("/Users/channy/Downloads/통합본/260825/전형정보_통합_정제.xlsx")
DEF_GOLD = Path("/Users/channy/Downloads/어디가 골든셋/2027_최저관련_정제_0715_일단최종5시.xlsx")
DEF_OUT = Path("/Users/channy/Downloads/통합본/260825/정제_검토필요_정합성검증.xlsx")

POLICY["최저-원본결손(전체)"] = (
    "어디가 상세페이지 최저학력기준란이 '대학에서 입력된 정보가 없습니다' — 대학이 아예 미입력",
    "어디가 재크롤로는 해결 불가(2026-08-25 재크롤 확인). 모집요강/대전모 등 외부 소스 보완 필요")

NOTES = [
    ("골든셋 오류 의심", "경북대학교[본교] 약학과",
     "학생부교과(지역인재 기초생활수급자등대상자전형) 외 1건",
     "원문: '국어, 영어, 사회/과학탐구 영역 중 상위 2개 등급과 수학 영역 등급 합이 6이내'",
     "골든 3합5 / 우리 3합6",
     "원문 숫자는 6. 동일 문구를 쓰는 경북대 다른 전형(의예·수의예·치의예 등)은 골든도 전부 '3합6'으로 "
     "우리와 일치함. 이 건만 5로 들어가 있어 골든셋 입력 오류로 판단. 우리 값(3합6) 유지."),
    ("골든셋 환산 누락 의심", "성균관대학교[본교] 스포츠과학과", "실기/실적(예체능 특기자)",
     "원문 ④: '국어, 수학, 영어, 탐구, 탐구, 한국사 중 상위등급 2개 과목의 평균등급이 7 이내'",
     "골든 2합7 / 우리 2합7 (값 일치)",
     "값은 일치하나 골든 내부 규칙이 모순. 골든은 평균등급형을 다른 곳에서는 N×M으로 환산함"
     "(호남대 평균 5등급→2합10, 한국성서대 평균 3등급→2합6, 수원가톨릭 평균 5등급→3합15). "
     "이 건만 환산 없이 2합7. 일관 적용하면 2합14. 데이터랩스 확인 필요."),
    ("모집단위 개편(시점 차) — 대조 불가", "부경대 30 · 동국WISE 28 · 조선대 27 · 성균관 10 등",
     "골든에만 111건 / 우리에만 73건",
     "골든셋 스냅샷 2026-07-15 ↔ 우리 크롤 2026-07-13·08-25. 같은 대학 안에서 모집단위명이 바뀜",
     "미조인 111건 중 퍼지매칭으로 값까지 확인된 건 12건뿐(전부 값 동일)",
     "3가지 유형: (1)표기 차이 — '건축학과<5년>'↔'건축학과', '의예과(경주권)'↔'의예과' "
     "(2)학과 개편 — '영어영문학부'↔'영어문화・산업전공', '생태공학전공'↔'생태공학과', "
     "'컴퓨터·인공지능공학부'↔'컴퓨터공학전공' (3)통합모집 단위 — "
     "'사회계열(법학과,중국학과,정치외교학과)'는 우리 쪽에 대응 단위 없음. "
     "양방향 건수(111 vs 73)가 같은 대학에 몰려 있어 수집 누락이 아니라 개편으로 판단"
     "(부경대 '학생부교과(일반)' 기준 골든 60 / 우리 55 / 공통 49). "
     "대학·전형·모집단위 표기는 크롤 원문을 기준으로 하므로 별도 시트로 관리하지 않음."),
    ("계열 추정 적용", "연세대학교[본교] 7개 모집단위 × 2전형 = 14건", "추천형 · 활동우수형",
     "원문에 인문계열(2합4+영3=3합7) / 자연계열(2합5+영3=3합8) 두 기준이 병기, 모집단위 계열 정보는 크롤 데이터에 없음",
     "골든과 14건 전원 일치",
     "학과명 키워드 규칙으로 계열 추정(간호·식품영양·실내건축·응용통계=자연 / 아동가족·의류환경·통합디자인=인문). "
     "②열에 [계열추정:자연|인문] 표기로 추적 가능. 규칙 순서 주의('의류환경'이 '환경'에 걸려 자연으로 "
     "오분류되지 않도록 인문 키워드를 앞에 둠). 다른 대학에서 같은 패턴 발생 시 스팟체크 권장."),
]

HEAD = PatternFill("solid", fgColor="1F4E78")
HEAD2 = PatternFill("solid", fgColor="C55A11")
HFONT = Font(bold=True, color="FFFFFF", size=9)
LEFT = Alignment(vertical="top", wrap_text=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _loose(s):
    """모집단위명 느슨 정규화 — 괄호·꺾쇠·연한 표기 제거 + 학과/학부/전공 접미 제거."""
    s = nk(s)
    s = re.sub(r"[<（(\[][^>）)\]]*[>）)\]]", "", s)
    s = re.sub(r"\d+년", "", s)
    return re.sub(r"(학과|학부|전공|과)$", "", s)


def nk(s):
    s = "" if s is None else str(s)
    return re.sub(r"\s+", "", re.sub(r"\((주간|야간)\)", "", s))


def nv(s):
    return re.sub(r"\s+", "", "" if s is None else str(s))


def _style(sh, widths, head_fill=HEAD, heights=None):
    for c in sh[1]:
        c.fill = head_fill; c.font = HFONT; c.alignment = CEN
    for r in sh.iter_rows(min_row=2):
        for c in r:
            c.alignment = LEFT; c.font = Font(size=9)
    for j, w in enumerate(widths, 1):
        sh.column_dimensions[get_column_letter(j)].width = w
    if heights:
        for i in range(2, sh.max_row + 1):
            sh.row_dimensions[i].height = heights
    sh.freeze_panes = "A2"
    if sh.max_row > 1:
        sh.auto_filter.ref = f"A1:{get_column_letter(sh.max_column)}{sh.max_row}"


def build(in_path=DEF_IN, gold_path=DEF_GOLD, out_path=DEF_OUT):
    # ── 골든셋 ──
    wb = openpyxl.load_workbook(gold_path, read_only=True)
    gold, gdup = {}, 0
    for r in wb["최저정제"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        k = (nk(r[0]), nk(r[1]), nk(r[4]))
        if k in gold:
            gdup += 1
            continue
        gold[k] = (nv(r[18]), r[16], r[17], r[15])
    wb.close()

    # ── 신규 정제본 ──
    wb = openpyxl.load_workbook(in_path, read_only=True)
    review, mine = [], {}
    for r in wb["정제"].iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            continue
        review.extend(_flagged(r))
        mine.setdefault((nk(r[2]), nk(r[6]), nk(r[4])), (nv(r[9]), r[10], r[8], r[2], r[4], r[6]))
    wb.close()

    both = set(gold) & set(mine)
    only_g = sorted(set(gold) - set(mine))
    honam = {k for k in both if k[0].startswith("호남대")}

    mismatch, ok, cause, missing = [], 0, Counter(), []
    for k in sorted(both):
        g1, g2, gadd, gdet = gold[k]
        m1, m2, mdet, U, T, M = mine[k]
        if k in honam and "검증필요" in m1:
            continue
        if g1 == m1 or (not g1 and m1 == "최저없음"):
            ok += 1
            continue
        if not (mdet or "").strip():
            missing.append((U, T, M, g1, g2, gadd))   # 원본 결손 → ①검토필요로만 관리
            continue
        if "검증필요" in m1:
            c = "B.우리 파싱 실패"
        elif m1 == "최저없음":
            c = "C.우리만 최저없음"
        elif not g1:
            c = "D.골든 값 공란"
        else:
            gm = re.match(r"^(\d+)합(\d+)$", g1); mm = re.match(r"^(\d+)합(\d+)$", m1)
            if gm and mm:
                c = ("H.N·M 모두 불일치" if gm.group(1) != mm.group(1) and gm.group(2) != mm.group(2)
                     else "E.영역수(N) 불일치" if gm.group(1) != mm.group(1) else "F.합(M) 불일치")
            else:
                c = "G.기타"
        cause[c] += 1
        mismatch.append((c, U, T, M, g1, m1, g2, m2, gadd, (mdet or "")[:200]))

    # ── ① 검토필요 (정제 플래그 + 원본 결손) ──
    for U, T, M, g1, g2, gadd in missing:
        사유, 해결책 = POLICY["최저-원본결손(전체)"]
        review.append(("최저-원본결손(전체)", U, T, M, 사유, "(공란)", "최저 없음",
                       f"{해결책} / 골든셋 참고값: {g1}"))

    out = openpyxl.Workbook()
    sh = out.active; sh.title = "검토필요"
    sh.append(["항목", "대학명", "전형명", "모집단위명", "사유(왜 검토필요)", "[원본]",
               "[현재 정제]", "해결책(권장)", "검토결과(직접입력)"])
    for row in review:
        sh.append(list(row) + [""])
    _style(sh, [22, 18, 26, 20, 40, 40, 34, 46, 20])

    # ── ② 정합성_불일치 ──
    s2 = out.create_sheet("정합성_불일치")
    s2.append(["원인", "대학명", "전형명", "모집단위명", "골든①", "신규①",
               "골든 선택필수구성", "신규②", "골든 추가", "[원본] 세부내용"])
    for row in mismatch:
        s2.append(list(row))
    _style(s2, [26, 20, 26, 22, 10, 10, 20, 34, 34, 60])

    # ③ 골든셋만/우리에만 있는 모집단위 시트는 두지 않음 —
    #   대학·전형·모집단위 표기는 크롤 원문이 기준이고, 골든셋(2026-07-15 스냅샷)과의
    #   모집단위 개편 차이는 정제 이슈가 아니므로 ④특이사항에 요약만 남긴다.

    # ── ③ 특이사항 ──
    s4 = out.create_sheet("특이사항")
    s4.append(["구분", "대상", "범위", "원문/근거", "값 비교", "판단·조치"])
    for n in NOTES:
        s4.append(list(n))
    _style(s4, [22, 34, 26, 52, 24, 72], head_fill=HEAD2, heights=90)

    out.save(out_path)
    tot = ok + len(mismatch)
    return out_path, dict(
        검토필요=len(review), 검토필요_항목별=dict(Counter(r[0] for r in review)),
        골든고유키=len(gold), 조인=len(both),
        모집단위개편_대조불가=len(only_g), 원본결손_평가제외=len(missing) + len(honam),
        평가대상=tot, 일치=ok, 일치율=round(ok / tot * 100, 2),
        불일치=len(mismatch), 불일치_원인=dict(cause))


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    out = next((Path(a.split("=", 1)[1]) for a in sys.argv[1:] if a.startswith("--out=")), DEF_OUT)
    p, stat = build(Path(args[0]) if args else DEF_IN,
                    Path(args[1]) if len(args) > 1 else DEF_GOLD, out)
    print("저장:", p)
    for k, v in stat.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()

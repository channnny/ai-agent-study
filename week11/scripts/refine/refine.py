"""전형정보 통합본 → 시트3 '정제' 추가 (순수 결정적 프로토타입).

정제 6종: ①N합N ②영역조합 ③교과반영영역 ④진로A/B/C 5a전형요소별 5b학년/요소별.
난케이스는 오답 대신 '검증필요:<사유>' 플래그. (설계: docs/superpowers/specs/)

프로토타입 단일 파일 — 검증 후 sources/parsers/build 모듈로 분리 예정.
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── RAW 컬럼 인덱스 (0-based, iter_rows values_only) ──────────────
# 시트1 전형일정및방법(46열)
S1_SEL_MODEL, S1_SEL_METHOD, S1_SEL_RATE = 14, 15, 16
S1_RB = {"학생부": 17, "수능": 18, "면접": 19, "논술": 20, "적성": 21,
         "1단계성적": 22, "실기": 23, "서류": 24, "기타": 25}
S1_CHOI_영역수, S1_CHOI_세부 = 44, 45
# 시트2 전형요소(55열)
S2_학년공통, S2_공통비율, S2_1학년, S2_2학년, S2_3학년 = 33, 34, 35, 36, 37
S2_요소 = {"교과": 38, "출결": 39, "자격": 40, "활동": 41, "봉사": 42, "기타": 43}
S2_서류학생부, S2_반영교과, S2_진로선택, S2_각주 = 44, 49, 52, 54

PLACEHOLDER = "대학에서 입력된 정보가 없습니다."


def _c(v):
    """셀 정규화: None/placeholder/공백 → ''."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == PLACEHOLDER else s


def _num(v):
    """'100', '90 ' → 정수 문자열. 아니면 ''."""
    s = _c(v)
    m = re.search(r"\d+", s)
    return m.group(0) if m else ""


# ── 종합전형 판별 ────────────────────────────────────────────────
def is_jonghap(전형명: str, s2) -> bool:
    if "종합" in 전형명:
        return True
    # 서류 평가요소(학교생활기록부) 채워짐 + 교과반영교과 없음 → 종합
    return bool(_c(s2[S2_서류학생부])) and not _c(s2[S2_반영교과])


# ── ③ 교과반영영역 ──────────────────────────────────────────────
_GYO_ORDER = [("국", ["국어"]), ("영", ["영어"]), ("수", ["수학"]),
              ("사", ["사회", "역사", "도덕"]), ("과", ["과학"]), ("한", ["한국사"])]


def refine_gyogwa(반영교과: str, jonghap: bool) -> str:
    if jonghap:
        return "해당없음(종합)"
    raw = _c(반영교과)
    if not raw:
        return "미반영"
    parts = [p.strip() for p in re.split(r"[/,]", raw) if p.strip()]
    out = []
    for code, keys in _GYO_ORDER:
        for p in parts:
            if any(k in p for k in keys):
                # '한국사'는 '한'으로만(사회의 역사와 구분)
                if code == "사" and "한국사" in p and "사회" not in p:
                    continue
                out.append(code)
                break
    return "".join(dict.fromkeys(out)) or "미반영"


# ── 5a 전형요소별비율 ───────────────────────────────────────────
def _decompose_hakbu(hakbu_val: str, s2, flags: list) -> list:
    """교과전형 학생부 → 요소별(교과/출결/…) 분해. 학생부=100이면 요소 그대로."""
    elems = [(lbl, _num(s2[idx])) for lbl, idx in S2_요소.items() if _num(s2[idx])]
    if not elems:
        return [f"교과{hakbu_val}"]
    if hakbu_val == "100":
        return [f"{lbl}{v}" for lbl, v in elems]
    # 학생부<100 + 요소분해 → 불확실, 교과로 collapse + 플래그
    if len(elems) > 1:
        flags.append("5a학생부분해")
    return [f"교과{hakbu_val}"]


def _ratio_components(cells: dict, s2, jonghap: bool, flags: list) -> str:
    """한 단계(또는 일괄)의 반영비율 요소 문자열. cells: {요소명: 값}."""
    comps = []
    for name, val in cells.items():
        v = _num(val)
        if not v:
            continue
        if name == "학생부":
            if jonghap:
                comps.append(f"서류{v}")
            else:
                comps += _decompose_hakbu(v, s2, flags)
        elif name == "1단계성적":
            comps.append(f"1단계{v}")
        else:
            comps.append(f"{name}{v}")
    return "+".join(comps)


def refine_5a(s1, s2, jonghap: bool, flags: list) -> str:
    model = _c(s1[S1_SEL_MODEL])
    if not model:
        return "해당없음(종합)" if jonghap else "미반영"
    is_stage = "단계" in model
    if not is_stage:
        cells = {k: s1[i] for k, i in S1_RB.items()}
        body = _ratio_components(cells, s2, jonghap, flags)
        return f"[일괄]{body}" if body else "검증필요:5a빈값"
    # 단계별: 각 반영비율 컬럼을 ' / '로 분할 → 단계별
    rates = _c(s1[S1_SEL_RATE]).split("/")            # 예 '400 / 100'
    n_stage = max(len([x for x in rates if x.strip()]), 2)
    stage_cells = {k: _c(s1[i]).split("/") for k, i in S1_RB.items()}
    out = []
    for st in range(n_stage):
        cells = {k: (vals[st] if st < len(vals) else "") for k, vals in stage_cells.items()}
        body = _ratio_components(cells, s2, jonghap, flags)
        if st == 0:
            mult = _num(rates[0]) if rates else ""
            baesu = str(int(int(mult) / 100)) if mult and int(mult) % 100 == 0 else ""
            head = f"[1단계({baesu}배수)]" if baesu else "[1단계]"
        else:
            head = f"[{st+1}단계]"
        out.append(head + (body or "검증필요"))
    if any("검증필요" in o for o in out):
        flags.append("5a단계별")
    return ";".join(out)


# ── 5b 학년별/요소별비율 ────────────────────────────────────────
def refine_5b(s2, jonghap: bool, flags: list) -> str:
    if jonghap:
        return "해당없음(종합)"
    공통 = _num(s2[S2_공통비율])
    if _c(s2[S2_학년공통]) or 공통:
        year = f"전학년공통{공통 or '100'}"
    else:
        ys = [(y, _num(s2[i])) for y, i in [("1학년", S2_1학년), ("2학년", S2_2학년), ("3학년", S2_3학년)] if _num(s2[i])]
        year = "·".join(f"{y}{v}" for y, v in ys)
    elems = [(lbl, _num(s2[idx])) for lbl, idx in S2_요소.items() if _num(s2[idx])]
    if not year and not elems:
        return "미반영"
    esum = sum(int(v) for _, v in elems)
    if elems and esum != 100:
        flags.append(f"5b합{esum}")
    elem_s = "+".join(f"{lbl}{v}" for lbl, v in elems)
    return f"[학년]{year} [요소]{elem_s}".strip()


# ── ①② 수능최저 ────────────────────────────────────────────────
# 최저 구조화 컬럼(0-based): 국선택35 수선택37 영39 탐선택40 탐과목41 제2외42 한국사43
S1_국, S1_수, S1_영, S1_탐, S1_탐과목, S1_한 = 35, 37, 39, 40, 41, 43
_TAM_과 = ["물리", "화학", "생명", "지구과학", "과학"]
_TAM_사 = ["지리", "역사", "윤리", "사회", "경제", "정치", "세계", "수산", "해운", "농업", "공업", "상업", "가사"]


def _tamgu_label(과목: str) -> str:
    사 = any(k in 과목 for k in _TAM_사)
    과 = any(k in 과목 for k in _TAM_과)
    if 과 and not 사:
        return "과"
    if 사 and not 과:
        return "사"
    return "탐"


def _build_yeongyeok(s1, N: str, K: str) -> str:
    """구조화 컬럼 → '국,수,영,탐(1) 중 N개 등급합 K' + 필수 포함 표기."""
    areas, 필수 = [], []
    for label, idx in [("국", S1_국), ("수", S1_수), ("영", S1_영)]:
        v = _c(s1[idx])
        if v:
            areas.append(label)
            if "필수" in v:
                필수.append(label)
    탐과목 = _c(s1[S1_탐과목])
    if _c(s1[S1_탐]) or 탐과목:
        areas.append(f"{_tamgu_label(탐과목)}(1)")
        if "필수" in _c(s1[S1_탐]):
            필수.append("탐")
    body = ",".join(areas) + f" 중 {N}개 등급합 {K}"
    # 필수 표기는 '선택 중 일부 필수'일 때만 의미(전 영역 필수/무필수면 생략)
    if 필수 and len(필수) < len(areas):
        body += f" ({','.join(필수)} 포함)"
    han = _c(s1[S1_한])
    if han and "필수" not in han:            # 등급컷 표기 → 한N (필수반영=응시필수는 제외)
        hm = re.search(r"\d+", han)
        if hm:
            body += f" 한{hm.group(0)}"
    return body


def refine_choi(s1, flags: list) -> tuple:
    세부 = _c(s1[S1_CHOI_세부])
    영역수 = _num(s1[S1_CHOI_영역수])
    if not 세부 and not 영역수:
        return "최저 없음", "최저 없음"
    txt = 세부.replace("!", "")
    # 계열별(인문/자연/예체능) 또는 다중 조건(1. 2. …) → 본질적 검증필요
    복잡 = bool(re.search(r"(인문|자연|예체능|국제)\s*계열", txt)) or \
        len(re.findall(r"(?:^|[\s\n])\d+\.\s", txt)) >= 2
    if 복잡:
        flags.append("최저계열별")
        tag = f"검증필요:계열별 [{txt[:30]}]"
        return tag, tag
    # 단순 단일조건: N=영역수, K=합/등급 숫자 추출
    mK = (re.search(r"등급\s*합\s*(\d+)", txt) or re.search(r"합\s*(?:이|의)?\s*(\d+)", txt)
          or re.search(r"(\d+)\s*등급\s*(?:이내|이하)", txt))
    if 영역수 and mK:
        return f"{영역수}합{mK.group(1)}", _build_yeongyeok(s1, 영역수, mK.group(1))
    flags.append("최저파싱")
    tag = f"검증필요:최저 [{txt[:30]}]"
    return tag, tag


# ── ④ 진로 A/B/C ────────────────────────────────────────────────
def refine_jinro(s2, jonghap: bool, flags: list) -> str:
    if jonghap:
        return "해당없음(종합)"
    진로 = _c(s2[S2_진로선택]).replace("/", "").strip()
    if not 진로:
        return "미반영"
    각주 = _c(s2[S2_각주])
    m = re.search(r"A\s*[:(]?\s*(\d+)\s*등급.*?B\s*[:(]?\s*(\d+)\s*등급.*?C\s*[:(]?\s*(\d+)\s*등급", 각주, re.S)
    if m:
        return f"A={m.group(1)},B={m.group(2)},C={m.group(3)}등급"
    m = re.search(r"A\s*[:(]?\s*(\d+)\s*점.*?B\s*[:(]?\s*(\d+)\s*점.*?C\s*[:(]?\s*(\d+)\s*점", 각주, re.S)
    if m:
        return f"A={m.group(1)},B={m.group(2)},C={m.group(3)}점"
    flags.append("진로소스없음")
    return "내부확인"


# ── [원본] 직렬화 컬럼 ──────────────────────────────────────────
def raw_choi(s1):
    return _c(s1[S1_CHOI_세부])


def raw_jinro(s2):
    j, note = _c(s2[S2_진로선택]), _c(s2[S2_각주])
    return f"진로선택: {j} / 반영방법각주: {note}".strip()


def raw_ratio(s1, s2):
    rb = " ".join(f"{k}={_c(s1[i])}" for k, i in S1_RB.items() if _c(s1[i]))
    yr = f"학년공통={_c(s2[S2_학년공통])} 공통비율={_c(s2[S2_공통비율])}"
    el = " ".join(f"{k}={_c(s2[i])}" for k, i in S2_요소.items() if _c(s2[i]))
    return f"선발모형={_c(s1[S1_SEL_MODEL])} 선발비율={_c(s1[S1_SEL_RATE])} | 반영비율: {rb} || 학년/요소: {yr} {el}"


# ── 시트3 헤더 ──────────────────────────────────────────────────
IDENT = ["adiga_selcntnm", "학년도", "대학명", "대학코드", "전형명", "전형코드", "모집단위명", "모집단위코드"]
GROUPS = [("기본정보 (RAW A~H)", 8), ("정제①② 수능최저", 3), ("정제③ 교과반영영역", 2),
          ("정제④ 진로 A/B/C", 2), ("정제 5a/5b 반영비율", 3)]
COLNAMES = IDENT + [
    "[원본] 최저학력기준(세부내용)", "[정제①] N합N", "[정제②] 영역조합",
    "[원본] 반영교과", "[정제③] 교과반영영역",
    "[원본] 진로선택+반영방법각주", "[정제④] 진로 A/B/C",
    "[원본] 반영비율/학년·요소", "[정제5a] 전형요소별비율", "[정제5b] 학년별/요소별비율"]

# 참고 산출물(전형정보_통합_정제.xlsx > 정제) 스타일 그대로
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")   # row2 컬럼명
GRP_FILL = PatternFill("solid", fgColor="2E75B6")    # row1 그룹
GRP_FONT = Font(bold=True, color="FFFFFF", size=10)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
RAW_FILL = PatternFill("solid", fgColor="FFF2CC")    # [원본] 연노랑
REF_FILL = PatternFill("solid", fgColor="E2EFDA")    # [정제] 연초록
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(vertical="top", wrap_text=True)
DATA_FONT = Font(size=9)
COL_WIDTHS = [20, 6, 16, 9, 24, 12, 14, 11, 44, 14, 30, 34, 14, 46, 22, 50, 30, 30]
RAW_COLS = {9, 12, 14, 16}                           # [원본] 열(1-based)
REF_COLS = {10, 11, 13, 15, 17, 18}                  # [정제] 열


def refine_row(s1, s2):
    """RAW 두 행(values tuple) → (정제 18값, flags)."""
    ident = [s1[i] for i in range(8)]
    전형명 = _c(s1[4])
    jh = is_jonghap(전형명, s2)
    flags = []
    n합, 영역 = refine_choi(s1, flags)
    row = ident + [
        raw_choi(s1), n합, 영역,
        _c(s2[S2_반영교과]), refine_gyogwa(s2[S2_반영교과], jh),
        raw_jinro(s2), refine_jinro(s2, jh, flags),
        raw_ratio(s1, s2), refine_5a(s1, s2, jh, flags), refine_5b(s2, jh, flags)]
    return row, flags


def build(in_path: Path, out_path: Path = None, limit: int = None):
    wb = openpyxl.load_workbook(in_path)
    s1 = list(wb["전형일정및방법"].iter_rows(min_row=4, values_only=True))
    s2 = list(wb["전형요소"].iter_rows(min_row=4, values_only=True))
    # adiga_selcntnm(0)로 매칭
    m2 = {r[0]: r for r in s2 if r[0]}
    if "정제" in wb.sheetnames:
        del wb["정제"]
    ws = wb.create_sheet("정제")
    # 헤더 2행
    ci = 1
    for label, n in GROUPS:
        ws.cell(1, ci, label)
        if n > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + n - 1)
        for k in range(n):
            g = ws.cell(1, ci + k); g.fill = GRP_FILL; g.font = GRP_FONT; g.alignment = CEN
        ci += n
    for j, name in enumerate(COLNAMES, 1):
        c = ws.cell(2, j, name); c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CEN
    # 데이터 ([원본]=연노랑 / [정제]=연초록 / 식별=무색)
    flagcount = {}
    n = 0
    rows = s1 if limit is None else s1[:limit]
    for r1 in rows:
        if not r1[0]:
            continue
        r2 = m2.get(r1[0], [None] * 55)
        vals, flags = refine_row(r1, r2)
        for f in flags:
            flagcount[f] = flagcount.get(f, 0) + 1
        for j, v in enumerate(vals, 1):
            cell = ws.cell(3 + n, j, v)
            cell.alignment = LEFT; cell.font = DATA_FONT
            if j in RAW_COLS:
                cell.fill = RAW_FILL
            elif j in REF_COLS:
                cell.fill = REF_FILL
        n += 1
    # 열너비 · 행높이 · 틀고정(식별 8열 + 헤더 2행)
    for j, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 34
    ws.freeze_panes = "I3"
    if out_path is None:
        out_dir = Path(__file__).resolve().parents[2] / "output"   # week11/output
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / (in_path.stem + "_정제.xlsx")
    wb.save(out_path)
    return out_path, n, flagcount


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    in_path = Path(args[0]) if args else Path(
        "/Users/channy/Documents/workspaces/ai/ai-agent-study/week10/output/test_피드백/전형정보_통합.xlsx")
    limit = None
    for a in sys.argv[1:]:
        if a.startswith("--limit="):
            limit = int(a.split("=")[1])
    out, n, fc = build(in_path, limit=limit)
    print(f"저장: {out}")
    print(f"정제 {n}행")
    print("검증필요/플래그:", {k: v for k, v in sorted(fc.items(), key=lambda x: -x[1])})

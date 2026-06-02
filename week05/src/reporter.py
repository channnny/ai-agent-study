"""6시트 평가 리포트 (Excel) 작성.

시트:
  1) summary           — row=metric × col=person
  2) by_university     — row=대학 × col=person별 metric
  3) by_column         — row=DATA 컬럼 × col=person별 match_rate
  4) missing_rows      — golden에 있고 person에 없음 (long-format)
  5) extra_rows        — person에만 있음
  6) mismatched_cells  — matched 행의 셀 불일치 (long-format)

# adapted from leejihyun/evaluation_report.xlsx 양식
"""
from __future__ import annotations
import pandas as pd
from pathlib import Path
from datetime import datetime

from .config import DATA_COLUMNS, PERSON_KOR
from .matcher import PersonResult


SUMMARY_METRICS_ORDER = [
    ("pk_match_rate",   "PK 매칭률"),
    ("cell_match_rate", "셀 일치율"),
    ("n_matched",       "matched 행 수"),
    ("n_missing",       "missing 행 수"),
    ("n_extra",         "extra 행 수"),
    ("n_golden_total",  "골든 전체 행 수"),
    ("coverage_pct",    "커버리지(%)"),
    ("n_failed_univ",   "fail 대학 수"),
    ("n_missing_univ",  "missing 대학 수"),
    ("pk_dod_pass",     "PK DoD 통과 (≥85%)"),
    ("cell_dod_pass",   "셀 DoD 통과 (≥90%)"),
]


def _format_rate(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return "✓" if v else "✗"
    if isinstance(v, float) and 0 <= v <= 1:
        return f"{v*100:.2f}%"
    if isinstance(v, float):
        return round(v, 2)
    return v


def _bar(pct, width=10):
    """유니코드 막대. pct는 0~100."""
    if pct is None:
        return ""
    f = max(0, min(width, int(round(pct / 100 * width))))
    return "█" * f + "░" * (width - f)


def _light(pct, good, mid):
    """신호등: good 이상 🟢 / mid 이상 🟡 / 미만 🔴 / None ⬜."""
    if pct is None:
        return "⬜"
    return "🟢" if pct >= good else ("🟡" if pct >= mid else "🔴")


def _dash_cell(pct, good, mid):
    """대시보드 셀: '🟢 89.6% ████████░░'."""
    if pct is None:
        return "—"
    return f"{_light(pct, good, mid)} {pct:.1f}%  {_bar(pct)}"


def build_summary(results: list[PersonResult]) -> pd.DataFrame:
    """대시보드형 종합 시트.

    상단 = 핵심 3지표(커버리지·PK·셀)를 신호등+막대+기준선으로.
    하단 = 상세 행 수 지표.
    """
    def col(r):
        return PERSON_KOR.get(r.person, r.person)

    rows = []

    # ── 핵심 3지표 (신호등 + 막대) ──
    rows.append({"지표": "📊 커버리지 (몇 대학 시도)", **{
        col(r): _dash_cell(r.summary.get("coverage_pct"), 80, 40) for r in results}, "목표": "높을수록"})
    rows.append({"지표": "📊 PK 매칭률 (같은 행 찾음)", **{
        col(r): _dash_cell((r.summary.get("pk_match_rate") or 0) * 100, 85, 40) for r in results}, "목표": "≥ 85%"})
    rows.append({"지표": "📊 셀 일치율 (긁은 값이 정확)", **{
        col(r): _dash_cell((r.summary.get("cell_match_rate") or 0) * 100, 90, 50) for r in results}, "목표": "≥ 90%"})
    rows.append({"지표": "📊 셀 충진율 (얼마나 채웠나)", **{
        col(r): _dash_cell((r.summary.get("cell_fill_rate") or 0) * 100, 80, 50) for r in results}, "목표": "높을수록"})

    # ── 구분 ──
    rows.append({"지표": "", **{col(r): "" for r in results}, "목표": ""})
    rows.append({"지표": "─ 상세 (행 수) ─", **{col(r): "" for r in results}, "목표": ""})

    detail = [
        ("matched 행 수 (정확히 찾음)", "n_matched"),
        ("missing 행 수 (못 찾음→누락행)", "n_missing"),
        ("extra 행 수 (잘못 긁음→잉여행)", "n_extra"),
        ("골든 전체 행 수", "n_golden_total"),
        ("비교한 셀 수 (양쪽 값 있음)", "n_cell_compared"),
        ("미수집 셀 수 (한쪽만 값)", "n_cell_fill_gap"),
        ("미수집 대학 수 (→미수집대학)", "n_missing_univ"),
    ]
    for label, key in detail:
        row = {"지표": label}
        for r in results:
            v = r.summary.get(key)
            row[col(r)] = f"{v:,}" if isinstance(v, (int, float)) else v
        row["목표"] = ""
        rows.append(row)

    # ── 종합 판정 ──
    rows.append({"지표": "", **{col(r): "" for r in results}, "목표": ""})
    judge = {"지표": "🏁 종합 판정 (DoD)"}
    for r in results:
        pk_ok = r.summary.get("pk_dod_pass")
        cell_ok = r.summary.get("cell_dod_pass")
        judge[col(r)] = "✅ 통과" if (pk_ok and cell_ok) else "❌ 미달"
    judge["목표"] = "PK85·셀90"
    rows.append(judge)

    return pd.DataFrame(rows, columns=["지표"] + [col(r) for r in results] + ["목표"])


def _univ_cell(m: dict) -> str:
    """대학별 시트 한 셀: 신호등 + PK·셀 요약, 또는 미수집."""
    if not m:
        return "⬜ 미수집"
    st = m.get("status")
    if st == "missing":
        return "⬜ 미수집"
    pk = m.get("pk_match_rate")
    cell = m.get("cell_match_rate")
    if pk is None and m.get("n_golden", 0) == 0:
        return "▫ 골든에 없음"
    light = _light((pk or 0) * 100, 85, 40)
    pk_s = f"{pk*100:.0f}%" if pk is not None else "0%"
    cell_s = f"{cell*100:.0f}%" if cell is not None else "—"
    matched = f"{m.get('n_matched', 0)}/{m.get('n_golden', 0)}"
    return f"{light} PK {pk_s} · 셀 {cell_s}  ({matched})"


def build_by_university(results: list[PersonResult]) -> pd.DataFrame:
    """row=대학(unvCd), col=사람별 1컬럼(신호등+PK·셀 요약).

    정렬: 골든 행수 많은 대학 먼저(데이터 풍부) → 전부 미수집 대학은 아래로.
    """
    all_univ = set()
    univ_names: dict[str, str] = {}
    golden_rows: dict[str, int] = {}
    for r in results:
        for unv_cd, m in r.by_university.items():
            all_univ.add(unv_cd)
            if m.get("univ_name"):
                univ_names[unv_cd] = m["univ_name"]
            golden_rows[unv_cd] = max(golden_rows.get(unv_cd, 0), m.get("n_golden", 0) or 0)

    # 정렬 키: 데이터 있는 대학(누구라도 missing 아님) 먼저, 그 안에서 골든행수 desc
    def sort_key(u):
        any_data = any(
            r.by_university.get(u, {}).get("status") not in (None, "missing")
            for r in results
        )
        return (0 if any_data else 1, -golden_rows.get(u, 0), univ_names.get(u, ""))

    rows = []
    for unv_cd in sorted(all_univ, key=sort_key):
        row = {"unvCd": unv_cd, "대학": univ_names.get(unv_cd, "")}
        for r in results:
            label = PERSON_KOR.get(r.person, r.person)
            row[label] = _univ_cell(r.by_university.get(unv_cd, {}))
        rows.append(row)
    return pd.DataFrame(rows)


def build_by_column(results: list[PersonResult]) -> pd.DataFrame:
    rows = []
    for col in DATA_COLUMNS:
        # 모든 사람이 비교 0건(전원 None)인 컬럼은 숨김 — 예: 대학별환산_총점
        # (골든에 없어 비교 불가). 노이즈 제거.
        if all(r.by_column.get(col) is None for r in results):
            continue
        row = {"컬럼": col}
        for r in results:
            label = PERSON_KOR.get(r.person, r.person)
            row[label] = _format_rate(r.by_column.get(col))
        rows.append(row)
    return pd.DataFrame(rows)


def build_long_rows(results: list[PersonResult], attr: str) -> pd.DataFrame:
    """missing_rows / extra_rows / mismatched_cells 통합."""
    rows = []
    for r in results:
        rows.extend(getattr(r, attr))
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def build_diff_rows(results: list[PersonResult]) -> pd.DataFrame:
    """누락행 + 잉여행을 한 시트로 통합.

    같은 (사람·대학·모집단위)로 정렬 → '전형명만 다른' 누락↔잉여 쌍이 인접해
    PK 매칭 실패 원인(전형명 분류 차이)이 한눈에 드러난다.
    """
    rows = []
    for r in results:
        person = PERSON_KOR.get(r.person, r.person)
        for m in r.missing_rows:
            rows.append({
                "구분": "🔴 누락 (골든O·크롤러X)",
                "사람": person, "unvCd": m.get("unvCd"), "대학": m.get("대학"),
                "전형": m.get("전형"), "모집단위": m.get("모집단위"),
            })
        for e in r.extra_rows:
            rows.append({
                "구분": "🔵 잉여 (크롤러O·골든X)",
                "사람": person, "unvCd": e.get("unvCd"), "대학": e.get("대학"),
                "전형": e.get("전형"), "모집단위": e.get("모집단위"),
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=["사람", "대학", "모집단위", "전형", "구분", "unvCd"])
    # 같은 사람·대학·모집단위 안에서 누락/잉여가 붙도록 정렬
    df = df.sort_values(["사람", "대학", "모집단위", "전형"],
                        key=lambda s: s.fillna("")).reset_index(drop=True)
    return df


def _verdict(bigo: str) -> str:
    """불일치 '비고' 텍스트 → 직관적 판정 이모지."""
    b = str(bigo or "")
    if "콤마" in b:
        return "🟡 거의같음 (콤마 포맷)"
    if "근접" in b:
        return "🟡 거의같음 (반올림)"
    if "null" in b:
        return "⬜ 한쪽만 값있음"
    return "🔴 값 다름"


_VERDICT_ORDER = {"🔴": 0, "⬜": 1, "🟡": 2}


def build_mismatch(results: list[PersonResult]) -> pd.DataFrame:
    """불일치셀 — 판정 이모지 + 골든↔크롤러 값 인접 + 심각한 것 위로 정렬."""
    rows = []
    for r in results:
        person = PERSON_KOR.get(r.person, r.person)
        for c in r.mismatched_cells:
            gv = c.get("golden_value")
            pv = c.get("person_value")
            # 한쪽만 값이 있으면 무조건 ⬜ (비고 텍스트보다 우선)
            if (gv is None) != (pv is None):
                verdict = "⬜ 한쪽만 값있음"
            else:
                verdict = _verdict(c.get("비고"))
            rows.append({
                "판정": verdict,
                "사람": person, "대학": c.get("대학"),
                "전형": c.get("전형"), "모집단위": c.get("모집단위"),
                "항목": c.get("컬럼"),
                "골든값": "(빈칸)" if gv is None else gv,
                "크롤러값": "(빈칸)" if pv is None else pv,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=["판정", "사람", "대학", "전형", "모집단위", "항목", "골든값", "크롤러값"])
    df = df.sort_values(
        ["사람", "판정", "대학", "항목"],
        key=lambda s: s.map(lambda v: _VERDICT_ORDER.get(str(v)[:1], 9)) if s.name == "판정" else s.fillna(""),
    ).reset_index(drop=True)
    return df


def build_uncovered(results: list[PersonResult]) -> pd.DataFrame:
    """커버리지 100%가 안 되는 이유 — 데이터를 못 낸 골든 대학 명단.

    유찬·이지현 둘 다 못 낸 대학 = 어디가 2025 미게시 추정(소스 문제).
    한쪽만 못 낸 대학 = 해당 크롤러 미수집.
    """
    by_person = {r.person: r for r in results}
    yuchan = by_person.get("yuchan")
    lee = by_person.get("lee")

    # 모든 unvCd + 대학명 수집
    names: dict[str, str] = {}
    for r in results:
        for u, m in r.by_university.items():
            if m.get("univ_name"):
                names[u] = m["univ_name"]
            names.setdefault(u, "")

    rows = []
    for u in sorted(names):
        # 골든에 실재하는 대학만 (어느 사람 기준이든 골든 행 수 > 0)
        ng = max((r.by_university.get(u, {}).get("n_golden", 0) or 0) for r in results)
        if ng == 0:
            continue
        y_st = yuchan.by_university.get(u, {}).get("status") if yuchan else None
        l_st = lee.by_university.get(u, {}).get("status") if lee else None
        y_miss = (y_st == "missing")
        l_miss = (l_st == "missing")
        if not (y_miss or l_miss):
            continue  # 둘 다 데이터 냄 → 제외

        if y_miss and l_miss:
            cause = "① 어디가 2025 미게시 추정 (유찬·이지현 공통 누락)"
        elif y_miss:
            cause = "② 유찬만 미수집 (복잡 테이블 등)"
        else:
            cause = "③ 이지현만 미수집"
        rows.append({
            "unvCd": u,
            "대학명": names[u],
            "골든 행수": ng,
            "유찬": y_st or "-",
            "이지현": l_st or "-",
            "추정 원인": cause,
        })
    # 공통 누락(①) 먼저 보이도록 정렬
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["추정 원인", "대학명"]).reset_index(drop=True)
    return df


def build_glossary() -> pd.DataFrame:
    """리포트 맨 앞 — 용어·지표·읽는 법 설명. 회의에서 처음 보는 사람용."""
    rows = [
        ("■ 이 리포트는?", "3명(유찬·이지현·임지현)의 어디가 크롤러 출력을 작년 골든셋(2025_수시_입시결과_통합본)과 비교해 정확도를 측정한 결과입니다."),
        ("", ""),
        ("■ 신호등·막대 읽는 법", "각 표에 공통으로 쓰입니다."),
        ("🟢 / 🟡 / 🔴 / ⬜", "🟢=목표 달성 / 🟡=절반 이상(아쉬움) / 🔴=목표 미달 / ⬜=데이터 없음(미수집)."),
        ("████░░░░░░", "비율 막대. 채워진 칸이 많을수록 높음 (한 칸 = 10%)."),
        ("", ""),
        ("■ 핵심 깔때기 3단계 (①→②→③)", "셋 다 높아야 진짜 정확. 하나만 높으면 함정."),
        ("① 커버리지", "골든 173개 대학 중 '데이터를 낸' 대학 비율. = 애초에 평가 테이블에 올라올 자격. (신호등: 80%↑🟢 / 40%↑🟡)"),
        ("② PK 매칭률", "PK=(대학코드,전형,모집단위) 3개 조합. '둘 다 가진 PK 수 ÷ 골든 전체 PK 수'. = 빠뜨리지 않고 같은 행을 찾았는가?(양). 목표 ≥85%."),
        ("③ 셀 일치율", "PK 매칭된 행에서 '양쪽 다 값이 있는 셀'만 비교해 일치한 비율. = 긁어온 값이 정확한가?(질). 한쪽만 값있는 셀(크롤러 미수집)은 분모에서 제외 — 그건 정확도가 아니라 충진율 문제. 목표 ≥90%."),
        ("■ 보조 지표", ""),
        ("④ 셀 충진율", "비교 대상 셀 중 '양쪽 다 값이 있던' 비율. 낮으면 = 골든엔 있는 항목을 크롤러가 덜 긁음(예: 어디가가 50%컷 미공개). 일치율과 분리해서 봐야 공정."),
        ("🏁 종합 판정 (DoD)", "PK ≥85% '그리고' 셀 일치율 ≥90% 면 ✅ 통과, 아니면 ❌ 미달."),
        ("", ""),
        ("■ 커버리지가 100%가 안 되는 이유", "→ 대학별 상세 명단은 '미수집대학' 시트."),
        ("(1) 어디가 미게시", "강릉원주대·가톨릭대(성의/성신교정)·대구예술대 등은 2025 전형 결과가 어디가에 아직 안 올라옴 → 유찬·이지현 모두 공통으로 못 냄 (크롤러 문제 아님, 소스 문제)."),
        ("(2) 유찬 복잡 테이블 스킵", "경북대·안양대·중부대·춘천교대 등은 '단과대학+모집단위' 다단 헤더라 유찬 어댑터가 아직 파싱 못 함 (W06 보강 예정)."),
        ("", ""),
        ("■ 남은 PK 격차(~43%)의 주원인", "크롤링 실패가 아니라 전형명 분류 차이. 예: 골든 '해람인재' vs 크롤러 '학생부종합(해람인재)' — 같은 전형인데 이름이 달라 한쪽은 누락·한쪽은 잉여로 잡힘. '누락·잉여' 시트에서 인접 쌍으로 확인 가능. W06에서 전형 분류 사전 합의로 해소."),
        ("", ""),
        ("■ 탭(시트)별 보는 법", ""),
        ("종합 (대시보드)", "3명을 신호등+막대로 한눈에. 위 3줄(📊)이 핵심 지표, 아래는 상세 행 수, 맨 끝 🏁이 최종 판정."),
        ("대학별", "대학마다 사람별 1칸 = '신호등 PK% · 셀% (맞은행/골든행)'. 배경색=신호등. 데이터 많은 대학이 위, 미수집은 아래. 같은 대학명 2줄=캠퍼스 분리(unvCd 다름)."),
        ("항목별", "데이터 항목별 일치율(매칭+양쪽 값 있는 셀 기준). 낮은 항목 = 그 항목 파싱 우선 개선."),
        ("  └ 반영교과 0%", "자동비교 한계 항목. 골든은 상세 기입('국수영과,전교과,동일비율,진로,미반영'), 어디가/크롤러는 요약('전교과')이라 정보량이 달라 전부 불일치. 사람 검수 또는 W06 파싱 개선 대상 — 다른 항목 정확도와 분리해서 볼 것."),
        ("미수집대학", "커버리지 미달 대학 명단 + 추정 원인(①소스 미게시 / ②③크롤러 미수집)."),
        ("누락·잉여", "🔴누락(골든O·크롤러X) + 🔵잉여(크롤러O·골든X)를 한 시트에. 같은 대학·모집단위로 정렬돼 전형명만 다른 쌍이 인접 → PK 격차 원인이 보임."),
        ("불일치셀", "PK·셀 양쪽 다 값이 있는데 값이 다른 칸만. 🔴값다름(진짜 오류, 위쪽) / 🟡거의같음(콤마·반올림=사실상 정답). 한쪽만 값있는 셀(미수집)은 여기 없고 충진율로 집계. 골든값↔크롤러값 나란히."),
    ]
    return pd.DataFrame(rows, columns=["항목", "설명"])


# ──────────────────────────────────────────────────────────────
# 스타일링 (openpyxl)
# ──────────────────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")   # 진파랑
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_DESC_FILL   = PatternFill("solid", fgColor="D9E1F2")    # 연파랑
_DESC_FONT   = Font(italic=True, color="1F3864", size=10)
_PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")    # 연초록
_FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")    # 연빨강
_MISS_FILL   = PatternFill("solid", fgColor="EDEDED")    # 회색
_SECTION_FONT = Font(bold=True, color="1F3864", size=11)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _disp_width(s: str) -> int:
    """한글 등 전각 문자는 2폭으로 계산."""
    return sum(2 if ord(c) > 0x1100 else 1 for c in str(s))


def _autofit(ws, df: pd.DataFrame, header_row: int, max_w: int = 60):
    for j, col in enumerate(df.columns, start=1):
        vals = [str(col)] + [str(v) for v in df[col] if v is not None]
        w = max((_disp_width(v) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), max_w)


def _write_styled(writer, df: pd.DataFrame, sheet: str, desc: str):
    """설명행(1행) + 헤더(2행) + 데이터 시트 작성 + 스타일."""
    if df.empty:
        df = pd.DataFrame([{"info": "해당 행 없음"}])
    df.to_excel(writer, sheet_name=sheet, index=False, startrow=1)
    ws = writer.sheets[sheet]
    n_cols = len(df.columns)

    # 1행: 설명 (병합)
    ws.cell(1, 1, "📌 " + desc)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1)
    c.fill = _DESC_FILL
    c.font = _DESC_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 42

    # 2행: 헤더
    for j in range(1, n_cols + 1):
        h = ws.cell(2, j)
        h.fill = _HEADER_FILL
        h.font = _HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = _BORDER

    # 데이터 영역 테두리 + 정렬
    for r in range(3, ws.max_row + 1):
        for j in range(1, n_cols + 1):
            cell = ws.cell(r, j)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _autofit(ws, df, header_row=2)
    ws.freeze_panes = "A3"  # 설명+헤더 고정


def _colorize_univ(ws, df: pd.DataFrame):
    """대학별 시트 사람 컬럼(신호등 셀)을 내용 따라 색칠."""
    person_cols = [j for j, c in enumerate(df.columns, start=1)
                   if str(c) not in ("unvCd", "대학")]
    for r in range(3, ws.max_row + 1):
        for j in person_cols:
            v = str(ws.cell(r, j).value or "")
            if "⬜" in v:
                ws.cell(r, j).fill = _MISS_FILL
            elif "🟢" in v:
                ws.cell(r, j).fill = _PASS_FILL
            elif "🔴" in v:
                ws.cell(r, j).fill = _FAIL_FILL
            elif "🟡" in v:
                ws.cell(r, j).fill = PatternFill("solid", fgColor="FFEB9C")  # 연노랑


def _colorize_diff(ws, df: pd.DataFrame):
    """누락·잉여 시트의 '구분' 컬럼 색칠."""
    if df.empty:
        return
    gj = [j for j, c in enumerate(df.columns, start=1) if c == "구분"]
    if not gj:
        return
    j = gj[0]
    for r in range(3, ws.max_row + 1):
        v = str(ws.cell(r, j).value or "")
        if "누락" in v:
            ws.cell(r, j).fill = _FAIL_FILL
        elif "잉여" in v:
            ws.cell(r, j).fill = PatternFill("solid", fgColor="DDEBF7")  # 연파랑


def _colorize_verdict(ws, df: pd.DataFrame):
    """불일치셀 시트의 '판정' 컬럼 색칠."""
    if df.empty:
        return
    pj = [j for j, c in enumerate(df.columns, start=1) if c == "판정"]
    if not pj:
        return
    j = pj[0]
    for r in range(3, ws.max_row + 1):
        v = str(ws.cell(r, j).value or "")
        if v.startswith("🔴"):
            ws.cell(r, j).fill = _FAIL_FILL
        elif v.startswith("🟡"):
            ws.cell(r, j).fill = PatternFill("solid", fgColor="FFEB9C")
        elif v.startswith("⬜"):
            ws.cell(r, j).fill = _MISS_FILL


def _highlight_dashboard(ws):
    """종합 시트의 핵심 3지표(📊) 행을 연노랑 배경으로 강조."""
    hl = PatternFill("solid", fgColor="FFF2CC")
    for r in range(3, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "")
        if label.startswith("📊") or label.startswith("🏁"):
            for j in range(1, ws.max_column + 1):
                if ws.cell(r, j).fill.fgColor.rgb in ("00000000", None):
                    ws.cell(r, j).fill = hl


def _style_glossary(ws, df: pd.DataFrame):
    """용어설명 시트 전용 스타일 — 섹션 헤더(■) 강조."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    # 헤더행
    for j in (1, 2):
        h = ws.cell(1, j)
        h.fill = _HEADER_FILL
        h.font = _HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1)
        b = ws.cell(r, 2)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if a.value and str(a.value).startswith("■"):
            a.font = _SECTION_FONT
            a.fill = _DESC_FILL
            b.fill = _DESC_FILL


def write_report(results: list[PersonResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    glossary   = build_glossary()
    summary    = build_summary(results)
    by_univ    = build_by_university(results)
    by_col     = build_by_column(results)
    uncovered  = build_uncovered(results)
    diff       = build_diff_rows(results)
    mismatched = build_mismatch(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 용어설명 (자체가 설명이라 별도 desc 없이)
        glossary.to_excel(writer, sheet_name="용어설명", index=False)
        _style_glossary(writer.sheets["용어설명"], glossary)

        _write_styled(writer, summary, "종합",
            "🟢=목표달성 🟡=절반이상 🔴=미달 ⬜=없음.  깔때기: 커버리지(173곳 중 몇 곳 시도)→PK매칭률(같은 행 찾음)→셀일치율(값 정확).  현재 유찬·이지현은 커버리지·셀 일치율은 목표 달성(90%↑)이나, PK는 전형명 분류 차이로 미달(누락·잉여 시트 참조). PK 85% 도달은 전형명 표준 합의 필요. 임지현은 8개 표본만.")
        _highlight_dashboard(writer.sheets["종합"])
        _write_styled(writer, by_univ, "대학별",
            "대학별 정확도 한 셀 요약: '신호등 PK% · 셀% (매칭행/골든행)'. "
            "예) '🔴 PK 37% · 셀 74% (226/619)' = 골든 619행 중 226행을 찾음(PK 37%), 그 중 값 일치 74%. "
            "신호등은 PK 기준: 🟢85%↑ 🟡40%↑ 🔴40%미만 ⬜미수집. 셀 '—'는 매칭행이 0이라 계산 불가. "
            "데이터 많은 대학이 위, 미수집은 아래. 같은 대학명 2줄=캠퍼스 분리(unvCd 다름).")
        _colorize_univ(writer.sheets["대학별"], by_univ)
        _write_styled(writer, by_col, "항목별",
            "8개 데이터 항목별 일치율(매칭된 행 한정). 숫자가 낮은 항목 = 그 항목 파싱이 부정확 → 우선 개선 대상. 예: '학생부등급_50컷'이 낮으면 등급 파싱 로직 점검.")
        _write_styled(writer, uncovered, "미수집대학",
            "커버리지가 100%가 안 되는 대학 명단. '골든 행수'=골든셋에 있는 그 대학 행 수(놓친 양). "
            "추정 원인 ①=어디가에 2025 결과 미게시(소스 문제, 누구도 못 함) / ②=유찬만 미수집(복잡 테이블 등) / ③=이지현만 미수집. "
            "'유찬'·'이지현' 칸의 missing=그 사람이 못 냄.")
        _write_styled(writer, diff, "누락·잉여",
            "🔴누락=골든엔 있는데 크롤러가 못 찾음 / 🔵잉여=크롤러엔 있는데 골든엔 없음. "
            "같은 사람·대학·모집단위로 정렬돼 전형명만 다른 누락↔잉여 쌍이 인접 → PK 격차의 주원인이 보임. "
            "예: 한 학과에 🔴누락 '해람인재' + 🔵잉여 '학생부종합(해람인재)'가 나란히 = 같은 전형인데 이름만 달라 매칭 실패. "
            "[데이터>필터]로 사람·대학 좁혀 보세요.")
        _colorize_diff(writer.sheets["누락·잉여"], diff)
        _write_styled(writer, mismatched, "불일치셀",
            "PK는 맞았는데 값이 다른 칸. 판정: 🔴값다름(진짜 오류) / 🟡거의같음(콤마·반올림, 사실상 정답) / ⬜한쪽만 값있음. 🔴부터 정렬 → 위쪽이 실제 고칠 것. 🟡이 많으면 실제 정확도는 셀일치율보다 높음.")
        _colorize_verdict(writer.sheets["불일치셀"], mismatched)

    print(f"\n✓ 평가 리포트: {output_path}")
    print(f"  시트 7개: 용어설명 · 종합 · 대학별 · 항목별 · 미수집대학 · 누락·잉여 · 불일치셀")

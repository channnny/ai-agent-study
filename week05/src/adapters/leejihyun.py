"""이지현 산출물(`per_university/{unvCd}.xlsx`) → 캐노니컬 DataFrame.

unvCd 매칭을 위해 통합 워크북(adiga_2027.xlsx) 대신 per_university 디렉토리를
순회한다. 파일명이 unvCd. 시트 구조는 통합 워크북과 동일.

구조: 4시트 워크북. 우리는 `수시 입시결과` 시트만 사용.
헤더: R2가 컬럼명. 데이터는 R3+.

# adapted from leejihyun/output/per_university/{unvCd}.xlsx 실측 + schema_v3.yaml
"""
from __future__ import annotations
import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict

from ..config import PK_COLUMNS, DATA_COLUMNS, CANONICAL_COLUMNS, GROUP_LABEL_COL
from ..normalizer import Normalizer


SHEET_NAME = "수시 입시결과"

# 이지현 R2 헤더 → 캐노니컬 컬럼명
# (없으면 컬럼명 그대로 매칭 시도)
LEE_HEADER_ALIAS = {
    # 그룹 헤더 결합 (R1 group + R2 detail) — 이지현은 R2만 가짐
    # 학생부등급 그룹: '최고', '평균', '50컷', '70컷', '80컷', '90컷', '최저'
    # 대학별환산 그룹: '최고', '평균', '50컷', '70컷', '80컷', '100컷', '총점'
    # 그룹 단순 결합 시 충돌 — 그래서 R1 + R2 점결합 필요. MVP는 인덱스 기반으로 처리.
}


def _build_col_map(r1: tuple, r2: tuple) -> dict[int, str]:
    """R1 그룹 헤더 + R2 세부 헤더 → 캐노니컬 컬럼 인덱스 매핑."""
    group = [None] * len(r2)
    cur = None
    for i, v in enumerate(r1):
        if v is not None and str(v).strip():
            cur = str(v).strip()
        group[i] = cur

    m: dict[int, str] = {}
    for i, name in enumerate(r2):
        if name is None:
            continue
        nm = str(name).strip()
        g = group[i] if i < len(group) else None
        if nm == "대학":           m[i] = "대학"
        elif nm == "전형":         m[i] = "전형"
        elif nm == "모집단위":     m[i] = "모집단위"
        elif nm == "모집인원":     m[i] = "모집인원"
        elif nm == "경쟁률":       m[i] = "경쟁률"
        elif nm == "충원합격순위": m[i] = "충원합격순위"
        elif nm == "반영교과":     m[i] = "반영교과"
        elif g == "학생부등급":
            if nm == "평균":   m[i] = "학생부등급_평균"
            elif nm == "50컷": m[i] = "학생부등급_50컷"
            elif nm == "70컷": m[i] = "학생부등급_70컷"
        elif g == "대학별환산":
            if nm == "총점":   m[i] = "대학별환산_총점"
    return m


def _parse_one_file(fp: Path, unv_cd: str, normalizer: Normalizer) -> list[dict]:
    """per_university/{unvCd}.xlsx 1개 → 행 리스트."""
    try:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    except Exception:
        return []
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    r1 = next(rows_iter, None)
    r2 = next(rows_iter, None)
    if not r2:
        return []
    col_map = _build_col_map(r1 or (), r2)

    records = []
    for row in rows_iter:
        if not row or all(c is None for c in row):
            continue
        rec: dict = {col: None for col in CANONICAL_COLUMNS}
        rec["unvCd"] = unv_cd
        for i, v in enumerate(row):
            cn = col_map.get(i)
            if cn is None:
                continue
            if cn == "전형":
                rec[cn] = normalizer.jeonghyeong(v)
            elif cn == "대학":
                rec[cn] = str(v).strip() if v else None
            elif cn == "모집단위":
                rec[cn] = normalizer.pk(v)
            elif cn == "모집인원":
                rec[cn] = normalizer.integer(v)
            elif cn in ("경쟁률", "학생부등급_평균", "학생부등급_50컷",
                        "학생부등급_70컷", "대학별환산_총점"):
                rec[cn] = normalizer.number(v)
            elif cn == "충원합격순위":
                n = normalizer.integer(v)
                rec[cn] = n if n is not None else (normalizer.cell(v) or None)
            elif cn == "반영교과":
                rec[cn] = normalizer.reflected_subjects(v)
            else:
                rec[cn] = normalizer.cell(v)

        if not all(rec.get(k) for k in PK_COLUMNS):
            continue
        records.append(rec)
    return records


def load(per_university_dir: Path, normalizer: Normalizer) -> dict[str, pd.DataFrame]:
    """이지현 per_university/{unvCd}.xlsx 전체 → {unvCd: DataFrame}."""
    if not per_university_dir.exists():
        return {}

    by_univ: dict[str, list[dict]] = defaultdict(list)
    for fp in sorted(per_university_dir.glob("*.xlsx")):
        if fp.parent != per_university_dir:
            continue
        unv_cd = fp.stem.strip()  # 파일명 = unvCd
        recs = _parse_one_file(fp, unv_cd, normalizer)
        if recs:
            by_univ[unv_cd].extend(recs)

    return {
        unv_cd: pd.DataFrame(rows, columns=CANONICAL_COLUMNS)
        for unv_cd, rows in by_univ.items()
        if rows
    }

"""유찬 산출물(`week03/output/<대학>/<탭>_전형결과.xlsx`) → 캐노니컬 DataFrame.

구조 (가천대 기준 — "단순" 패턴):
  R1: ['모집단위', '<전형명>', '', '', '', '']    ← 첫 셀이 모집단위 라벨, 둘째가 전형명
  R2: ['모집인원', '경쟁률', '충원합격 순위', '학생부등급', '반영교과', '']
  R3+: ['<모집단위>', '<모집인원>', '<경쟁률>', '<충원합격 순위>', '<등급>', '<반영교과>']
  → R2의 헤더가 데이터 col 1+ 위치에 해당 (R3 col 0 = 모집단위, col 1+ = 데이터)

복잡 패턴 (경북대):
  R1: ['단과대학', '모집단위', '지원 및 등록 현황', '등록기준', '최저기준통과', '입학자 학생부 등급', ...]
  → 다단 헤더 + 그룹. MVP에선 단순 패턴만 처리, 복잡 패턴은 스킵+로그.

매핑 키워드:
  - 모집인원: "모집인원" or "모집 인원"
  - 경쟁률: "경쟁률" (단, "실질경쟁률" 등 변형 제외)
  - 충원합격순위: "충원합격" or "충원" + "순위"
  - 학생부등급_평균: "학생부등급" + "평균"
  - 반영교과: "반영" or "교과목"

# adapted from week03/crawl_adiga.py 실측 출력
"""
from __future__ import annotations
import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict

from ..config import PK_COLUMNS, CANONICAL_COLUMNS, GROUP_LABEL_COL
from ..normalizer import Normalizer


# 캐노니컬 컬럼 → R2 헤더 매칭 키워드 (포함 검사)
HEADER_KEYWORDS = {
    "모집인원":      [["모집", "인원"]],   # "모집인원" or "모집 인원"
    "경쟁률":         [["경쟁률"]],         # but exclude when prefixed with 실질/최초
    "충원합격순위":   [["충원", "순위"], ["충원합격"]],
    "학생부등급_평균": [["등급"]],         # 단일 등급 컬럼이면 평균으로 가정
    "반영교과":       [["반영", "교과"], ["교과목"]],
}

EXCLUDE_PATTERNS = {
    "경쟁률": ["실질", "최초"],   # "실질경쟁률", "최초경쟁률"는 일반 경쟁률 아님
}


def _match_canonical(header: str) -> str | None:
    """R2 헤더 텍스트 → 캐노니컬 컬럼명. 일치 안 되면 None."""
    if not header:
        return None
    h = str(header).strip()
    for canonical, patterns in HEADER_KEYWORDS.items():
        for kw_group in patterns:
            if all(kw in h for kw in kw_group):
                # exclude pattern 체크
                if canonical in EXCLUDE_PATTERNS:
                    if any(ex in h for ex in EXCLUDE_PATTERNS[canonical]):
                        continue
                return canonical
    return None


def _parse_simple_table(ws, unv_cd: str, university: str, normalizer: Normalizer) -> list[dict]:
    """단순 패턴(가천대 식) 테이블 1개 파싱."""
    rows_iter = ws.iter_rows(values_only=True)
    r1 = next(rows_iter, None)
    r2 = next(rows_iter, None)
    if not r1 or not r2:
        return []

    # R1 col 0 = "모집단위" 라벨인지 확인. 아니면 복잡 패턴 → 스킵
    if not r1 or len(r1) < 2:
        return []
    if r1[0] is None or "모집단위" not in str(r1[0]):
        return []

    # 전형명 추출: R1 col 1 — normalizer.jeonghyeong()로 정규화
    raw_jeonghyeong = r1[1]
    if raw_jeonghyeong is None:
        return []
    jeonghyeong = normalizer.jeonghyeong(raw_jeonghyeong)
    if not jeonghyeong:
        return []

    # R2 헤더 매핑: R2[i] → 데이터 col (i+1)
    # 단순 패턴에서 R3 col 0 = 모집단위 데이터, col 1+ = 다른 컬럼 데이터
    # 따라서 R2[0] = col 1 헤더, R2[1] = col 2 헤더, ...
    col_to_canonical: dict[int, str] = {0: "모집단위"}  # col 0은 항상 모집단위
    for i, h in enumerate(r2):
        canonical = _match_canonical(h)
        if canonical:
            data_col = i + 1  # shift right by 1
            col_to_canonical[data_col] = canonical

    # 데이터 수집
    records = []
    for row in rows_iter:
        if not row or all(c is None for c in row):
            continue
        rec: dict = {c: None for c in CANONICAL_COLUMNS}
        rec["unvCd"] = unv_cd
        rec["대학"] = university
        rec["전형"] = jeonghyeong

        for i, v in enumerate(row):
            cn = col_to_canonical.get(i)
            if cn is None:
                continue
            if cn == "모집단위":
                rec[cn] = normalizer.pk(v)
            elif cn == "모집인원":
                rec[cn] = normalizer.integer(v)
            elif cn in ("경쟁률", "학생부등급_평균", "학생부등급_50컷",
                        "학생부등급_70컷", "대학별환산_총점"):
                rec[cn] = normalizer.number(v)
            elif cn == "충원합격순위":
                n = normalizer.integer(v)
                rec[cn] = n if n is not None else (normalizer.cell(v) or None)
            elif cn == "반영교과":
                rec[cn] = normalizer.reflected_subjects(v)
            else:
                rec[cn] = normalizer.cell(v)

        if not all(rec.get(k) for k in PK_COLUMNS):
            continue
        records.append(rec)
    return records


def load(output_root: Path, normalizer: Normalizer) -> dict[str, pd.DataFrame]:
    """`week03/output/{unvCd}_{대학명}/*.xlsx` 전체 → {unvCd: DataFrame}.

    폴더명 형식: `{unvCd}_{safe_univ}` (예: `0000063_가천대학교`).
    """
    by_univ: dict[str, list[dict]] = defaultdict(list)
    skipped_complex = []

    if not output_root.exists():
        return {}

    for univ_dir in sorted(output_root.iterdir()):
        if not univ_dir.is_dir():
            continue
        # 폴더명에서 unvCd 추출: "0000063_가천대학교" → ("0000063", "가천대학교")
        folder = univ_dir.name
        if "_" in folder:
            unv_cd, univ_name = folder.split("_", 1)
        else:
            unv_cd, univ_name = folder, folder

        for xlsx in sorted(univ_dir.glob("*.xlsx")):
            try:
                wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            except Exception:
                continue

            for sn in wb.sheetnames:
                ws = wb[sn]
                # 단순 패턴 시도
                recs = _parse_simple_table(ws, unv_cd, univ_name, normalizer)
                if recs:
                    by_univ[unv_cd].extend(recs)
                else:
                    skipped_complex.append(f"{folder}/{xlsx.name}#{sn}")

    if skipped_complex:
        import sys
        print(f"  [유찬 어댑터] 스킵된 복잡 테이블: {len(skipped_complex)}", file=sys.stderr)
        for s in skipped_complex[:5]:
            print(f"    - {s}", file=sys.stderr)
        if len(skipped_complex) > 5:
            print(f"    ... 외 {len(skipped_complex)-5}건", file=sys.stderr)

    return {
        unv_cd: pd.DataFrame(rows, columns=CANONICAL_COLUMNS)
        for unv_cd, rows in by_univ.items()
        if rows
    }

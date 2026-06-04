"""골든셋(`2025_어디가입결_통합본.xlsx`) → 캐노니컬 DataFrame.

구조: 시트 '수시' 1개, 22컬럼.
  R1 = 그룹 헤더 (col6='대학별환산', col13='학생부등급')
  R2 = 세부 헤더 (대학·전형·모집단위·모집인원·경쟁률·충원합격순위 + 그룹 하위)
  R3+ = 데이터

unvCd 컬럼이 없으므로 PK는 대학명 기반 (config.PK_COLUMNS = [대학,전형,모집단위]).
새 골든·유찬 크롤러 둘 다 어디가 기반이라 대학명이 직접 일치한다.

컬럼 매핑 (0-indexed):
  0 대학 · 1 전형 · 2 모집단위 · 3 모집인원 · 4 경쟁률 · 5 충원합격순위
  12 대학별환산_총점 · 14 학생부등급_평균 · 15 학생부등급_50컷 · 16 학생부등급_70컷
  21 반영교과

# adapted from 2025_어디가입결_통합본.xlsx 헤더 실측
"""
from __future__ import annotations
import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict

from ..config import PK_COLUMNS, DATA_COLUMNS, CANONICAL_COLUMNS
from ..normalizer import Normalizer


# 0-indexed 컬럼 → 캐노니컬 컬럼명
GOLDEN_COL_MAP = {
    0:  "대학",
    1:  "전형",
    2:  "모집단위",
    3:  "모집인원",
    4:  "경쟁률",
    5:  "충원합격순위",
    12: "대학별환산_총점",
    14: "학생부등급_평균",
    15: "학생부등급_50컷",
    16: "학생부등급_70컷",
    21: "반영교과",
}


def load(golden_path: Path, normalizer: Normalizer) -> dict[str, pd.DataFrame]:
    """골든셋 → {대학명: DataFrame}. 데이터는 R3부터 (R1·R2는 헤더)."""
    wb = openpyxl.load_workbook(golden_path, read_only=True, data_only=True)
    ws = wb["수시"] if "수시" in wb.sheetnames else wb.active

    by_univ: dict[str, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):  # R3부터 데이터
        if not row or all(c is None for c in row):
            continue
        univ_name = row[0]
        if not univ_name:
            continue
        univ_name = str(univ_name).strip()

        rec: dict = {c: None for c in CANONICAL_COLUMNS}
        for idx, cn in GOLDEN_COL_MAP.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if cn == "대학":
                rec[cn] = univ_name
            elif cn == "전형":
                rec[cn] = normalizer.jeonghyeong(v)
            elif cn == "모집단위":
                rec[cn] = normalizer.pk(v)
            elif cn == "모집인원":
                rec[cn] = normalizer.integer(v)
            elif cn == "경쟁률":
                rec[cn] = normalizer.number(v)
            elif cn == "충원합격순위":
                n = normalizer.integer(v)
                rec[cn] = n if n is not None else (normalizer.cell(v) or None)
            elif cn.startswith("학생부등급_") or cn == "대학별환산_총점":
                rec[cn] = normalizer.number(v)
            elif cn == "반영교과":
                rec[cn] = normalizer.reflected_subjects(v)
            else:
                rec[cn] = normalizer.cell(v)

        # PK 결측 행 제외
        if not all(rec.get(k) for k in PK_COLUMNS):
            continue
        by_univ[univ_name].append(rec)

    return {
        univ: pd.DataFrame(rows, columns=CANONICAL_COLUMNS)
        for univ, rows in by_univ.items()
        if rows
    }

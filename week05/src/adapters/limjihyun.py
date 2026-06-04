"""임지현 산출물(`outputs/{unvCd}.xlsx`) → 캐노니컬 DataFrame.

구조: 대학별 단일 파일, 단일 시트 `수시`.
헤더: R1 그룹 헤더(대학별환산, 학생부등급), R2 세부 헤더. 데이터 R3+.
22컬럼.

매핑: 컬럼명 + 그룹 결합으로 캐노니컬 매칭.

# adapted from 임지현 outputs/0000063.xlsx 실측
"""
from __future__ import annotations
import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict

from ..config import PK_COLUMNS, CANONICAL_COLUMNS, GROUP_LABEL_COL
from ..normalizer import Normalizer


SHEET_NAME = "수시"


def _build_col_map(r1: tuple, r2: tuple) -> dict[int, str]:
    """그룹 헤더 + 세부 헤더 → 캐노니컬 매핑."""
    # R1 forward-fill (병합 셀 대신)
    group = [None] * len(r2)
    cur = None
    for i, v in enumerate(r1):
        if v is not None and str(v).strip():
            cur = str(v).strip()
        group[i] = cur

    m: dict[int, str] = {}
    for i, name in enumerate(r2):
        if name is None:
            continue
        nm = str(name).strip()
        g = group[i] if i < len(group) else None

        if nm == "대학":           m[i] = "대학"
        elif nm == "전형":         m[i] = "전형"
        elif nm == "모집단위":     m[i] = "모집단위"
        elif nm == "모집인원":     m[i] = "모집인원"
        elif nm == "경쟁률":       m[i] = "경쟁률"
        elif nm == "충원합격순위": m[i] = "충원합격순위"
        elif nm == "반영교과":     m[i] = "반영교과"
        elif g == "대학별환산":
            if nm == "총점":   m[i] = "대학별환산_총점"
        elif g == "학생부등급":
            if nm == "평균":   m[i] = "학생부등급_평균"
            elif nm == "50컷": m[i] = "학생부등급_50컷"
            elif nm == "70컷": m[i] = "학생부등급_70컷"
    return m


def load(outputs_dir: Path, normalizer: Normalizer) -> dict[str, pd.DataFrame]:
    """outputs/{unvCd}.xlsx 전체 → {대학명: DataFrame}."""
    by_univ: dict[str, list[dict]] = defaultdict(list)

    if not outputs_dir.exists():
        return {}

    for fp in sorted(outputs_dir.glob("*.xlsx")):
        # 백업 폴더 등 제외
        if fp.parent != outputs_dir:
            continue

        unv_cd = fp.stem.strip()  # 파일명 = unvCd

        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        except Exception:
            continue
        if SHEET_NAME not in wb.sheetnames:
            continue
        ws = wb[SHEET_NAME]

        rows_iter = ws.iter_rows(values_only=True)
        r1 = next(rows_iter, None)
        r2 = next(rows_iter, None)
        if not r2 or not r1:
            continue

        col_map = _build_col_map(r1, r2)

        for row in rows_iter:
            if not row or all(c is None for c in row):
                continue
            rec: dict = {c: None for c in CANONICAL_COLUMNS}
            for i, v in enumerate(row):
                cn = col_map.get(i)
                if cn is None:
                    continue
                if cn == "전형":
                    rec[cn] = normalizer.jeonghyeong(v)
                elif cn == "대학":
                    rec[cn] = normalizer.university(v)   # 표기 통일
                elif cn == "모집단위":
                    rec[cn] = normalizer.pk(v)
                elif cn == "모집인원":
                    rec[cn] = normalizer.integer(v)
                elif cn in ("경쟁률", "학생부등급_평균", "학생부등급_50컷",
                            "학생부등급_70컷", "대학별환산_총점"):
                    rec[cn] = normalizer.number(v)
                elif cn == "충원합격순위":
                    n = normalizer.integer(v)
                    rec[cn] = n if n is not None else (normalizer.cell(v) or None)
                elif cn == "반영교과":
                    rec[cn] = normalizer.reflected_subjects(v)
                else:
                    rec[cn] = normalizer.cell(v)

            if not all(rec.get(k) for k in PK_COLUMNS):
                continue
            uname = rec.get("대학")
            if uname:
                by_univ[uname].append(rec)   # 그룹키 = 대학명(정규화)

    return {
        uname: pd.DataFrame(rows, columns=CANONICAL_COLUMNS)
        for uname, rows in by_univ.items()
        if rows
    }

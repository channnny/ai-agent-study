"""입시결과 상세정보 raw xlsx → 가독성 포맷(다단 병합 헤더) 적용.

규칙(유찬 요청):
  1) "-" → null
  2) 성적 아닌 텍스트("미제출 사유 …")는 셀 null + '비고' 열로 이동
  3) 모든 파일 우측에 '비고' 열
  4) 학생부위주 행 → 수능 열 null, 수능위주 행 → 학생부 열 null
  5) 헤더 배경색 + 상단(및 식별열) 고정
  6) 열 너비 내용 맞춤
  7) 가독성(테두리·정렬·필터)
  + 원본 표와 동일한 4단 병합 헤더 (학생부/수능 → 환산점수·환산등급·백분위 → 50%… → 사탐/과탐/직탐)

사용: python format_detail.py <raw.xlsx> <out.xlsx>
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ID_COLS = ["대학코드", "대학명", "모집 시기", "전형유형", "전형명", "학과명"]
COMMON_NUM = ["모집인원", "경쟁률", "충원인원"]
IDENTITY = ID_COLS + COMMON_NUM
NOTE_KW = ("미제출", "사유")
WIDE_COLS = {"대학명", "전형명", "학과명", "비고"}
NR = 4  # 헤더 단 수

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
ID_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LEFT = Alignment(horizontal="left", vertical="center")
A_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── 데이터 정리 ──────────────────────────────────────────────
def _is_null(s: str) -> bool:
    return s in ("", "-", "nan", "None", "NaT", "–", "ㅡ")


def _num(v):
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", "").replace("%", "").strip())
        return None if f != f else f   # NaN(f!=f) → None
    except (ValueError, TypeError):
        return None


def process(df: pd.DataFrame) -> pd.DataFrame:
    keep = [c for c in df.columns if not re.match(r"^(col\d+|Unnamed)", str(c))]
    df = df[keep].copy()
    hak = [c for c in df.columns if str(c).startswith("학생부")]
    su = [c for c in df.columns if str(c).startswith("수능")]
    idcols = [c for c in ID_COLS if c in df.columns]
    # 공통 수치 블록(모집인원·경쟁률·충원 등)은 어디가 표 변동(예: 모집인원 컬럼 3분할)에 대비해
    # 하드코딩하지 않고, 식별/학생부/수능이 아닌 나머지 값 컬럼을 원본 순서대로 동적 수집.
    mid = [c for c in df.columns if c not in idcols and c not in hak and c not in su]
    data_cols = mid + hak + su

    bigos = []
    for idx, row in df.iterrows():
        # (rule #4 폐기) 전형유형에 따른 반대 블록 비우기는 적용하지 않음.
        #   학생부위주 행의 수능 열 / 수능위주 행의 학생부 열은 원본 그대로 두되,
        #   "-"·미제출 텍스트는 아래 공통 규칙(1·2)으로 처리된다.
        notes = []
        for c in data_cols:
            v = df.at[idx, c]
            s = ("" if v is None else str(v)).strip()
            if _is_null(s):
                df.at[idx, c] = None
            elif any(k in s for k in NOTE_KW):
                notes.append(re.sub(r"^미제출\s*사유\s*[:：]?\s*", "", s).strip() or s)
                df.at[idx, c] = None
        bigos.append("; ".join(dict.fromkeys(notes)) if notes else None)

    df["비고"] = bigos
    # (행 가감 없음) 어디가 팝업 표를 그대로 보존 — 모집인원 0 placeholder 행도 유지.
    order = [c for c in idcols + mid + hak + su + ["비고"] if c in df.columns]
    return df[order]


# ── 다단 헤더 구조 복원 ──────────────────────────────────────
def _parse(col: str):
    """flat 컬럼명 → 4단 경로 (없는 단은 None)."""
    if col in IDENTITY or col == "비고":
        return (col, None, None, None)
    for g in ("학생부 환산점수", "학생부 환산등급", "수능 환산점수"):
        if col.startswith(g + " "):
            top, mid = g.split(" ")
            return (top, mid, col[len(g) + 1:], None)
    if col.startswith("수능 백분위 "):
        rest = col[len("수능 백분위 "):]
        pct, sub = rest.split(" ", 1)
        if sub.startswith("탐구1 ") or sub.startswith("탐구2 "):
            grp, leaf = sub.split(" ", 1)
            return ("수능", "백분위 " + pct, grp, leaf)
        return ("수능", "백분위 " + pct, sub, None)
    return (col, None, None, None)


def _header_merges(columns):
    """(r1,c1,r2,c2,text) 병합 셀 목록 (0-based)."""
    paths = [_parse(c) for c in columns]
    deepest = [max(i for i in range(4) if p[i] is not None) for p in paths]
    merges = []
    n = len(columns)
    for r in range(NR):
        c = 0
        while c < n:
            if r > deepest[c]:
                c += 1
                continue
            key = tuple(paths[c][: r + 1])
            c2 = c
            while c2 + 1 < n and r <= deepest[c2 + 1] and tuple(paths[c2 + 1][: r + 1]) == key:
                c2 += 1
            r2 = NR - 1 if r == deepest[c] else r
            merges.append((r, c, r2, c2, paths[c][r]))
            c = c2 + 1
    return merges


# ── 엑셀 쓰기 ────────────────────────────────────────────────
def _disp_len(s: str) -> int:
    return sum(2 if ord(ch) > 0x1100 else 1 for ch in str(s))


def write_formatted(df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    cols = list(df.columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "입시결과"

    # 헤더 (4단 병합)
    for r1, c1, r2, c2, text in _header_merges(cols):
        ws.cell(row=r1 + 1, column=c1 + 1, value=text)
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1 + 1, start_column=c1 + 1, end_row=r2 + 1, end_column=c2 + 1)
    for r in range(1, NR + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = A_CENTER
            cell.border = BORDER

    # 데이터
    id_idx = {i for i, c in enumerate(cols) if c in ID_COLS}
    hakmyung_i = cols.index("학과명") if "학과명" in cols else -1
    bigo_i = cols.index("비고") if "비고" in cols else -1
    for ri, row in enumerate(df.itertuples(index=False), start=NR + 1):
        for ci, val in enumerate(row):
            cell = ws.cell(row=ri, column=ci + 1, value=(None if pd.isna(val) else val))
            cell.border = BORDER
            if ci == bigo_i:
                cell.alignment = A_LEFT_WRAP
            elif ci in id_idx:
                cell.alignment = A_LEFT
                cell.fill = ID_FILL
                if ci == hakmyung_i:
                    cell.font = Font(bold=True)
            else:
                cell.alignment = A_CENTER

    # 고정(헤더4행+식별열) · 필터 · 열너비
    # 헤더 4행 + 식별열(학과명까지) 고정 → 가로 스크롤해도 대학·학과 식별 가능
    fz_col = (cols.index("학과명") + 2) if "학과명" in cols else 1
    ws.freeze_panes = f"{get_column_letter(fz_col)}{NR + 1}"
    last = get_column_letter(len(cols))
    ws.auto_filter.ref = f"A{NR}:{last}{NR + len(df)}"
    for ci, col in enumerate(cols, 1):
        data_w = max((_disp_len(v) for v in df[col].dropna().astype(str)), default=0)
        leaf = _parse(col)[max(i for i in range(4) if _parse(col)[i] is not None)]
        if col in WIDE_COLS:
            width = min(max(data_w, 12) + 2, 46)
        else:
            width = min(max(data_w, _disp_len(leaf), 5) + 1, 14)
        ws.column_dimensions[get_column_letter(ci)].width = width

    wb.save(out_path)


def main():
    raw, out = Path(sys.argv[1]), Path(sys.argv[2])
    df = process(pd.read_excel(raw))
    write_formatted(df, out)
    print(f"출력 {df.shape} | 비고 {int(df['비고'].notna().sum())}행 | 저장 {out}")


if __name__ == "__main__":
    main()

"""W06 — '입시결과 상세정보'(classUnivAdmssPopup.do) 표준 양식 크롤 + 포맷 + 리포트.

대학별:
  1) 메인 페이지(searchSyr=2027) → '전형 결과' 버튼 upCd(20/30/40) 수집
  2) upCd → slcnTypeCd(20→02 종합, 30→01 교과, 그외→05 수능)
  3) POST classUnivAdmssPopup.do (searchSyr=학년도=2026) → 표2(80열 표준), 총건수 기반 전 페이지 순회
  4) 유형별 concat + 앞열에 대학코드·대학명[캠퍼스] 추가 → format_detail로 포맷
  5) 개별 저장: {대학코드}_{대학명[캠퍼스]} 입시결과 상세.xlsx  (raw 없음)

전체:
  - (전체통합) 입시결과 상세.xlsx  — 모든 대학 합본
  - (크롤링리포트).xlsx           — 시도/성공/실패 + 대학별 어디가총건수 vs 크롤건수 교차검증

실행: cd week06 && ../week05/.venv/bin/python scripts/crawl_2027_detail.py
"""
from __future__ import annotations

import collections
import csv
import json
import math
import pickle
import random
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "week03"))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # scripts/ (format_detail)
import crawl_adiga as C  # noqa: E402
import format_detail as F  # noqa: E402

POPUP_URL = "https://www.adiga.kr/ucp/cls/uni/classUnivAdmssPopup.do"
UPCD_TO_SLCN = {"20": "02", "30": "01"}  # 그 외(40 수능 등) → "05"
SLCN_NAME = {"02": "종합", "01": "교과", "05": "수능"}
OUT = ROOT / "week06" / "output" / "crawl_2027_detail"
SUMMARY_DIR = OUT / "summary"   # 통합본 + 리포트
DETAIL_DIR = OUT / "detail"     # 대학별 폴더
PROGRESS_DIR = OUT / ".progress"  # 대학별 resume 캐시(meta json + 처리 df pkl)

# 표본 10개 (2027 데이터 보유 확인된 대학)
SAMPLE = ["0000056", "0000063", "0000146", "0000003", "0000120",
          "0000200", "0000099", "0000150", "0000088", "0000111"]


# ── 대학명(캠퍼스 구분) ───────────────────────────────────
def _load_names():
    names, bases = {}, collections.Counter()
    with open(C.INPUT_FILE, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            u, nm = row["unv_cd"].strip(), row["univ_name"].strip()
            names[u] = nm
            bases[re.split(r"[\[\(]", nm)[0]] += 1
    return names, {b for b, c in bases.items() if c > 1}


NAMES, SPLIT_BASES = _load_names()


def _pad(s: str, width: int) -> str:
    """표시폭(한글=2칸) 기준 좌측정렬 패딩. 넘치면 표시폭 기준 자르고 '…'."""
    s = str(s)
    if F._disp_len(s) > width:
        out = ""
        for ch in s:
            if F._disp_len(out + ch) > width - 2:  # ".." 자리(2칸) 확보
                break
            out += ch
        s = out + ".."
    return s + " " * (width - F._disp_len(s))


def display_name(unv: str) -> str:
    full = NAMES.get(unv, "")
    base = re.split(r"[\[\(]", full)[0]
    if base not in SPLIT_BASES:
        return base
    paren = re.search(r"\(([^)]+)\)", full)
    brack = re.search(r"\[([^\]]+)\]", full)
    campus = paren.group(1) if paren else (brack.group(1) if brack else "")
    return f"{base}[{campus}]" if campus else base


def _safe(s: str) -> str:
    return re.sub(r'[/\\:*?"<>|]', "", str(s)).strip()


def _pathname(name: str) -> str:
    """폴더·파일명용: 대학명[캠퍼스] → 대학명_캠퍼스 (괄호 제거, 언더스코어)."""
    return _safe(re.sub(r"\[([^\]]+)\]", r"_\1", name))


# ── 네트워크 ─────────────────────────────────────────────
WORKERS = 8                       # 대학 단위 동시 처리 수
DELAY_RANGE = (0.3, 0.8)          # 정상 요청 사이 지연(초) — 봇 패턴 완화

# 실브라우저 흉내: 최신 UA + 표준 헤더(Accept/Language 등)로 핑거프린팅 회피.
BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate",
}


# 서킷브레이커: 전 워커 통틀어 연속 실패가 임계 넘으면 차단 의심 → 신규 요청 중단(fail-fast).
CIRCUIT_THRESHOLD = 8             # 연속 실패 N회(=1웨이브) → 전역 중단
_BLOCK = threading.Event()
_FAIL_LOCK = threading.Lock()
_CONSEC_FAIL = [0]


def _retry_after(r):
    """429 Retry-After 헤더(초) 파싱. http-date 형식은 무시."""
    v = r.headers.get("Retry-After")
    try:
        return float(v) if v else None
    except ValueError:
        return None


def _note_success():
    with _FAIL_LOCK:
        _CONSEC_FAIL[0] = 0


def _note_failure():
    """연속 실패 누적 → 임계 도달 시 차단 의심 플래그 set(1회 경고)."""
    with _FAIL_LOCK:
        _CONSEC_FAIL[0] += 1
        if _CONSEC_FAIL[0] >= CIRCUIT_THRESHOLD and not _BLOCK.is_set():
            _BLOCK.set()
            print(f"\n🛑 연속 실패 {_CONSEC_FAIL[0]}회 — 차단 의심. 신규 요청 중단(서킷브레이커). "
                  f"남은 대학은 빠르게 error 처리 후 종료. 캐시는 보존되니 나중에 이어받기 가능.",
                  flush=True)


def _request(method, url, **kw):
    """순간 네트워크 오류(DNS·타임아웃·5xx) 3회 재시도(점증 백오프). 429는 Retry-After 존중.
    연속 실패가 임계 넘으면(차단 의심) 전역 중단되어 이후 요청은 즉시 실패."""
    if _BLOCK.is_set():
        raise RuntimeError("서킷브레이커 작동 중 — 크롤 중단(차단 의심)")
    kw.setdefault("headers", {})
    kw["headers"] = {**BROWSER_HEADERS, "Connection": "close", **kw["headers"]}
    kw.setdefault("timeout", (5, 15))
    last = None
    for attempt in range(3):
        try:
            time.sleep(random.uniform(*DELAY_RANGE))  # 요청 간 지연 + 지터
            r = requests.request(method, url, **kw)
            if r.status_code == 429:  # 레이트리밋 — 서버 지시대로 더 길게 대기 후 재시도
                wait = _retry_after(r) or 5.0 * (attempt + 1)
                time.sleep(min(wait, 60))
                last = requests.HTTPError("429 Too Many Requests")
                continue
            r.raise_for_status()       # 403·5xx 등 → 아래 except로
            _note_success()
            return r
        except requests.RequestException as e:
            last = e
            time.sleep(1.5 * (attempt + 1))  # 1.5s, 3s
    _note_failure()
    raise last


def _fetch_main(unv: str):
    """메인 페이지(searchSyr=2027) soup."""
    params = {"menuId": "PCUVTINF2000", "unvCd": unv, "searchSyr": "2027"}
    r = _request("GET", C.BASE_URL, params=params)
    return BeautifulSoup(r.text, "html.parser")


# ── 팝업 크롤 ─────────────────────────────────────────────
def _result_buttons(soup):
    out = []
    for acc in soup.find_all("button", class_="accordionBtn"):
        oc = acc.get("onclick", "")
        if "fnItemSearchInclude" in oc and "전형 결과" in acc.get_text():
            m = re.search(r'fnItemSearchInclude\([^,]+,[^,]+,\s*"([^"]*)",\s*"([^"]*)",\s*"([^"]*)"', oc)
            if m:
                out.append(m.groups())
    return out


PAGE_SIZE = 15
_TOTAL_RE = re.compile(r'class="total"[^>]*>\s*총\s*<strong[^>]*>\s*([0-9,]+)')


def _total_count(html: str):
    m = _TOTAL_RE.search(html)
    return int(m.group(1).replace(",", "")) if m else None


def _fetch_detail(syr: str, unv: str, upcd: str):
    """(DataFrame|None, 어디가_총건수). 총건수 기반 전 페이지 순회(wrap 방지)."""
    slcn = UPCD_TO_SLCN.get(upcd, "05")
    base = {"searchSyr": str(int(syr) - 1), "unvCd": unv, "ruCd": "X", "slcnTypeCd": slcn}
    frames, n_pages, total, pg = [], 1, None, 1
    while pg <= n_pages and pg <= 60:
        data = {**base, "pagination.currentPage": str(pg)}
        r = _request("POST", POPUP_URL, data=data, headers={"Referer": C.BASE_URL})
        html = r.text
        if pg == 1:
            total = _total_count(html)
            n_pages = max(1, math.ceil(total / PAGE_SIZE)) if total else 1
        soup = BeautifulSoup(html, "html.parser")
        tbls = soup.find_all("table")
        if len(tbls) < 2:
            break
        t = tbls[1]
        tbody = t.find("tbody")
        if not tbody or not tbody.find_all("tr"):
            break
        grid = C._table_to_grid(t)
        thead = t.find("thead")
        n_hint = len(thead.find_all("tr")) if thead else None
        _, flat, rows = C._flatten_table(grid, n_hint)
        # 빈 전형 안내 행("조회된 자료가 없습니다.")은 실데이터 아님 → 제거(어디가 총건수=0)
        rows = [rr for rr in rows if not any("조회된 자료가 없습니다" in str(c) for c in rr)]
        if rows:
            rows = [rr + [""] * (len(flat) - len(rr)) for rr in rows]
            frames.append(pd.DataFrame(rows, columns=flat))
        pg += 1
    df = pd.concat(frames, ignore_index=True) if frames else None
    return df, total


def crawl_one(unv: str) -> dict:
    """반환: {unv, name, status, expected, crawled, per_type, df, error}."""
    name = display_name(unv) or unv
    rec = {"unv": unv, "name": name, "status": "", "expected": 0, "crawled": 0,
           "per_type": {}, "df": None, "error": ""}
    soup = _fetch_main(unv)
    if not name or name == unv:
        rec["name"] = name = C.get_university_name(soup)
    seen = {}
    for syr, u, upcd in _result_buttons(soup):
        seen.setdefault(upcd, (syr, u, upcd))
    if not seen:
        rec["status"] = "no_section"   # '전형 결과' 섹션 자체가 없음(미게시)
        return rec
    frames = []
    for upcd, (syr, u, _) in sorted(seen.items()):
        nm = SLCN_NAME.get(UPCD_TO_SLCN.get(upcd, "05"), upcd)
        df, total = _fetch_detail(syr, u, upcd)
        n = len(df) if df is not None else 0
        rec["per_type"][nm] = {"expected": total or 0, "crawled": n}
        rec["expected"] += total or 0
        rec["crawled"] += n
        if df is not None and n:
            frames.append(df)
    if not frames:
        rec["status"] = "no_data"
        return rec
    merged = pd.concat(frames, ignore_index=True)
    merged.insert(0, "대학명", rec["name"])    # 매핑용 식별 열
    merged.insert(0, "대학코드", unv)
    rec["df"] = merged
    rec["status"] = "ok" if rec["crawled"] == rec["expected"] else "mismatch"
    return rec


# ── 리포트 ────────────────────────────────────────────────
def _write_report(records, path: Path) -> None:
    n = len(records)
    n_ok = sum(1 for r in records if r["status"] == "ok")
    n_mis = sum(1 for r in records if r["status"] == "mismatch")
    n_nodata = sum(1 for r in records if r["status"] in ("no_data", "no_section"))
    n_err = sum(1 for r in records if r["status"] == "error")
    fail_names = [f"{r['name']}({r['unv']})" for r in records if r["status"] in ("error", "mismatch")]

    summary = pd.DataFrame([
        {"항목": "크롤링 시도 대학", "값": f"{n}개"},
        {"항목": "성공 (어디가 총건수와 완전 일치)", "값": f"{n_ok}개"},
        {"항목": "건수 불일치 (크롤 ≠ 어디가)", "값": f"{n_mis}개"},
        {"항목": "데이터 없음 (미게시·표없음)", "값": f"{n_nodata}개"},
        {"항목": "크롤 오류 (실패)", "값": f"{n_err}개"},
        {"항목": "총 크롤링 행수", "값": f"{sum(r['crawled'] for r in records):,}행"},
        {"항목": "실패/불일치 대학", "값": ", ".join(fail_names) if fail_names else "없음"},
    ])

    STAT_KOR = {"ok": "✅ 완전", "mismatch": "⚠️ 불일치", "no_data": "▫️ 표없음",
                "no_section": "▫️ 미게시", "error": "❌ 오류"}
    rows = []
    for r in records:
        detail = " / ".join(f"{k} {v['crawled']}/{v['expected']}" for k, v in r["per_type"].items())
        rows.append({
            "대학코드": r["unv"], "대학명": r["name"], "상태": STAT_KOR.get(r["status"], r["status"]),
            "어디가 총건수": r["expected"], "크롤링 건수": r["crawled"],
            "일치": "O" if (r["status"] == "ok") else ("X" if r["status"] in ("mismatch", "error") else "-"),
            "유형별(크롤/어디가)": detail,
            "비고": r["error"] or {"no_section": "전형결과 섹션 없음(미게시 추정)", "no_data": "표 없음",
                                  "mismatch": "건수 불일치 — 점검 필요", "ok": ""}.get(r["status"], ""),
        })
    detail_df = pd.DataFrame(rows)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="요약", index=False)
        detail_df.to_excel(w, sheet_name="대학별 완전성", index=False)
        for sh, df in [("요약", summary), ("대학별 완전성", detail_df)]:
            ws = w.sheets[sh]
            ws.freeze_panes = "A2"
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(1, ci)
                cell.fill, cell.font = hfill, hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
                width = max([F._disp_len(col)] + [F._disp_len(v) for v in df[col].astype(str)])
                ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 8), 60)


# ── 메인 ──────────────────────────────────────────────────
def _has_data(status):
    """status상 처리 df(pkl)가 있어야 하는 경우."""
    return status in ("ok", "mismatch")


def _save_progress(rec, pdf):
    """대학별 resume 캐시 저장. pkl 먼저 → json(커밋 마커) 순서로 원자성 확보."""
    PROGRESS_DIR.mkdir(parents=True, exist_ok=True)
    if pdf is not None:
        (PROGRESS_DIR / f"{rec['unv']}.pkl").write_bytes(pickle.dumps(pdf))
    meta = {k: v for k, v in rec.items() if k != "df"}
    (PROGRESS_DIR / f"{rec['unv']}.json").write_text(  # 마지막에 기록 = 완료 표시
        json.dumps(meta, ensure_ascii=False), encoding="utf-8")


def _load_progress(unv):
    """완결된 캐시면 (rec, pdf), 아니면 None(재크롤). pdf는 no_data 등에서 None."""
    mp = PROGRESS_DIR / f"{unv}.json"
    if not mp.exists():
        return None
    try:
        rec = json.loads(mp.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, ValueError):
        return None
    rec["df"] = None
    pp = PROGRESS_DIR / f"{unv}.pkl"
    pdf = pickle.loads(pp.read_bytes()) if pp.exists() else None
    if _has_data(rec.get("status", "")) and pdf is None:
        return None  # 데이터 있어야 하는데 pkl 누락(부분 기록) → 재크롤
    return rec, pdf


_PRINT_LOCK = threading.Lock()
_DONE = [0]


def _process_one(unv, total, fresh):
    """단일 대학: 캐시 확인 → 크롤 → 저장. 반환 (rec, pdf, skipped)."""
    cached = None if fresh else _load_progress(unv)
    if cached is not None:
        rec, pdf = cached
        skipped = True
    else:
        try:
            rec = crawl_one(unv)
        except Exception as e:
            rec = {"unv": unv, "name": display_name(unv) or unv, "status": "error",
                   "expected": 0, "crawled": 0, "per_type": {}, "df": None,
                   "error": f"{type(e).__name__}: {e}"}
        pdf = None
        if rec["df"] is not None:
            try:  # 포맷/쓰기 실패가 풀(executor) 전체를 막지 않도록 격리
                pdf = F.process(rec["df"])
                stem = f"{rec['unv']}_{_pathname(rec['name'])}"   # 0000003_강원대학교_본교
                folder = DETAIL_DIR / stem
                folder.mkdir(parents=True, exist_ok=True)
                F.write_formatted(pdf, folder / f"{stem}_입시결과상세.xlsx")
            except Exception as e:
                pdf = None
                rec["status"] = "error"
                rec["error"] = f"format/write {type(e).__name__}: {e}"
        _save_progress(rec, pdf)  # 즉시 캐시 → 중단돼도 이어받기
        skipped = False
    with _PRINT_LOCK:
        _DONE[0] += 1
        mark = ("↻" if skipped else
                {"ok": "✓", "mismatch": "⚠", "no_data": "·", "no_section": "·", "error": "✗"}.get(rec["status"], "?"))
        idx = f"[{_DONE[0]}/{total}]"
        print(f"  {mark} {_pad(rec['name'], 26)} ({unv})  {idx:<10} "
              f"크롤 {rec['crawled']:4d} / 어디가 {rec['expected']:4d}  [{rec['status']}]"
              + (f"  {rec['error']}" if rec["error"] else ""), flush=True)
    return unv, rec, pdf, skipped


def main():
    # 기본 = csv 전체 대학(모든 캠퍼스 코드). --sample → 표본 10개만.
    targets = SAMPLE if "--sample" in sys.argv else list(NAMES.keys())
    fresh = "--fresh" in sys.argv  # 캐시 무시하고 처음부터
    SUMMARY_DIR.mkdir(parents=True, exist_ok=True)
    DETAIL_DIR.mkdir(parents=True, exist_ok=True)
    print(f"=== 입시결과 상세정보(표준양식) 크롤 — 대상 {len(targets)}개 대학 (동시 {WORKERS}) ===")
    print(f"출력: {OUT}  (summary/ + detail/)\n")
    _DONE[0] = 0
    t0 = time.time()
    results = {}
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = [ex.submit(_process_one, unv, len(targets), fresh) for unv in targets]
        for fut in as_completed(futs):
            unv, rec, pdf, skipped = fut.result()
            results[unv] = (rec, pdf, skipped)

    # csv 원래 순서로 재조립 (통합본·리포트 행 순서 보존)
    records, processed, n_skip = [], [], 0
    for unv in targets:
        rec, pdf, skipped = results[unv]
        records.append(rec)
        if pdf is not None:
            processed.append(pdf)
        if skipped:
            n_skip += 1

    if n_skip:
        print(f"\n(캐시 이어받기 {n_skip}개 / 신규 크롤 {len(targets) - n_skip}개)")
    if processed:
        combined = pd.concat(processed, ignore_index=True)
        F.write_formatted(combined, SUMMARY_DIR / "입시결과상세_전체.xlsx")
        print(f"\n✓ 통합본: {len(combined):,}행 → summary/입시결과상세_전체.xlsx")

    _write_report(records, SUMMARY_DIR / "크롤링_리포트.xlsx")
    n_ok = sum(1 for r in records if r["status"] == "ok")
    n_mis = sum(1 for r in records if r["status"] == "mismatch")
    n_nd = sum(1 for r in records if r["status"] in ("no_data", "no_section"))
    n_er = sum(1 for r in records if r["status"] == "error")
    print(f"\n=== 요약 ===  시도 {len(records)} | 성공 {n_ok} | 불일치 {n_mis} | 데이터없음 {n_nd} | 오류 {n_er}")
    for r in records:
        if r["status"] in ("error", "mismatch", "no_data", "no_section"):
            print(f"  ⚠ {r['name']}({r['unv']}): {r['status']} (크롤 {r['crawled']}/어디가 {r['expected']}) {r['error']}")
    m, s = divmod(int(time.time() - t0), 60)
    print(f"\n⏱  소요 시간: {m}분 {s}초")


if __name__ == "__main__":
    main()

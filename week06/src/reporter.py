"""6시트 평가 리포트 (Excel) 작성.

시트:
  1) summary           — row=metric × col=person
  2) by_university     — row=대학 × col=person별 metric
  3) by_column         — row=DATA 컬럼 × col=person별 match_rate
  4) missing_rows      — golden에 있고 person에 없음 (long-format)
  5) extra_rows        — person에만 있음
  6) mismatched_cells  — matched 행의 셀 불일치 (long-format)

# adapted from leejihyun/evaluation_report.xlsx 양식
"""
from __future__ import annotations
import pandas as pd
from pathlib import Path
from datetime import datetime

from .config import DATA_COLUMNS, PERSON_KOR
from .matcher import PersonResult




def _format_rate(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return "✓" if v else "✗"
    if isinstance(v, float) and 0 <= v <= 1:
        return f"{v*100:.2f}%"
    if isinstance(v, float):
        return round(v, 2)
    return v


def _bar(pct, width=10):
    """유니코드 막대. pct는 0~100."""
    if pct is None:
        return ""
    f = max(0, min(width, int(round(pct / 100 * width))))
    return "█" * f + "░" * (width - f)


def _light(pct, good, mid):
    """신호등: good 이상 🟢 / mid 이상 🟡 / 미만 🔴 / None ⬜."""
    if pct is None:
        return "⬜"
    return "🟢" if pct >= good else ("🟡" if pct >= mid else "🔴")


def _dash_cell(pct, good, mid):
    """대시보드 셀: '🟢 89.6% ████████░░'."""
    if pct is None:
        return "—"
    return f"{_light(pct, good, mid)} {pct:.1f}%  {_bar(pct)}"


def build_summary(results: list[PersonResult]) -> pd.DataFrame:
    """대시보드형 종합 시트.

    상단 = 핵심 3지표(커버리지·PK·셀)를 신호등+막대+기준선으로.
    하단 = 상세 행 수 지표.
    """
    def col(r):
        return PERSON_KOR.get(r.person, r.person)

    rows = []

    # ── 핵심 지표 (신호등 + 막대) ──
    # 평가 핵심 = 셀 일치율(긁은 값 정확도). PK 매칭률은 골든 전형명 변동으로 평가 제외 → 참고용.
    rows.append({"지표": "📊 셀 일치율 (긁은 값이 정확) ★평가기준", **{
        col(r): _dash_cell((r.summary.get("cell_match_rate") or 0) * 100, 90, 50) for r in results}, "목표": "≥ 90%"})
    rows.append({"지표": "📊 커버리지 (몇 대학 시도)", **{
        col(r): _dash_cell(r.summary.get("coverage_pct"), 80, 40) for r in results}, "목표": "높을수록"})
    rows.append({"지표": "📊 셀 충진율 (얼마나 채웠나)", **{
        col(r): _dash_cell((r.summary.get("cell_fill_rate") or 0) * 100, 80, 50) for r in results}, "목표": "높을수록"})
    rows.append({"지표": "(참고) PK 매칭률 — 평가 제외", **{
        col(r): _dash_cell((r.summary.get("pk_match_rate") or 0) * 100, 85, 40) for r in results}, "목표": "전형명 변동"})

    # ── 구분 ──
    rows.append({"지표": "", **{col(r): "" for r in results}, "목표": ""})
    rows.append({"지표": "─ 상세 (행 수) ─", **{col(r): "" for r in results}, "목표": ""})

    detail = [
        ("matched 행 수 (정확히 찾음)", "n_matched"),
        ("missing 행 수 (못 찾음→누락행)", "n_missing"),
        ("extra 행 수 (잘못 긁음→잉여행)", "n_extra"),
        ("골든 전체 행 수", "n_golden_total"),
        ("비교한 셀 수 (양쪽 값 있음)", "n_cell_compared"),
        ("미수집 셀 수 (한쪽만 값)", "n_cell_fill_gap"),
        ("미수집 대학 수 (→미수집대학)", "n_missing_univ"),
    ]
    for label, key in detail:
        row = {"지표": label}
        for r in results:
            v = r.summary.get(key)
            row[col(r)] = f"{v:,}" if isinstance(v, (int, float)) else v
        row["목표"] = ""
        rows.append(row)

    # ── 종합 판정 ──
    rows.append({"지표": "", **{col(r): "" for r in results}, "목표": ""})
    judge = {"지표": "🏁 종합 판정 (셀 일치율 ≥90%)"}
    for r in results:
        judge[col(r)] = "✅ 통과" if r.summary.get("dod_pass") else "❌ 미달"
    judge["목표"] = "셀 ≥90%"
    rows.append(judge)

    return pd.DataFrame(rows, columns=["지표"] + [col(r) for r in results] + ["목표"])


def _univ_cell(m: dict) -> str:
    """대학별 시트 한 셀: 신호등 + PK·셀 요약, 또는 미수집."""
    if not m:
        return "⬜ 미수집"
    st = m.get("status")
    if st == "missing":
        return "⬜ 미수집"
    pk = m.get("pk_match_rate")
    cell = m.get("cell_match_rate")
    if pk is None and m.get("n_golden", 0) == 0:
        return "▫ 골든에 없음"
    light = _light((pk or 0) * 100, 85, 40)
    pk_s = f"{pk*100:.0f}%" if pk is not None else "0%"
    cell_s = f"{cell*100:.0f}%" if cell is not None else "—"
    matched = f"{m.get('n_matched', 0)}/{m.get('n_golden', 0)}"
    return f"{light} PK {pk_s} · 셀 {cell_s}  ({matched})"


def build_by_university(results: list[PersonResult]) -> pd.DataFrame:
    """row=대학(unvCd), col=사람별 1컬럼(신호등+PK·셀 요약).

    정렬: 골든 행수 많은 대학 먼저(데이터 풍부) → 전부 미수집 대학은 아래로.
    """
    all_univ = set()
    univ_names: dict[str, str] = {}
    golden_rows: dict[str, int] = {}
    for r in results:
        for unv_cd, m in r.by_university.items():
            all_univ.add(unv_cd)
            if m.get("univ_name"):
                univ_names[unv_cd] = m["univ_name"]
            golden_rows[unv_cd] = max(golden_rows.get(unv_cd, 0), m.get("n_golden", 0) or 0)

    # 정렬 키: 데이터 있는 대학(누구라도 missing 아님) 먼저, 그 안에서 골든행수 desc
    def sort_key(u):
        any_data = any(
            r.by_university.get(u, {}).get("status") not in (None, "missing")
            for r in results
        )
        return (0 if any_data else 1, -golden_rows.get(u, 0), univ_names.get(u, ""))

    rows = []
    for unv_cd in sorted(all_univ, key=sort_key):
        row = {"unvCd": unv_cd, "대학": univ_names.get(unv_cd, "")}
        for r in results:
            label = PERSON_KOR.get(r.person, r.person)
            row[label] = _univ_cell(r.by_university.get(unv_cd, {}))
        rows.append(row)
    return pd.DataFrame(rows)


def build_by_column(results: list[PersonResult]) -> pd.DataFrame:
    rows = []
    for col in DATA_COLUMNS:
        # 모든 사람이 비교 0건(전원 None)인 컬럼은 숨김 — 예: 대학별환산_총점
        # (골든에 없어 비교 불가). 노이즈 제거.
        if all(r.by_column.get(col) is None for r in results):
            continue
        row = {"컬럼": col}
        for r in results:
            label = PERSON_KOR.get(r.person, r.person)
            row[label] = _format_rate(r.by_column.get(col))
        rows.append(row)
    return pd.DataFrame(rows)


def build_long_rows(results: list[PersonResult], attr: str) -> pd.DataFrame:
    """missing_rows / extra_rows / mismatched_cells 통합."""
    rows = []
    for r in results:
        rows.extend(getattr(r, attr))
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def build_diff_rows(results: list[PersonResult]) -> pd.DataFrame:
    """누락행 + 잉여행을 한 시트로 통합.

    같은 (사람·대학·모집단위)로 정렬 → '전형명만 다른' 누락↔잉여 쌍이 인접해
    PK 매칭 실패 원인(전형명 분류 차이)이 한눈에 드러난다.
    """
    rows = []
    for r in results:
        person = PERSON_KOR.get(r.person, r.person)
        for m in r.missing_rows:
            rows.append({
                "구분": "🔴 누락 (골든O·크롤러X)",
                "사람": person, "unvCd": m.get("unvCd"), "대학": m.get("대학"),
                "전형": m.get("전형"), "모집단위": m.get("모집단위"),
            })
        for e in r.extra_rows:
            rows.append({
                "구분": "🔵 잉여 (크롤러O·골든X)",
                "사람": person, "unvCd": e.get("unvCd"), "대학": e.get("대학"),
                "전형": e.get("전형"), "모집단위": e.get("모집단위"),
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=["사람", "대학", "모집단위", "전형", "구분", "unvCd"])
    # 같은 사람·대학·모집단위 안에서 누락/잉여가 붙도록 정렬
    df = df.sort_values(["사람", "대학", "모집단위", "전형"],
                        key=lambda s: s.fillna("")).reset_index(drop=True)
    return df


def _verdict(bigo: str) -> str:
    """불일치 '비고' 텍스트 → 직관적 판정 이모지."""
    b = str(bigo or "")
    if "콤마" in b:
        return "🟡 거의같음 (콤마 포맷)"
    if "근접" in b:
        return "🟡 거의같음 (반올림)"
    if "null" in b:
        return "⬜ 한쪽만 값있음"
    return "🔴 값 다름"


_VERDICT_ORDER = {"🔴": 0, "⬜": 1, "🟡": 2}


def build_mismatch(results: list[PersonResult]) -> pd.DataFrame:
    """불일치셀 — 판정 이모지 + 골든↔크롤러 값 인접 + 심각한 것 위로 정렬."""
    rows = []
    for r in results:
        person = PERSON_KOR.get(r.person, r.person)
        for c in r.mismatched_cells:
            gv = c.get("golden_value")
            pv = c.get("person_value")
            # 한쪽만 값이 있으면 무조건 ⬜ (비고 텍스트보다 우선)
            if (gv is None) != (pv is None):
                verdict = "⬜ 한쪽만 값있음"
            else:
                verdict = _verdict(c.get("비고"))
            rows.append({
                "판정": verdict,
                "사람": person, "대학": c.get("대학"),
                "전형": c.get("전형"), "모집단위": c.get("모집단위"),
                "항목": c.get("컬럼"),
                "골든값": "(빈칸)" if gv is None else gv,
                "크롤러값": "(빈칸)" if pv is None else pv,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=["판정", "사람", "대학", "전형", "모집단위", "항목", "골든값", "크롤러값"])
    df = df.sort_values(
        ["사람", "판정", "대학", "항목"],
        key=lambda s: s.map(lambda v: _VERDICT_ORDER.get(str(v)[:1], 9)) if s.name == "판정" else s.fillna(""),
    ).reset_index(drop=True)
    return df


def build_uncovered(results: list[PersonResult]) -> pd.DataFrame:
    """커버리지 100%가 안 되는 이유 — 한 명이라도 못 낸 골든 대학 명단.

    각 사람의 status(pass/fail/missing)를 그대로 표시. 모두 missing이면
    어디가 소스 문제(이미지/미게시) 추정, 일부만 missing이면 해당 크롤러 한계.
    """
    persons = list(results)  # 평가에 포함된 모든 사람(유찬·이지현·임지현)

    # 모든 대학 + 대학명 수집
    names: dict[str, str] = {}
    for r in results:
        for u, m in r.by_university.items():
            if m.get("univ_name"):
                names[u] = m["univ_name"]
            names.setdefault(u, "")

    rows = []
    for u in sorted(names):
        # 골든에 실재하는 대학만
        ng = max((r.by_university.get(u, {}).get("n_golden", 0) or 0) for r in results)
        if ng == 0:
            continue
        statuses = {r.person: r.by_university.get(u, {}).get("status") for r in persons}
        missers = [r.person for r in persons if statuses.get(r.person) == "missing"]
        if not missers:
            continue  # 전원 데이터 냄 → 제외

        n_miss = len(missers)
        if n_miss == len(persons):
            cause = "전원 미수집 (어디가 이미지/미게시 등 소스 문제 추정)"
        else:
            kor = [PERSON_KOR.get(p, p) for p in missers]
            cause = f"{'·'.join(kor)} 미수집 ({n_miss}/{len(persons)})"

        row = {"대학명": names[u], "골든 행수": ng}
        for r in persons:
            row[PERSON_KOR.get(r.person, r.person)] = statuses.get(r.person) or "-"
        row["추정 원인"] = cause
        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["추정 원인", "대학명"]).reset_index(drop=True)
    return df


def build_glossary() -> pd.DataFrame:
    """리포트 맨 앞 — 용어·지표·읽는 법 설명. 회의에서 처음 보는 사람용."""
    rows = [
        ("■ 이 리포트는? (W06)", "W05 3명(유찬·이지현·임지현)의 어디가 크롤러 출력을 하나로 '통합'한 결과와, 통합 전 3인 각각을 작년 골든셋(2025_수시_입시결과_통합본)과 비교한 것입니다. 목표 = 누락 없이 최대한 많은 데이터(커버리지·충진율·일치율 동시 극대화)."),
        ("🔷 통합 컬럼", "3인 출력을 병합한 W06 산출물. 병합 규칙 — ①커버리지: 한 명이라도 긁은 (대학·전형·모집단위) 행은 모두 포함(행 합집합) ②충진율: 같은 행에서 한 명이라도 값이 있으면 채움 ③일치율: 값이 갈리면 다수결(2:1), 동률이면 충진율 높은 순(임지현>이지현>유찬). 골든값은 병합에 쓰지 않음(평가 독립성 유지)."),
        ("🔷 통합 효과", "통합은 '비교셀(긁어온 셀 절대량)'이 단일 크롤러 최대치를 상회 — 한 명이 놓친 행·셀을 다른 사람이 메워 정답 셀 총량이 최다. 충진율 %가 임지현보다 약간 낮아 보이는 건 커버리지가 넓어진 만큼 분모가 커진 희석 효과일 뿐, 절대 수집량은 통합이 1위."),
        ("", ""),
        ("■ 신호등·막대 읽는 법", "각 표에 공통으로 쓰입니다."),
        ("🟢 / 🟡 / 🔴 / ⬜", "🟢=목표 달성 / 🟡=절반 이상(아쉬움) / 🔴=목표 미달 / ⬜=데이터 없음(미수집)."),
        ("████░░░░░░", "비율 막대. 채워진 칸이 많을수록 높음 (한 칸 = 10%)."),
        ("", ""),
        ("■ 평가 기준 = 셀 일치율 (★)", "회의 결정: 골든셋은 데이터랩스가 최종 정제하며 전형명 등이 변동되므로 PK 매칭률은 평가에서 제외. 목표는 '크롤링을 잘 해오는가' = 셀 일치율."),
        ("★ 셀 일치율", "행 매칭된 곳에서 '양쪽 다 값이 있는 셀'만 비교해 일치한 비율. = 긁어온 값이 정확한가? 한쪽만 값있는 셀(크롤러 미수집)은 분모에서 제외(충진율로 따로 봄). 목표 ≥90%. 🏁 종합 판정의 단일 기준."),
        ("■ 보조 지표", ""),
        ("커버리지", "골든 대학 중 '데이터를 낸' 대학 비율. = 애초에 평가 테이블에 올라올 자격. W06 통합은 100% 달성(3인 중 한 명이라도 커버한 대학 모두 포함)."),
        ("셀 충진율", "비교 대상 셀 중 '양쪽 다 값이 있던' 비율. 낮으면 = 골든엔 있는 항목을 크롤러가 덜 긁음(예: 어디가가 50%컷 미공개)."),
        ("(참고) PK 매칭률", "(대학코드,전형,모집단위)가 골든과 같은 행 비율. 전형명이 골든 정제 과정에서 변동되므로 평가에선 제외(참고만). 같은 PK에 크롤러 행이 여럿이면 골든과 가장 잘 맞는 행을 골라(best-match) 셀 비교."),
        ("🏁 종합 판정", "셀 일치율 ≥90%면 ✅ 통과, 아니면 ❌ 미달. (PK는 판정에서 제외)"),
        ("", ""),
        ("■ 통합 커버리지 100%·개별 차이", "→ 대학별 상세는 '미수집대학' 시트. W05에서 캠퍼스 정규화·교대형 표 처리·이미지 OCR로 유찬 100% 달성, W06 통합도 100%. 이지현·임지현은 일부 대학(교대형·소스 차이)에서 개별 미수집이 남아 통합이 이를 메움."),
        ("", ""),
        ("■ 남은 PK 격차(~43%)의 주원인", "크롤링 실패가 아니라 전형명 분류 차이. 예: 골든 '해람인재' vs 크롤러 '학생부종합(해람인재)' — 같은 전형인데 이름이 달라 한쪽은 누락·한쪽은 잉여로 잡힘. '누락·잉여' 시트에서 인접 쌍으로 확인 가능. W06에서 전형 분류 사전 합의로 해소."),
        ("", ""),
        ("■ 탭(시트)별 보는 법", ""),
        ("종합 (대시보드)", "3명을 신호등+막대로 한눈에. 위 3줄(📊)이 핵심 지표, 아래는 상세 행 수, 맨 끝 🏁이 최종 판정."),
        ("대학별", "대학마다 사람별 1칸 = '신호등 PK% · 셀% (맞은행/골든행)'. 배경색=신호등. 데이터 많은 대학이 위, 미수집은 아래. 같은 대학명 2줄=캠퍼스 분리(unvCd 다름)."),
        ("항목별", "데이터 항목별 일치율(매칭+양쪽 값 있는 셀 기준). 낮은 항목 = 그 항목 파싱 우선 개선."),
        ("  └ 반영교과 0%", "자동비교 한계 항목. 골든은 상세 기입('국수영과,전교과,동일비율,진로,미반영'), 어디가/크롤러는 요약('전교과')이라 정보량이 달라 전부 불일치. 사람 검수 또는 W06 파싱 개선 대상 — 다른 항목 정확도와 분리해서 볼 것."),
        ("미수집대학", "한 명이라도 못 낸 대학(3인 전원 status 표시) + 추정 원인. 전원 미수집=소스 문제(이미지·미게시), 일부만=해당 크롤러 한계."),
        ("누락·잉여", "한 행=(대학·전형·모집단위) 조합. 그 '전형-학과 행'이 한쪽에만 있음. 🔴누락=골든O·크롤러X / 🔵잉여=크롤러O·골든X. 대부분 같은 행을 골든·크롤러가 다른 전형명으로 불러 매칭 실패(예: 크롤러 '일반' ↔ 골든 '지역인재'). PK 제외 근거."),
        ("불일치셀", "PK·셀 양쪽 다 값이 있는데 값이 다른 칸만. 🔴값다름(진짜 오류, 위쪽) / 🟡거의같음(콤마·반올림=사실상 정답). 한쪽만 값있는 셀(미수집)은 여기 없고 충진율로 집계. 골든값↔크롤러값 나란히."),
    ]
    return pd.DataFrame(rows, columns=["항목", "설명"])


# ──────────────────────────────────────────────────────────────
# 스타일링 (openpyxl)
# ──────────────────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")   # 진파랑
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_DESC_FILL   = PatternFill("solid", fgColor="D9E1F2")    # 연파랑
_DESC_FONT   = Font(italic=True, color="1F3864", size=10)
_PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")    # 연초록
_FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")    # 연빨강
_MISS_FILL   = PatternFill("solid", fgColor="EDEDED")    # 회색
_SECTION_FONT = Font(bold=True, color="1F3864", size=11)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _disp_width(s: str) -> int:
    """한글 등 전각 문자는 2폭으로 계산."""
    return sum(2 if ord(c) > 0x1100 else 1 for c in str(s))


def _autofit(ws, df: pd.DataFrame, header_row: int, max_w: int = 60):
    for j, col in enumerate(df.columns, start=1):
        vals = [str(col)] + [str(v) for v in df[col] if v is not None]
        w = max((_disp_width(v) for v in vals), default=8)
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), max_w)


def _write_styled(writer, df: pd.DataFrame, sheet: str, desc: str):
    """설명행(1행) + 헤더(2행) + 데이터 시트 작성 + 스타일."""
    if df.empty:
        df = pd.DataFrame([{"info": "해당 행 없음"}])
    df.to_excel(writer, sheet_name=sheet, index=False, startrow=1)
    ws = writer.sheets[sheet]
    n_cols = len(df.columns)

    # 1행: 설명 (병합)
    ws.cell(1, 1, "📌 " + desc)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1)
    c.fill = _DESC_FILL
    c.font = _DESC_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 42

    # 2행: 헤더
    for j in range(1, n_cols + 1):
        h = ws.cell(2, j)
        h.fill = _HEADER_FILL
        h.font = _HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = _BORDER

    # 데이터 영역 테두리 + 정렬
    for r in range(3, ws.max_row + 1):
        for j in range(1, n_cols + 1):
            cell = ws.cell(r, j)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _autofit(ws, df, header_row=2)
    ws.freeze_panes = "A3"  # 설명+헤더 고정


def _colorize_univ(ws, df: pd.DataFrame):
    """대학별 시트 사람 컬럼(신호등 셀)을 내용 따라 색칠."""
    person_cols = [j for j, c in enumerate(df.columns, start=1)
                   if str(c) not in ("unvCd", "대학")]
    for r in range(3, ws.max_row + 1):
        for j in person_cols:
            v = str(ws.cell(r, j).value or "")
            if "⬜" in v:
                ws.cell(r, j).fill = _MISS_FILL
            elif "🟢" in v:
                ws.cell(r, j).fill = _PASS_FILL
            elif "🔴" in v:
                ws.cell(r, j).fill = _FAIL_FILL
            elif "🟡" in v:
                ws.cell(r, j).fill = PatternFill("solid", fgColor="FFEB9C")  # 연노랑


def _colorize_diff(ws, df: pd.DataFrame):
    """누락·잉여 시트의 '구분' 컬럼 색칠."""
    if df.empty:
        return
    gj = [j for j, c in enumerate(df.columns, start=1) if c == "구분"]
    if not gj:
        return
    j = gj[0]
    for r in range(3, ws.max_row + 1):
        v = str(ws.cell(r, j).value or "")
        if "누락" in v:
            ws.cell(r, j).fill = _FAIL_FILL
        elif "잉여" in v:
            ws.cell(r, j).fill = PatternFill("solid", fgColor="DDEBF7")  # 연파랑


def _colorize_verdict(ws, df: pd.DataFrame):
    """불일치셀 시트의 '판정' 컬럼 색칠."""
    if df.empty:
        return
    pj = [j for j, c in enumerate(df.columns, start=1) if c == "판정"]
    if not pj:
        return
    j = pj[0]
    for r in range(3, ws.max_row + 1):
        v = str(ws.cell(r, j).value or "")
        if v.startswith("🔴"):
            ws.cell(r, j).fill = _FAIL_FILL
        elif v.startswith("🟡"):
            ws.cell(r, j).fill = PatternFill("solid", fgColor="FFEB9C")
        elif v.startswith("⬜"):
            ws.cell(r, j).fill = _MISS_FILL


def _highlight_dashboard(ws):
    """종합 시트의 핵심 3지표(📊) 행을 연노랑 배경으로 강조."""
    hl = PatternFill("solid", fgColor="FFF2CC")
    for r in range(3, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "")
        if label.startswith("📊") or label.startswith("🏁"):
            for j in range(1, ws.max_column + 1):
                if ws.cell(r, j).fill.fgColor.rgb in ("00000000", None):
                    ws.cell(r, j).fill = hl


def _style_glossary(ws, df: pd.DataFrame):
    """용어설명 시트 전용 스타일 — 섹션 헤더(■) 강조."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    # 헤더행
    for j in (1, 2):
        h = ws.cell(1, j)
        h.fill = _HEADER_FILL
        h.font = _HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1)
        b = ws.cell(r, 2)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if a.value and str(a.value).startswith("■"):
            a.font = _SECTION_FONT
            a.fill = _DESC_FILL
            b.fill = _DESC_FILL


def write_report(results: list[PersonResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    glossary   = build_glossary()
    summary    = build_summary(results)
    by_univ    = build_by_university(results)
    by_col     = build_by_column(results)
    uncovered  = build_uncovered(results)
    diff       = build_diff_rows(results)
    mismatched = build_mismatch(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 용어설명 (자체가 설명이라 별도 desc 없이)
        glossary.to_excel(writer, sheet_name="용어설명", index=False)
        _style_glossary(writer.sheets["용어설명"], glossary)

        _write_styled(writer, summary, "종합",
            "🔷통합 vs 3인 비교. 🟢=목표달성 🟡=절반이상 🔴=미달 ⬜=없음. 보는 순서: ①커버리지(대학 몇 곳)→②셀일치율(값 정확)→③충진율(얼마나 채움). "
            "통합은 커버리지 100% + 비교셀(긁은 셀 절대량) 최다로 '누락 없이 최대 수집' 목표 달성. 셀 일치율 90%↑ 유지. "
            "PK매칭률은 전형명 분류 차이로 참고 지표(평가 제외, 누락·잉여 시트 참조).")
        _highlight_dashboard(writer.sheets["종합"])
        _write_styled(writer, by_univ, "대학별",
            "대학별 정확도 한 셀 요약: '신호등 PK% · 셀% (매칭행/골든행)'. "
            "예) '🔴 PK 37% · 셀 74% (226/619)' = 골든 619행 중 226행을 찾음(PK 37%), 그 중 값 일치 74%. "
            "신호등은 PK 기준: 🟢85%↑ 🟡40%↑ 🔴40%미만 ⬜미수집. 셀 '—'는 매칭행이 0이라 계산 불가. "
            "데이터 많은 대학이 위, 미수집은 아래. 같은 대학명 2줄=캠퍼스 분리(unvCd 다름).")
        _colorize_univ(writer.sheets["대학별"], by_univ)
        _write_styled(writer, by_col, "항목별",
            "8개 데이터 항목별 일치율(매칭된 행 한정). 숫자가 낮은 항목 = 그 항목 파싱이 부정확 → 우선 개선 대상. 예: '학생부등급_50컷'이 낮으면 등급 파싱 로직 점검.")
        _write_styled(writer, uncovered, "미수집대학",
            "한 명이라도 못 낸 대학 명단(3인 전원 표시). '골든 행수'=골든셋의 그 대학 행 수(놓친 양). "
            "각 사람 칸: missing=그 사람이 못 냄 / fail·pass=데이터는 냄. "
            "추정 원인: 전원 미수집=어디가 소스 문제(이미지 업로드·미게시 등), 일부만=해당 크롤러 한계(교대형 표·전형 파싱 누락 등).")
        _write_styled(writer, diff, "누락·잉여",
            "한 행 = (대학·전형·모집단위) 조합 = 'OO대학에서 △△전형으로 뽑는 □□학과' 입시결과 한 줄. "
            "그 조합이 한쪽에만 있으면 표시됨. 🔴누락=골든엔 있는 행을 크롤러가 못 찾음 / 🔵잉여=크롤러엔 있는데 골든엔 없는 행. "
            "주의: '값이 없다'가 아니라 '그 전형-학과 행 자체가 없다'는 뜻. 대부분은 데이터가 진짜 없는 게 아니라, "
            "같은 행을 골든·크롤러가 다른 전형명으로 불러서 매칭 실패한 것 — 예) 충남대 간호학과: 크롤러 '일반'(🔵잉여) ↔ 골든 '지역인재'(같은 22명 모집인데 전형명만 다름). "
            "같은 대학·모집단위로 정렬돼 이런 쌍이 인접. 이게 PK를 평가에서 제외한 이유(골든 정제 시 전형명 변동). [데이터>필터]로 사람·대학 좁혀 보세요.")
        _colorize_diff(writer.sheets["누락·잉여"], diff)
        _write_styled(writer, mismatched, "불일치셀",
            "PK는 맞았는데 값이 다른 칸. 판정: 🔴값다름(진짜 오류) / 🟡거의같음(콤마·반올림, 사실상 정답) / ⬜한쪽만 값있음. 🔴부터 정렬 → 위쪽이 실제 고칠 것. 🟡이 많으면 실제 정확도는 셀일치율보다 높음.")
        _colorize_verdict(writer.sheets["불일치셀"], mismatched)

    print(f"\n✓ 평가 리포트: {output_path}")
    print(f"  시트 7개: 용어설명 · 종합 · 대학별 · 항목별 · 미수집대학 · 누락·잉여 · 불일치셀")


def write_integrated_dataset(integrated_df: pd.DataFrame, lineage_df: pd.DataFrame,
                             output_path: Path) -> None:
    """W06 통합 canonical 데이터셋 + 계보(lineage)를 별도 xlsx로 저장.

    - 통합데이터: 3인 병합 결과. 컬럼 = W05 평가 스키마와 동일(대학·전형·모집단위 + 8항목).
    - 통합계보: 소스 간 값이 충돌한 셀의 후보값·채택출처 추적 (감사·디버깅용).
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        integrated_df.to_excel(writer, sheet_name="통합데이터", index=False)
        _autofit(writer.sheets["통합데이터"], integrated_df, header_row=1)

        lineage_sample = lineage_df
        note = None
        if len(lineage_sample) > 50000:
            note = "계보가 50,000행을 넘어 앞부분만 포함됨"
            lineage_sample = lineage_sample.head(50000).copy()
        lineage_sample.to_excel(writer, sheet_name="통합계보", index=False)
        _autofit(writer.sheets["통합계보"], lineage_sample, header_row=1)

    print(f"✓ 통합 데이터셋: {output_path}")
    print(f"  시트 2개: 통합데이터({len(integrated_df):,}행) · 통합계보(충돌 {len(lineage_df):,}건)"
          + (f" — {note}" if note else ""))

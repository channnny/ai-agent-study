"""2뷰 합본 미리보기 — 한 워크북에 열람(블록) + 관리(tidy) 시트.

  · 열람(전형별)   : 원본 표 블록 (사람 검수)
  · 전형일정및방법 : 1행/레코드 평탄 (관리·조인)
  · 전형요소       : 1행/레코드 tidy (관리)
공통 PK(unvCd_comScsbjtCd_ruSn) + 대학·전형·모집단위로 두 뷰 연결.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import parse_schedule as S  # noqa: E402
import parse_element as EL  # noqa: E402
import write_blocks as B  # noqa: E402

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")        # 소제목(컬럼명)
GROUP_FILL = PatternFill("solid", fgColor="2E5C8A")       # 대제목(그룹)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
GROUP_FONT = Font(bold=True, color="FFFFFF", size=10)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 전형일정및방법 23열의 대제목 그룹 (라벨, 컬럼수)
SCHED_GROUPS = [("식별", 4), ("전형일정", 6), ("전형 요소별 반영비율", 13)]


def _pk(u: dict) -> str:
    return f"{u['unvCd']}_{u['comScsbjtCd']}_{u['ruSn']}"


def _sheet(ws, rows: list[dict], columns: list[str] | None, groups=None):
    """2단 헤더(대제목 병합 + 소제목) + 데이터. groups=[(라벨,열수)...]."""
    if not rows:
        ws.append(["(데이터 없음)"])
        return
    cols = columns or list({k: None for r in rows for k in r})
    if groups is None:  # 식별 4열 + 나머지 한 그룹
        groups = [("식별", 4), ("전형요소", len(cols) - 4)]
    # row1 대제목(병합), row2 소제목
    ci = 1
    for label, n in groups:
        ws.cell(row=1, column=ci, value=label)
        if n > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + n - 1)
        for k in range(n):
            g = ws.cell(row=1, column=ci + k)
            g.fill = GROUP_FILL
            g.font = GROUP_FONT
            g.alignment = CEN
        ci += n
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=2, column=j, value=name)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = CEN
    for ri, r in enumerate(rows, start=3):  # 데이터 row3부터
        for j, name in enumerate(cols, 1):
            ws.cell(row=ri, column=j, value=r.get(name, ""))
    ws.freeze_panes = "A3"
    B.autofit(ws)


def build(units: list[dict], 대학명: str, out_path: Path):
    sched_all, elem_all, pages = [], [], {}
    for u in units:
        hs, he = B.fetch_pages(u)
        pages[id(u)] = (hs, he)
        pk = _pk(u)
        전형 = u.get("전형명", "").split(">")[-1].strip()
        sched = S.parse(hs, 대학명=대학명, 전형명=전형)
        for r in sched:
            if not r.get("모집단위명"):
                r["모집단위명"] = u.get("학과명", "")
            sched_all.append({"PK": pk, **r})
        for r in EL.parse(he, 대학명=대학명, 전형명=전형):
            elem_all.append({"PK": pk, "모집단위명": u.get("학과명", ""), **r})

    wb = Workbook()
    # 1) 열람(블록)
    ws_view = wb.active
    ws_view.title = "열람(전형별)"
    B.render_blocks(ws_view, units, 대학명, pages=pages)
    # 2) 관리 tidy (2단 헤더: 대제목 병합 + 소제목)
    _sheet(wb.create_sheet("전형일정및방법"), sched_all, ["PK"] + list(S.COLUMNS), groups=SCHED_GROUPS)
    _sheet(wb.create_sheet("전형요소"), elem_all, None)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"저장: {out_path}  | 열람 {ws_view.max_row}행 / 일정 {len(sched_all)}행 / 요소 {len(elem_all)}행")


def main():
    ROOT = Path(__file__).resolve().parents[2]
    units = json.load(open(ROOT / "week09" / "output" / "enum" / "가천대.json", encoding="utf-8"))
    ai = [u for u in units if u["학과명"].startswith("AI인문")][:5]
    print("샘플:", [u["전형명"].split(">")[-1].strip() for u in ai])
    build(ai, "가천대학교[본교]", ROOT / "week09" / "output" / "preview" / "가천대_샘플5_2뷰합본.xlsx")


if __name__ == "__main__":
    main()

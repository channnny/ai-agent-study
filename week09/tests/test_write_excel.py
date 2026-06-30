import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1] / "scripts"))
import write_excel as W, parse_schedule as S
import openpyxl

def test_two_tabs(tmp_path):
    sched=[{**{c:"" for c in S.COLUMNS}, "대학명":"가천대","전형명":"논술","모집단위명":"AI인문대학","선발모형":"일괄합산"}]
    elem=[{"대학명":"가천대","전형명":"논술","표이름":"학생부 교과성적 반영방법","반영교과":"국영수","비율":"100"}]
    out=tmp_path/"가천대.xlsx"
    W.write_university(out, sched, elem)
    wb=openpyxl.load_workbook(out)
    assert wb.sheetnames==["전형일정및방법","전형요소"]
    assert wb["전형일정및방법"].max_row>=2 and wb["전형요소"].max_row>=2

def test_combined(tmp_path):
    sched=[{**{c:"" for c in S.COLUMNS}, "대학명":"가천대"}]
    out=tmp_path/"통합.xlsx"; W.write_combined(out,[("가천대",sched,[])])
    assert "전형일정및방법" in openpyxl.load_workbook(out).sheetnames

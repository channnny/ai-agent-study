"""포맷 B — 전형(모집단위)별 블록. 어디가 원본 표를 colspan/rowspan 그대로 워크시트에 적층.

한 (전형×모집단위) = 한 블록:
  [블록 제목: 대학 / 전형 / 모집단위]
  ▸ 전형일정 (원본 표)
  ▸ 전형 요소별 반영비율 (원본 표)
  ▸ 지원자격 및 기타 / 최저학력기준 (원본 표)
  ▸ (전형요소) 학생부 학년별/요소별 반영비율 · 교과성적 반영방법 (원본 매트릭스)
블록 사이 빈 행. 원본 표 모양 100% 보존(병합셀 유지).
"""
import sys
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import net  # noqa: E402
import enumerate_admissions as E  # noqa: E402

DETAIL_URL = "https://www.adiga.kr/ucp/prc/uni/admssUnivDetail.do"
ELEMENT_URL = "https://www.adiga.kr/ucp/prc/uni/admssUnivDetailElement.do"
SYR = "2027"

# detail.do에서 살릴 원본 표(캡션). 성적분석 팝업 등 제외.
DETAIL_CAPTIONS = ["전형일정", "전형 요소별 반영비율", "지원자격 및 기타", "최저학력기준"]

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
TITLE_FONT = Font(bold=True, size=12, color="FFFFFF")
# 블록 경계 = 주황(테이블 헤더 청색과 강한 대비 → 전형 구분 확실)
TITLE_BAND = PatternFill("solid", fgColor="C55A11")    # 블록 제목 밴드(주황)
DIVIDER_FILL = PatternFill("solid", fgColor="ED7D31")  # 블록 구분 divider(밝은 주황)
CAP_FONT = Font(bold=True, size=10, color="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BANDW = 13  # 제목/divider 밴드 폭(열)


def _disp(s) -> int:
    return sum(2 if ord(ch) > 0x1100 else 1 for ch in str(s))


def autofit(ws, cap: int = 52, min_w: int = 6):
    """열별 최대 표시폭(한글=2)으로 너비 자동 맞춤."""
    from openpyxl.utils import get_column_letter as _gl
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max((_disp(line) for line in str(cell.value).split("\n")), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for col, w in widths.items():
        ws.column_dimensions[_gl(col)].width = min(max(w + 2, min_w), cap)


def _moonit(html: str) -> str:
    """전형일정 페이지 div.selectionInfo → 모집단위명 (예 'AI인문대학')."""
    import re
    soup = BeautifulSoup(html, "html.parser")
    si = soup.select_one("div.selectionInfo")
    if not si:
        return ""
    m = re.search(r"\]\s*(.+?)\s*/\s*\S*대학", si.get_text(" ", strip=True))
    return m.group(1).strip() if m else ""


def _grid(table):
    """HTML table → (matrix dict[(r,c)]=text, merges[(r1,c1,r2,c2)], nrows, ncols, header_is_th[r])."""
    rows = table.find_all("tr")
    cells, merges, occ = {}, [], set()
    header_row = []
    for r, tr in enumerate(rows):
        ths = tr.find_all(["th", "td"])
        header_row.append(all(c.name == "th" for c in ths) and bool(ths))
        c = 0
        for cell in ths:
            while (r, c) in occ:
                c += 1
            rs, cs = int(cell.get("rowspan", 1)), int(cell.get("colspan", 1))
            cells[(r, c)] = cell.get_text(" ", strip=True)
            if rs > 1 or cs > 1:
                merges.append((r, c, r + rs - 1, c + cs - 1))
            for dr in range(rs):
                for dc in range(cs):
                    occ.add((r + dr, c + dc))
            c += cs
    nrows = len(rows)
    ncols = max((c for (_, c) in cells), default=-1) + 1
    return cells, merges, nrows, ncols, header_row


def _render_table(ws, table, top: int, col0: int = 1) -> int:
    """table을 (top, col0)에서 원본 그대로 그림. 다음 빈 행 번호 반환."""
    cells, merges, nrows, ncols, header_row = _grid(table)
    for (r, c), txt in cells.items():
        cell = ws.cell(row=top + r, column=col0 + c, value=txt)
        cell.border = BORDER
        cell.alignment = CENTER if header_row[r] else LEFT
        if header_row[r]:
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
    for (r1, c1, r2, c2) in merges:
        ws.merge_cells(start_row=top + r1, start_column=col0 + c1,
                       end_row=top + r2, end_column=col0 + c2)
    # 빈 셀에도 테두리(표 모양 유지)
    for r in range(nrows):
        for c in range(ncols):
            cc = ws.cell(row=top + r, column=col0 + c)
            if cc.border.left.style is None:
                cc.border = BORDER
    return top + nrows


def _tables_from(html: str, captions=None, section_only=False):
    """캡션 매칭 표 또는 div.admssDtlSection 내 표 추출."""
    soup = BeautifulSoup(html, "html.parser")
    out = []
    if section_only:
        for t in soup.select("div.admssDtlSection table"):
            cap = t.find("caption")
            out.append((cap.get_text(strip=True) if cap else "", t))
    else:
        for t in soup.find_all("table"):
            cap = t.find("caption")
            name = cap.get_text(strip=True) if cap else ""
            if captions and not any(name.startswith(k) for k in captions):
                continue
            out.append((name, t))
    return out


def _params(u: dict) -> dict:
    return {"user": "", "cnrtYear": "2026", "unvSeCd": "10", "menuId": "PCPRCINF2000",
            "searchSyr": SYR, **{k: u[k] for k in E.PARAM_KEYS}}


def fetch_pages(u: dict) -> tuple:
    """(전형일정및방법 html, 전형요소 html) GET."""
    p = _params(u)
    hs = net._request("GET", DETAIL_URL, params=p).text
    ep = {k: v for k, v in p.items() if k not in ("user", "cnrtYear", "unvSeCd")}
    ep["admssInfoTabYn"] = ""
    he = net._request("GET", ELEMENT_URL, params=ep).text
    return hs, he


def render_blocks(ws, units: list[dict], 대학명: str, pages: dict | None = None):
    """전형별 블록을 워크시트에 적층. pages={id:(hs,he)} 주면 재요청 안 함."""
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 16
    row = 1
    for u in units:
        if pages is not None:
            hs, he = pages[id(u)]
        else:
            hs, he = fetch_pages(u)
        소속 = (u.get("전형명", "").split(">")[-1].strip()) or u.get("전형명", "")
        모집단위 = _moonit(hs) or u.get("학과명", "")
        # 블록 제목 밴드: 전체 폭 진청 배경 + 흰 굵은 글씨
        tcell = ws.cell(row=row, column=1, value=f"◆ {대학명}  |  {소속}  /  {모집단위}")
        tcell.font = TITLE_FONT
        for c in range(1, BANDW + 1):
            ws.cell(row=row, column=c).fill = TITLE_BAND
        ws.row_dimensions[row].height = 22
        row += 2

        blocks = _tables_from(hs, captions=DETAIL_CAPTIONS) + _tables_from(he, section_only=True)
        if not any(name for name, _ in blocks):
            ws.cell(row=row, column=1, value="(표 없음)")
            row += 1
        for name, table in blocks:
            cap = ws.cell(row=row, column=1, value=f"▸ {name}")
            cap.font = CAP_FONT
            row += 1
            row = _render_table(ws, table, row) + 1  # 표 + 빈 행
        # 블록 구분 divider: 전체 폭 주황 밴드(굵게)
        for c in range(1, BANDW + 1):
            ws.cell(row=row, column=c).fill = DIVIDER_FILL
        ws.row_dimensions[row].height = 14
        row += 3  # 블록 간 간격(빈 행 2)
    autofit(ws)


def render_university(units: list[dict], 대학명: str, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "전형정보"
    render_blocks(ws, units, 대학명)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"저장: {out_path}")


def main():
    import json
    ROOT = Path(__file__).resolve().parents[2]
    units = json.load(open(ROOT / "week10" / "output" / "enum" / "가천대.json", encoding="utf-8"))
    ai = [u for u in units if u["학과명"].startswith("AI인문")][:5]
    print("샘플 5 전형:", [u["전형명"].split(">")[-1].strip() for u in ai])
    render_university(ai, "가천대학교[본교]",
                      ROOT / "week10" / "output" / "preview" / "가천대_샘플5_포맷B.xlsx")


if __name__ == "__main__":
    main()

"""포맷 C — 구조화 엑셀. 데이터랩스 첨부 `가천대 (1).xlsx` 스키마와 정합.

원본 표를 평탄화해 3단 헤더(대=caption / 중=헤더행0 / 소=헤더행1) + 데이터행을 만든다.
식별 7열(학년도/대학명/대학코드/전형명/전형코드/모집단위명/모집단위코드) 을 앞에 붙인다.

시트:
  ① 전형일정및방법 (46열) — 전형일정 / 전형요소별반영비율 / 지원자격및기타 / 최저학력기준
  ② 전형요소        (26열) — 학생부학년별요소별 / 학생부교과성적반영방법
  ③ 열람(전형별)           — write_blocks.render_blocks 원본 블록 뷰
"""
import sys
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import net               # noqa: E402
import write_blocks as B  # noqa: E402
import enumerate_admissions as E  # noqa: E402

SYR = "2027"

# ── 스타일 상수 ────────────────────────────────────────────────
_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")  # 진청 — 헤더 배경
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── 첨부 정합 헤더 스키마 (대, 중, 소) ─────────────────────────
# 순서가 첨부 열 순서와 일치해야 한다.
_IDENT_HEADERS = [
    (None, None, "adiga_selcntnm"),
    (None, None, "학년도"),
    (None, None, "대학명"),
    (None, None, "대학코드"),
    (None, None, "전형명"),
    (None, None, "전형코드"),
    (None, None, "모집단위명"),
    (None, None, "모집단위코드"),
]
N_IDENT = len(_IDENT_HEADERS)  # 8

# 전형일정및방법 (39열) — 식별 7 + 이하 39 = 46열
# 첨부 가천대 (1).xlsx 스키마와 열 순서 완전 일치 (H=8 ~ AT=46)
_SCHED_HEADERS = [
    # 전형일정 (H~M = 6열)
    ("전형일정", "원서접수",      "인터넷"),         # H=8
    ("전형일정", "원서접수",      "현장"),            # I=9
    ("전형일정", "대학별고사일정", "논술등필답"),      # J=10
    ("전형일정", "대학별고사일정", "면접구술"),        # K=11
    ("전형일정", "대학별고사일정", "실기"),            # L=12
    ("전형일정", "합격자발표일",  "합격자발표일"),     # M=13 (row2~3 병합)
    # 전형 요소별 반영비율 (선발모형부터 — 빈 열 제거)
    ("전형 요소별 반영비율", "선발모형",    "선발모형"),    # O=15
    ("전형 요소별 반영비율", "선발방법",    "선발방법"),    # P=16
    ("전형 요소별 반영비율", "선발비율(%)", "선발비율(%)"), # Q=17
    ("전형 요소별 반영비율", "반영비율",    "학생부"),      # R=18
    ("전형 요소별 반영비율", "반영비율",    "수능"),        # S=19
    ("전형 요소별 반영비율", "반영비율",    "면접"),        # T=20
    ("전형 요소별 반영비율", "반영비율",    "논술"),        # U=21
    ("전형 요소별 반영비율", "반영비율",    "적성"),        # V=22
    ("전형 요소별 반영비율", "반영비율",    "1단계성적"),   # W=23
    ("전형 요소별 반영비율", "반영비율",    "실기"),        # X=24
    ("전형 요소별 반영비율", "반영비율",    "서류"),        # Y=25
    ("전형 요소별 반영비율", "반영비율",    "기타"),        # Z=26
    ("전형 요소별 반영비율", "기타내용",    "기타내용"),    # AA=27
    # 지원자격 및 기타 (AB~AH = 7열)
    ("지원자격 및 기타", "지원자격유형",   "지원자격유형"),  # AB=28
    ("지원자격 및 기타", "세부지원자격",   "세부지원자격"),  # AC=29
    ("지원자격 및 기타", "학력유형",       "학력유형"),      # AD=30
    ("지원자격 및 기타", "공통지원자격",   "공통지원자격"),  # AE=31
    ("지원자격 및 기타", "졸업년도",       "제한여부"),      # AF=32 (원본 하위셀1)
    ("지원자격 및 기타", "졸업년도",       "제한여부"),      # AG=33 (원본 하위셀2 — 라벨은 병합헤더 그대로, 값은 복제하지 않음)
    ("지원자격 및 기타", "기타(추가사항)", "기타(추가사항)"),# AH=34
    # 최저학력기준 (AI~AW = 15열) — 모든 수능 과목 [선택반영,과목명] 2열로 일관
    ("최저학력기준", "학생부",  "학생부"),                    # AI=35
    ("최저학력기준", "국어",    "선택반영"),                  # AJ=36
    ("최저학력기준", "국어",    "과목명"),                    # AK=37
    ("최저학력기준", "수학",    "선택반영"),                  # AL=38
    ("최저학력기준", "수학",    "과목명"),                    # AM=39
    ("최저학력기준", "영어",    "영어"),                      # AN=40 (단일 — 영어는 과목명 없음)
    ("최저학력기준", "탐구영역", "선택반영"),                 # AP=42
    ("최저학력기준", "탐구영역", "과목명"),                   # AQ=43
    ("최저학력기준", "제2외국어/한문", "제2외국어/한문"),     # 단일 — 과목명 없음
    ("최저학력기준", "한국사",  "한국사"),                    # 단일 — 과목명 없음
    ("최저학력기준", "반영영역수", "반영영역수"),             # AV=48
    ("최저학력기준", "세부내용", "세부내용"),                 # AW=49
]

# 전형요소 (47열) — 식별 8 + 이하 47 = 55열 (첨부 가천대 (1)_양식수정.xlsx 스키마 일치)
_ELEM_HEADERS = [
    # 수능 영역별 반영비율 (I~S)
    ("수능 영역별 반영비율", "활용지표", "활용지표"),
    ("수능 영역별 반영비율", "영역수", "영역수"),
    ("수능 영역별 반영비율", "국어", "반영비율"),
    ("수능 영역별 반영비율", "국어", "선택과목"),
    ("수능 영역별 반영비율", "수학", "반영비율"),
    ("수능 영역별 반영비율", "수학", "선택과목"),
    ("수능 영역별 반영비율", "영어", "영어"),
    ("수능 영역별 반영비율", "탐구영역", "반영비율"),
    ("수능 영역별 반영비율", "탐구영역", "선택과목"),
    ("수능 영역별 반영비율", "제2외국어/한문", "제2외국어/한문"),
    ("수능 영역별 반영비율", "한국사", "한국사"),
    # 수능 영역별 등급기준 (T~W) — 다중행(1~9등급) → 컬럼별 ' / ' join
    ("수능 영역별 등급기준", "구분", "구분"),
    ("수능 영역별 등급기준", "영어", "영어"),
    ("수능 영역별 등급기준", "제2외국어/한문", "제2외국어/한문"),
    ("수능 영역별 등급기준", "한국사", "한국사"),
    # 수능 가(감)산부여 (X~AC)
    ("수능 가(감)산부여", "국어", "국어"),
    ("수능 가(감)산부여", "수학", "수학"),
    ("수능 가(감)산부여", "영어", "영어"),
    ("수능 가(감)산부여", "탐구", "탐구"),
    ("수능 가(감)산부여", "제2외국어/한문", "제2외국어/한문"),
    ("수능 가(감)산부여", "한국사", "한국사"),
    # 수능 탐구영역 반영사항 (AD~AE)
    ("수능 탐구영역 반영사항", "반영과목수", "반영과목수"),
    ("수능 탐구영역 반영사항", "지정과목", "지정과목"),
    # 비고 (AF)
    ("비고", "비고", "비고"),
    # 학생부 학년별/요소별 반영비율 (AG~AR)
    ("학생부 학년별/요소별 반영비율", "적용대상 졸업년도",     "적용대상 졸업년도"),
    ("학생부 학년별/요소별 반영비율", "학년별 반영비율(100%)", "학년공통"),
    ("학생부 학년별/요소별 반영비율", "학년별 반영비율(100%)", "공통비율"),
    ("학생부 학년별/요소별 반영비율", "학년별 반영비율(100%)", "1학년"),
    ("학생부 학년별/요소별 반영비율", "학년별 반영비율(100%)", "2학년"),
    ("학생부 학년별/요소별 반영비율", "학년별 반영비율(100%)", "3학년"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "교과"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "출결"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "자격"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "활동"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "봉사활동"),
    ("학생부 학년별/요소별 반영비율", "요소별 반영비율(100%)", "기타"),
    # 서류 (AS~AV)
    ("서류", "학교생활기록부", "학교생활기록부"),
    ("서류", "특기자실적", "특기자실적"),
    ("서류", "기타", "기타"),
    ("서류", "기타내용", "기타내용"),
    # 학생부 교과성적 반영방법 (AW~BC)
    ("학생부 교과성적 반영방법", "학년",               "학년"),
    ("학생부 교과성적 반영방법", "반영교과",           "교과"),
    ("학생부 교과성적 반영방법", "반영교과",           "공통과목"),
    ("학생부 교과성적 반영방법", "반영교과",           "일반선택"),
    ("학생부 교과성적 반영방법", "반영교과",           "진로선택"),
    ("학생부 교과성적 반영방법", "점수산출 활용지표",   "점수산출 활용지표"),
    ("학생부 교과성적 반영방법", "반영교과 및 반영방법", "반영교과 및 반영방법"),
]

# caption → (스키마 내 컬럼 인덱스 목록, 원본표 ncols 기대)
# 원본표의 각 데이터 컬럼이 스키마의 몇 번째(0-based) 항목에 매핑되는지 순서대로 나열.
# None이면 해당 원본 컬럼을 스킵.

# ── 원본 표 → 스키마 값 추출 ────────────────────────────────
PLACEHOLDER = "대학에서 입력된 정보가 없습니다."


def _cell(cells: dict, r: int, c: int) -> str:
    """그리드 셀 읽기 — 원본 '정보 없음' 플레이스홀더는 빈칸으로 치환(피드백#3)."""
    v = cells.get((r, c), "") or ""
    return "" if PLACEHOLDER in v else v


def _join(vals) -> str:
    """다중행 값 → ' / ' join, 빈칸 보존(전 표 공통 원칙 — 피드백#1/#5/#7).
    예: [a,b,'',e] -> 'a / b /  / e'. 전부 빈값이면 ''. 단일값이면 그대로(구분자 없음)."""
    vals = list(vals) if vals else [""]
    if not any(v for v in vals):
        return ""
    return vals[0] if len(vals) == 1 else " / ".join(vals)


def _bigo(html: str) -> str:
    """전형요소 페이지의 '비고' 섹션 텍스트(피드백#2). caption 있는 <table>이 아니라
    admssDtlSection 내 <h4>비고</h4> + desc div 구조라 _tables_from()으로는 안 잡힘."""
    soup = BeautifulSoup(html, "html.parser")
    for sec in soup.select("div.admssDtlSection"):
        h4 = sec.select_one(".titArea h4")
        if h4 and h4.get_text(strip=True) == "비고":
            desc = sec.select_one(".desc")
            txt = (desc or sec).get_text(" ", strip=True)
            return "" if PLACEHOLDER in txt else txt
    return ""


def _expand_grid(cells, nrows, ncols, hrow):
    """header 행 수 탐지 후 데이터 행 반환. 헤더는 th 전체가 True인 행."""
    # 헤더 행 = hrow[r] == True
    data_rows = [r for r in range(nrows) if not hrow[r]]
    # 교과성적처럼 중간에 헤더행이 다시 나타나는 경우 제거
    # (예: row6 '반영교과 및 반영방법' th행 → 스킵)
    # data_rows는 이미 hrow==False인 행만 포함
    return data_rows


def extract_table_data(table, cap_name: str, schema_keys: list[str]) -> dict[str, list[str]]:
    """표 → {소제목 키: [값,...]}. 교과성적은 컬럼별 ' / ' join 처리."""
    cells, merges, nrows, ncols, hrow = B._grid(table)
    if nrows == 0 or ncols == 0:
        return {k: [""] for k in schema_keys}

    # 헤더 행 0, 1 (없으면 빈 dict)
    def hcell(r, c):
        return cells.get((r, c), "") or ""

    # 데이터 행 (hrow==False)
    data_rows = [r for r in range(nrows) if not hrow[r]]

    # 원본 표의 헤더행 수 파악 (연속된 hrow True 행들)
    n_header = 0
    for r in range(nrows):
        if hrow[r]:
            n_header += 1
        else:
            break

    # 헤더1(row0), 헤더2(row1) 라벨 — 병합 확장
    h1 = {}  # col → 중제목
    h2 = {}  # col → 소제목

    # 병합 맵 구축 (원본 표 내부)
    occ1 = {}  # (r,c) → 원본 값 (for 병합 확장)
    for (r1, c1, r2, c2) in merges:
        val = cells.get((r1, c1), "")
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                occ1[(r, c)] = val

    def _hval(r, c):
        if (r, c) in occ1:
            return occ1[(r, c)]
        return cells.get((r, c), "") or ""

    for c in range(ncols):
        h1[c] = _hval(0, c) if n_header >= 1 else ""
        h2[c] = _hval(1, c) if n_header >= 2 else ""

    # 표 caption 기준으로 소제목(schema_keys) 매핑
    # 방법: 소제목 라벨이 일치하는 컬럼을 찾아 매핑.
    # 동일 소제목이 여러 열이면 순서대로 배분.
    result: dict[str, list[str]] = {k: [] for k in schema_keys}

    if cap_name == "전형일정":
        # ncols=6: 인터넷, 현장, 논술등필답, 면접구술, 실기, 합격자발표일
        col_map = {
            "인터넷":    0,
            "현장":      1,
            "논술등필답": 2,
            "면접구술":  3,
            "실기":      4,
            "합격자발표일": 5,
        }
        for dr in data_rows:
            for k, c in col_map.items():
                if k in result:
                    result[k].append(_cell(cells, dr, c))

    elif cap_name == "전형 요소별 반영비율":
        # ncols=13: 선발모형,선발방법,선발비율(%),반영비율x9,기타내용
        # h2[3..11] = 학생부,수능,면접,논술,적성,1단계성적,실기,서류,기타
        key_map = [
            "선발모형", "선발방법", "선발비율(%)",
            "학생부", "수능", "면접", "논술", "적성", "1단계성적", "실기", "서류", "기타",
            "기타내용",
        ]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "지원자격 및 기타":
        # ncols=7: 지원자격유형, 세부지원자격, 학력유형, 공통지원자격, 졸업년도(col4,col5 2개 하위셀), 기타(추가사항)
        # 원본 졸업년도 하위 2셀을 AF(제한여부_a)/AG(제한여부_b)에 각각 그대로 매핑 — 값 복제 금지.
        key_map = [
            "지원자격유형", "세부지원자격", "학력유형", "공통지원자격",
            "제한여부_a", "제한여부_b", "기타(추가사항)",
        ]
        for dr in data_rows:
            vals = {k: _cell(cells, dr, i) for i, k in enumerate(key_map)}
            for k in ["지원자격유형", "세부지원자격", "학력유형", "공통지원자격", "기타(추가사항)"]:
                if k in result:
                    result[k].append(vals[k])
            if "제한여부_a" in result:
                result["제한여부_a"].append(vals["제한여부_a"])
            if "제한여부_b" in result:
                result["제한여부_b"].append(vals["제한여부_b"])

    elif cap_name == "최저학력기준":
        # 원본 표: col0=학생부, 세부내용=마지막 컬럼.
        # 수능 6개 과목(국어/수학/영어/탐구영역/제2외국어·한문/한국사)은
        # 헤더 colspan에 따라 [선택반영, 과목명] 2 하위셀 또는 [선택반영] 1 하위셀.
        # 헤더행1(수능 하위 과목명, row index 1)의 병합(colspan) 폭으로 소스 컬럼 폭을 판별한다.
        subj_order = ["국어", "수학", "영어", "탐구영역", "제2외국어/한문", "한국사"]
        # (r1,c1,r2,c2) merges 중 header row(=1, 두번째 헤더행)에서 시작하는 것 → 과목별 colspan
        span_by_col = {}  # 시작col → colspan
        for (r1, c1, r2, c2) in merges:
            if r1 == 1 == r2:
                span_by_col[c1] = c2 - c1 + 1

        # col1부터 시작해 과목 순서대로 소스 컬럼 폭 소비 (col0=학생부, 마지막=세부내용 제외)
        subj_cols: dict[str, list[int]] = {}
        c = 1
        last_col = ncols - 1  # 세부내용
        # 반영영역수는 항상 폭 1, 세부내용 직전 컬럼
        # 과목 6개 + 반영영역수(1) 가 col1..last_col-1 을 채운다
        for subj in subj_order:
            width = span_by_col.get(c, 1)
            subj_cols[subj] = [c] if width == 1 else [c, c + 1]
            c += width
        반영영역수_col = c

        for dr in data_rows:
            if "학생부" in result:
                result["학생부"].append(_cell(cells, dr, 0))
            for subj in subj_order:
                cols = subj_cols[subj]
                sel = _cell(cells, dr, cols[0])
                name = _cell(cells, dr, cols[1]) if len(cols) > 1 else ""
                if f"{subj}_선택반영" in result:
                    result[f"{subj}_선택반영"].append(sel)
                if f"{subj}_과목명" in result:
                    result[f"{subj}_과목명"].append(name)
            if "반영영역수" in result:
                result["반영영역수"].append(_cell(cells, dr, 반영영역수_col))
            if "세부내용" in result:
                result["세부내용"].append(_cell(cells, dr, last_col))

    elif cap_name in ("학생부 학년별/요소별 반영비율", "학생부학년별/요소별반영비율"):
        # ncols=12: 적용대상졸업년도, 학년공통, 공통비율, 1학년, 2학년, 3학년,
        #            교과, 출결, 자격, 활동, 봉사활동, 기타
        key_map = [
            "적용대상 졸업년도",
            "학년공통", "공통비율", "1학년", "2학년", "3학년",
            "교과", "출결", "자격", "활동", "봉사활동", "기타",
        ]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "수능 영역별 반영비율":
        # ncols=11: 활용지표,영역수,국어(반영비율/선택과목),수학(반영비율/선택과목),
        #           영어,탐구영역(반영비율/선택과목),제2외국어/한문,한국사
        key_map = [
            "활용지표", "영역수",
            "국어_반영비율", "국어_선택과목",
            "수학_반영비율", "수학_선택과목",
            "영어",
            "탐구영역_반영비율", "탐구영역_선택과목",
            "제2외국어/한문", "한국사",
        ]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "수능 영역별 등급기준":
        # ncols=4: 구분,영어,제2외국어/한문,한국사 — 데이터 9행(1~9등급), 컬럼별 다중행
        # (join은 build_row의 _first/_join에서 빈칸 보존 처리)
        key_map = ["구분", "영어", "제2외국어/한문", "한국사"]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "수능 가(감)산부여":
        # ncols=6: 국어,수학,영어,탐구,제2외국어/한문,한국사
        key_map = ["국어", "수학", "영어", "탐구", "제2외국어/한문", "한국사"]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "수능 탐구영역 반영사항":
        # ncols=2: 반영과목수,지정과목
        key_map = ["반영과목수", "지정과목"]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "서류":
        # ncols=4: 학교생활기록부,특기자실적,기타,기타내용
        key_map = ["학교생활기록부", "특기자실적", "기타", "기타내용"]
        for dr in data_rows:
            for i, k in enumerate(key_map):
                if k in result:
                    result[k].append(_cell(cells, dr, i))

    elif cap_name == "학생부 교과성적 반영방법":
        # ncols=6: 학년, 교과, 공통과목, 일반선택, 진로선택, 점수산출활용지표
        # 마지막에 '반영교과 및 반영방법' 섹션(th행 + 데이터1행) 처리
        # data_rows는 hrow==False인 행만 — 중간 th행(반영교과및방법) 제외됨
        #
        # 단, row7이 실제 '반영교과 및 반영방법' 내용이라 data_rows에 포함됨.
        # 그 행은 col0에 전체 내용, col1~5 빈 칸.
        # → 컬럼별 join: 일반 data_rows에서 col 0~5
        # 반영교과및방법 전용 행: col0의 긴 텍스트

        key_map_cols = {
            "학년":               0,
            "교과":               1,
            "공통과목":           2,
            "일반선택":           3,
            "진로선택":           4,
            "점수산출 활용지표":   5,
        }
        # 반영교과 및 반영방법 행 분리: col0 값이 '반영교과 및 반영방법' 이후 행
        # hrow에서 '반영교과 및 반영방법' th행의 다음 data 행을 찾는다
        method_row = None
        in_method = False
        for r in range(nrows):
            if hrow[r] and "반영교과" in (cells.get((r, 0), "") or ""):
                in_method = True
                continue
            if in_method and not hrow[r]:
                method_row = r
                break

        # 일반 데이터 행 (method_row 제외)
        normal_rows = [r for r in data_rows if r != method_row]

        # 컬럼별 다중행 append (join은 build_row의 _first/_join에서 빈칸 보존 처리)
        for k in ["학년", "교과", "공통과목", "일반선택", "진로선택", "점수산출 활용지표"]:
            if k in result:
                c = key_map_cols[k]
                for r in normal_rows:
                    result[k].append(_cell(cells, r, c))

        if "반영교과 및 반영방법" in result:
            if method_row is not None:
                # col0에 전체 텍스트
                result["반영교과 및 반영방법"].append(_cell(cells, method_row, 0))
            else:
                result["반영교과 및 반영방법"].append("")

    return result


# ── 시트별 데이터 키 목록 ────────────────────────────────────
# 전형일정및방법 — 원본표 컬럼 소제목(중복 허용)
_SCHED_KEYS = [
    # 전형일정
    "인터넷", "현장", "논술등필답", "면접구술", "실기", "합격자발표일",
    # 전형요소별
    "선발모형", "선발방법", "선발비율(%)",
    "학생부", "수능", "면접", "논술", "적성", "1단계성적", "실기", "서류", "기타", "기타내용",
    # 지원자격
    "지원자격유형", "세부지원자격", "학력유형", "공통지원자격", "제한여부_a", "제한여부_b", "기타(추가사항)",
    # 최저
    "학생부",
    "국어_선택반영", "국어_과목명", "수학_선택반영", "수학_과목명",
    "영어_선택반영", "영어_과목명", "탐구영역_선택반영", "탐구영역_과목명",
    "제2외국어/한문_선택반영", "제2외국어/한문_과목명",
    "한국사_선택반영", "한국사_과목명",
    "반영영역수", "세부내용",
]
# 전형요소 — 원본표 컬럼 소제목
_ELEM_KEYS = [
    "활용지표", "영역수",
    "국어_반영비율", "국어_선택과목", "수학_반영비율", "수학_선택과목", "영어",
    "탐구영역_반영비율", "탐구영역_선택과목", "제2외국어/한문", "한국사",
    "구분", "영어", "제2외국어/한문", "한국사",
    "국어", "수학", "영어", "탐구", "제2외국어/한문", "한국사",
    "반영과목수", "지정과목",
    "학교생활기록부", "특기자실적", "기타", "기타내용",
    "적용대상 졸업년도",
    "학년공통", "공통비율", "1학년", "2학년", "3학년",
    "교과", "출결", "자격", "활동", "봉사활동", "기타",
    "학년", "교과", "공통과목", "일반선택", "진로선택", "점수산출 활용지표", "반영교과 및 반영방법",
]


def _clean_moonit(raw: str) -> str:
    """B._moonit() 결과 선두 잡음 제거(피드백#4).
    전형명 자체에 대괄호가 중첩된 경우(예: '수능(일반전형[일반계열])') B._moonit의
    좌측-최우선 정규식이 전형명 내부 ']'에서 끊겨 ') [ 정시(가) ] IT융합공학전공'처럼
    잡음 섞인 문자열을 반환한다. 마지막 ']' 이후(= 실제 '[ 수시/정시(가/나/다) ]' 태그
    뒤) 텍스트만 남기고 선두 ')'·공백을 제거하면 순수 모집단위명만 남는다.
    정상 케이스(중첩 대괄호 없음)는 그대로 통과(no-op)."""
    import re
    s = raw or ""
    if "]" in s:
        s = s.rsplit("]", 1)[-1]
    return re.sub(r"^[)\s]+", "", s).strip()


def build_row(u: dict, 대학명: str, hs: str, he: str, sheet: str) -> tuple[list[tuple], list[str]]:
    """(전형일정및방법 | 전형요소) 한 행 생성.
    returns (headers: list[(대,중,소)], values: list[str])
    """
    # 식별 8열 (adiga_selcntnm 선두)
    전형명 = u.get("전형명", "").split(">")[-1].strip()
    전형코드 = f"{u['slcnGroupCd']}-{u['lclsfAftCd']}-{u['slcnTypeCd']}-{u['slcnCd']}"
    # adiga_selcntnm = 어디가 식별코드 조합 (데이터랩스 정합)
    selcntnm = (f"{u['unvCd']}_{u['comScsbjtCd']}_{u['slcnGroupCd']}-{u['lclsfAftCd']}-"
                f"{u['slcnTypeCd']}-{u['slcnCd']}-{u['rcmtMmntCd']}-{u['ruSn']}")
    모집단위명 = _clean_moonit(B._moonit(hs)) or u.get("학과명", "")
    ident_vals = [
        selcntnm, SYR, 대학명, u["unvCd"], 전형명, 전형코드, 모집단위명, u["comScsbjtCd"],
    ]

    if sheet == "전형일정및방법":
        headers = list(_IDENT_HEADERS) + list(_SCHED_HEADERS)
        # 원본 표 파싱
        detail_tables = {
            cap: tbl
            for cap, tbl in B._tables_from(hs, captions=B.DETAIL_CAPTIONS)
            if cap
        }
        # 전형일정 (6 열)
        sched_keys_1 = ["인터넷", "현장", "논술등필답", "면접구술", "실기", "합격자발표일"]
        d1 = extract_table_data(detail_tables.get("전형일정"), "전형일정", sched_keys_1) \
            if "전형일정" in detail_tables else {k: [""] for k in sched_keys_1}

        # 전형 요소별 반영비율 (13 열)
        sched_keys_2 = ["선발모형", "선발방법", "선발비율(%)",
                        "학생부", "수능", "면접", "논술", "적성", "1단계성적", "실기", "서류", "기타",
                        "기타내용"]
        d2 = extract_table_data(detail_tables.get("전형 요소별 반영비율"), "전형 요소별 반영비율", sched_keys_2) \
            if "전형 요소별 반영비율" in detail_tables else {k: [""] for k in sched_keys_2}

        # 지원자격 및 기타 (7→7 열 — 졸업년도 2 하위셀을 AF/AG에 각각 매핑, 복제 없음)
        sched_keys_3 = ["지원자격유형", "세부지원자격", "학력유형", "공통지원자격",
                        "제한여부_a", "제한여부_b", "기타(추가사항)"]
        d3 = extract_table_data(detail_tables.get("지원자격 및 기타"), "지원자격 및 기타", sched_keys_3) \
            if "지원자격 및 기타" in detail_tables else {k: [""] for k in sched_keys_3}

        # 최저학력기준 (15열 — 수능 6과목 전부 [선택반영,과목명] 2열)
        sched_keys_4 = ["학생부"]
        for subj in ("국어", "수학", "영어", "탐구영역", "제2외국어/한문", "한국사"):
            sched_keys_4 += [f"{subj}_선택반영", f"{subj}_과목명"]
        sched_keys_4 += ["반영영역수", "세부내용"]
        d4 = extract_table_data(detail_tables.get("최저학력기준"), "최저학력기준", sched_keys_4) \
            if "최저학력기준" in detail_tables else {k: [""] for k in sched_keys_4}

        def _first(d, k):
            # 이름은 _first이나 실제로는 다중행 전체를 ' / ' join(빈칸 보존) — 피드백#1/#5/#7
            return _join(d.get(k) or [""])

        values = list(ident_vals) + [
            # 전형일정 (H~M = 6열)
            _first(d1, "인터넷"), _first(d1, "현장"),
            _first(d1, "논술등필답"), _first(d1, "면접구술"), _first(d1, "실기"),
            _first(d1, "합격자발표일"),
            # 전형요소별 (선발모형부터 — 빈 열 제거)
            _first(d2, "선발모형"), _first(d2, "선발방법"), _first(d2, "선발비율(%)"),
            _first(d2, "학생부"), _first(d2, "수능"), _first(d2, "면접"),
            _first(d2, "논술"), _first(d2, "적성"), _first(d2, "1단계성적"),
            _first(d2, "실기"), _first(d2, "서류"), _first(d2, "기타"),
            _first(d2, "기타내용"),
            # 지원자격 (AB~AH = 7열)
            _first(d3, "지원자격유형"), _first(d3, "세부지원자격"),
            _first(d3, "학력유형"), _first(d3, "공통지원자격"),
            _first(d3, "제한여부_a"), _first(d3, "제한여부_b"),  # AF, AG — 원본 하위셀 그대로
            _first(d3, "기타(추가사항)"),
            # 최저 (AI~AW = 15열)
            _first(d4, "학생부"),                                        # AI=35
            _first(d4, "국어_선택반영"), _first(d4, "국어_과목명"),      # AJ=36, AK=37
            _first(d4, "수학_선택반영"), _first(d4, "수학_과목명"),      # AL=38, AM=39
            _first(d4, "영어_선택반영"),                                # AN=40 (단일)
            _first(d4, "탐구영역_선택반영"), _first(d4, "탐구영역_과목명"),  # AP=42, AQ=43
            _first(d4, "제2외국어/한문_선택반영"),                      # 단일
            _first(d4, "한국사_선택반영"),                              # 단일
            _first(d4, "반영영역수"), _first(d4, "세부내용"),            # AV=48, AW=49
        ]

    else:  # 전형요소
        headers = list(_IDENT_HEADERS) + list(_ELEM_HEADERS)
        elem_tables = {
            cap: tbl
            for cap, tbl in B._tables_from(he, section_only=True)
        }

        def _first(d, k):
            # 이름은 _first이나 실제로는 다중행 전체를 ' / ' join(빈칸 보존) — 피드백#1/#5/#7
            return _join(d.get(k) or [""])

        # 수능 영역별 반영비율
        e_keys_0a = ["활용지표", "영역수",
                     "국어_반영비율", "국어_선택과목", "수학_반영비율", "수학_선택과목", "영어",
                     "탐구영역_반영비율", "탐구영역_선택과목", "제2외국어/한문", "한국사"]
        cap0a = elem_tables.get("수능 영역별 반영비율")
        d0a = extract_table_data(cap0a, "수능 영역별 반영비율", e_keys_0a) \
            if cap0a is not None else {k: [""] for k in e_keys_0a}

        # 수능 영역별 등급기준 (다중행 ' / ' join)
        e_keys_0b = ["구분", "영어", "제2외국어/한문", "한국사"]
        cap0b = elem_tables.get("수능 영역별 등급기준")
        d0b = extract_table_data(cap0b, "수능 영역별 등급기준", e_keys_0b) \
            if cap0b is not None else {k: [""] for k in e_keys_0b}

        # 수능 가(감)산부여
        e_keys_0c = ["국어", "수학", "영어", "탐구", "제2외국어/한문", "한국사"]
        cap0c = elem_tables.get("수능 가(감)산부여")
        d0c = extract_table_data(cap0c, "수능 가(감)산부여", e_keys_0c) \
            if cap0c is not None else {k: [""] for k in e_keys_0c}

        # 수능 탐구영역 반영사항
        e_keys_0d = ["반영과목수", "지정과목"]
        cap0d = elem_tables.get("수능 탐구영역 반영사항")
        d0d = extract_table_data(cap0d, "수능 탐구영역 반영사항", e_keys_0d) \
            if cap0d is not None else {k: [""] for k in e_keys_0d}

        # 학생부 학년별/요소별 반영비율
        e_keys_1 = ["적용대상 졸업년도",
                    "학년공통", "공통비율", "1학년", "2학년", "3학년",
                    "교과", "출결", "자격", "활동", "봉사활동", "기타"]
        # caption 키 탐색 (공백 차이 대응)
        cap1 = next((c for c in elem_tables if "학년별" in c and "요소별" in c), None)
        d1 = extract_table_data(elem_tables[cap1], "학생부 학년별/요소별 반영비율", e_keys_1) \
            if cap1 else {k: [""] for k in e_keys_1}

        # 서류
        e_keys_2s = ["학교생활기록부", "특기자실적", "기타", "기타내용"]
        cap2s = elem_tables.get("서류")
        d2s = extract_table_data(cap2s, "서류", e_keys_2s) \
            if cap2s is not None else {k: [""] for k in e_keys_2s}

        # 학생부 교과성적 반영방법
        e_keys_2 = ["학년", "교과", "공통과목", "일반선택", "진로선택",
                    "점수산출 활용지표", "반영교과 및 반영방법"]
        cap2 = next((c for c in elem_tables if "교과성적" in c), None)
        d2 = extract_table_data(elem_tables[cap2], "학생부 교과성적 반영방법", e_keys_2) \
            if cap2 else {k: [""] for k in e_keys_2}

        values = list(ident_vals) + [
            # 수능 영역별 반영비율 (I~S = 11열)
            _first(d0a, "활용지표"), _first(d0a, "영역수"),
            _first(d0a, "국어_반영비율"), _first(d0a, "국어_선택과목"),
            _first(d0a, "수학_반영비율"), _first(d0a, "수학_선택과목"),
            _first(d0a, "영어"),
            _first(d0a, "탐구영역_반영비율"), _first(d0a, "탐구영역_선택과목"),
            _first(d0a, "제2외국어/한문"), _first(d0a, "한국사"),
            # 수능 영역별 등급기준 (T~W = 4열)
            _first(d0b, "구분"), _first(d0b, "영어"),
            _first(d0b, "제2외국어/한문"), _first(d0b, "한국사"),
            # 수능 가(감)산부여 (X~AC = 6열)
            _first(d0c, "국어"), _first(d0c, "수학"), _first(d0c, "영어"),
            _first(d0c, "탐구"), _first(d0c, "제2외국어/한문"), _first(d0c, "한국사"),
            # 수능 탐구영역 반영사항 (AD~AE = 2열)
            _first(d0d, "반영과목수"), _first(d0d, "지정과목"),
            # 비고 (AF = 1열) — 표가 아닌 admssDtlSection div(수능 하단) → _bigo()로 별도 추출(피드백#2)
            _bigo(he),
            # 학생부 학년별/요소별 반영비율 (AG~AR = 12열)
            _first(d1, "적용대상 졸업년도"),
            _first(d1, "학년공통"), _first(d1, "공통비율"),
            _first(d1, "1학년"), _first(d1, "2학년"), _first(d1, "3학년"),
            _first(d1, "교과"), _first(d1, "출결"), _first(d1, "자격"),
            _first(d1, "활동"), _first(d1, "봉사활동"), _first(d1, "기타"),
            # 서류 (AS~AV = 4열)
            _first(d2s, "학교생활기록부"), _first(d2s, "특기자실적"),
            _first(d2s, "기타"), _first(d2s, "기타내용"),
            # 학생부 교과성적 반영방법 (AW~BC = 7열)
            _first(d2, "학년"),
            _first(d2, "교과"), _first(d2, "공통과목"),
            _first(d2, "일반선택"), _first(d2, "진로선택"),
            _first(d2, "점수산출 활용지표"),
            _first(d2, "반영교과 및 반영방법"),
        ]

    return headers, [str(v) if v else "" for v in values]


def _write_headers(ws, headers: list[tuple]):
    """3단 헤더 행 작성 (병합 + 스타일).

    - 식별 7열(대/중 None): row1~3 세로 병합, 소제목만 표시.
    - 나머지: 대(row1)는 같은 값 연속 열 가로병합, 중(row2) 동일, 소(row3) 개별.
    - 모든 헤더 셀 테두리 + 진청 배경 + 흰 글씨.
    """
    from openpyxl.utils import get_column_letter as G

    n = len(headers)

    def _apply_style(cell):
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    # row1: 대제목
    for i, (da, _, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=da)
        _apply_style(cell)

    # row2: 중제목
    for i, (_, mid, _) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=mid)
        _apply_style(cell)

    # row3: 소제목
    for i, (_, _, sub) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=sub)
        _apply_style(cell)

    # 식별열: row1~3 세로병합 (소제목만 남김)
    for i in range(1, N_IDENT + 1):
        ws.cell(row=1, column=i).value = None
        ws.cell(row=2, column=i).value = None
        ws.cell(row=3, column=i).value = headers[i - 1][2]
        ws.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
        ws.cell(row=1, column=i).value = headers[i - 1][2]
        for r in (1, 2, 3):
            _apply_style(ws.cell(row=r, column=i))

    # 대제목(row1) 가로 병합
    _merge_row(ws, 1, headers, key_fn=lambda h: h[0], start_col=N_IDENT + 1)
    # 중제목(row2) 가로 병합 (같은 대제목 그룹 내에서)
    _merge_row(ws, 2, headers, key_fn=lambda h: (h[0], h[1]), start_col=N_IDENT + 1)

    # 합격자발표일(소제목 없음): row2~3 세로병합
    _fix_single_row_span(ws, headers)


def _merge_row(ws, row: int, headers: list[tuple], key_fn, start_col: int = 8):
    """row에서 key_fn 값이 같은 연속 열을 가로 병합."""
    n = len(headers)
    i = start_col - 1  # 0-based index
    while i < n:
        key = key_fn(headers[i])
        if not key:
            i += 1
            continue
        j = i + 1
        while j < n and key_fn(headers[j]) == key:
            j += 1
        col_start = i + 1
        col_end = j
        if col_end > col_start:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)
        i = j


def _fix_single_row_span(ws, headers: list[tuple]):
    """중제목이 소제목과 동일하거나 소제목이 없는 열 → row2~3 세로병합."""
    for i, (da, mid, sub) in enumerate(headers, start=1):
        if i < N_IDENT + 1:
            continue
        # 소제목이 없거나 중/소 동일하면 row2~3 병합
        if not sub or sub == mid:
            ws.cell(row=2, column=i).value = mid
            ws.cell(row=3, column=i).value = None
            try:
                ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
            except Exception:
                pass


def _filter(ws, ncols: int, ndata: int):
    """소제목 행(row3)에 autofilter — 컬럼별 필터 드롭다운."""
    last_col = get_column_letter(ncols)
    ws.auto_filter.ref = f"A3:{last_col}{3 + ndata}"


WORKERS = 6


def _crawl_one(u: dict, 대학명: str) -> dict:
    """단일 (전형×모집단위) 크롤+빌드. 상태/사유 포함 레코드 반환."""
    # 원본 전형구분(첫 세그먼트)이 '기타'면 — 재외국민/외국인/북한이탈/해외이수자 등(피드백#6).
    # **출력에서 제외**(수집 안 함)하고 리포트엔 '의도적 미수집(기타)'로 별도 집계.
    is_gita = u.get("전형명", "").split(">")[0].strip() == "기타"
    rec = {"unvCd": u["unvCd"], "대학명": 대학명,
           "전형명": u.get("전형명", "").split(">")[-1].strip(),
           "학과명": u.get("학과명", ""), "status": "ok", "error": "",
           "sched": None, "elem": None, "has_elem": False, "is_gita": is_gita}
    if is_gita:                     # 크롤 자체를 건너뜀(시간 절약) + 출력 제외
        rec["status"] = "기타제외"
        return rec
    try:
        hs, he = B.fetch_pages(u)
        _, rec["sched"] = build_row(u, 대학명, hs, he, "전형일정및방법")
        _, rec["elem"] = build_row(u, 대학명, hs, he, "전형요소")
        # 전형요소 실데이터 유무(식별 8열 뒤 값이 하나라도 차있으면 보유)
        rec["has_elem"] = any(v for v in rec["elem"][N_IDENT:])
    except Exception as e:
        rec["status"] = "error"
        rec["error"] = f"{type(e).__name__}: {e}"
    return rec


def _autofit_unmerged(ws):
    """열 너비 = 가로병합 안 된 셀만 기준. 병합된 대제목 헤더가 열을
    쓸데없이 넓히지 않게(병합 텍스트는 병합 범위 위로 오버플로 표시)."""
    hmerges = [(m.min_row, m.max_row, m.min_col, m.max_col)
               for m in ws.merged_cells.ranges if m.max_col > m.min_col]

    def in_hmerge(r, c):
        return any(r1 <= r <= r2 and c1 <= c <= c2 for (r1, r2, c1, c2) in hmerges)

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or in_hmerge(cell.row, cell.column):
                continue
            longest = max((B._disp(line) for line in str(cell.value).split("\n")), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 5), 52)


def _write_data_sheet(ws, headers, rows):
    _write_headers(ws, headers)
    for ri, row in enumerate(rows, start=4):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = LEFT
            cell.border = BORDER
    ws.freeze_panes = "A4"
    _autofit_unmerged(ws)
    _filter(ws, len(headers), len(rows))


def write_structured(records: list[dict], out_path: str | Path):
    """성공 레코드 → 구조화 2시트(전형일정및방법/전형요소)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    ok = [r for r in records if r["status"] == "ok"]
    wb = Workbook()
    ws1 = wb.active; ws1.title = "전형일정및방법"
    ws2 = wb.create_sheet("전형요소")
    # ok가 0건이어도 3단 헤더는 항상 기록 — 빈 껍데기(헤더조차 없음) 대신
    # '0행 유효 템플릿'을 남겨야 크롤 실패와 구분되고 통합 병합에도 안전하다.
    sched_h = list(_IDENT_HEADERS) + list(_SCHED_HEADERS)
    elem_h = list(_IDENT_HEADERS) + list(_ELEM_HEADERS)
    _write_data_sheet(ws1, sched_h, [r["sched"] for r in ok])
    _write_data_sheet(ws2, elem_h, [r["elem"] for r in ok])
    wb.save(out_path)
    print(f"  저장: {out_path}  (성공 {len(ok)}행)")


def _zero_reason(d: dict) -> str:
    """성공 0행 대학의 사유 분류(사람 검증용)."""
    if d["err"] and not d["gita"]:
        return "전 전형 fetch/parse 실패 — 확인 필요"
    if d["gita"] and not d["err"]:
        return "전 전형이 '기타'(의도적 제외) — 정상. 실제 모집은 타 캠퍼스 파일 확인"
    if d["gita"] and d["err"]:
        return "성공 0(기타제외 + 실패 혼재) — 실패분 확인 필요"
    return "출력 0행 — 원인 불명, 확인 필요"


def write_report(records: list[dict], out_path: str | Path, elapsed: float = 0.0,
                 problem_univs: list[dict] | None = None):
    """크롤링 리포트 — 종합 + 대학별 + 실패상세 + 검증(0행 경고).

    problem_univs: 레코드조차 없는 대학(열거0/예외). run_all이 전달.
      [{'unvCd','대학명','kind':'열거0'|'예외','detail'}, ...]
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    problem_univs = problem_univs or []
    n = len(records)
    n_ok = sum(1 for r in records if r["status"] == "ok")
    n_err = sum(1 for r in records if r["status"] == "error")
    n_elem = sum(1 for r in records if r["status"] == "ok" and r["has_elem"])
    # '기타' 유형(재외국민/외국인/북한이탈 등) — 정상 수집되나 원본 데이터 자체가
    # 희소한 경우가 많아 실패와 구분해 별도 집계(피드백#6, 실패로 분류하지 않음).
    n_gita = sum(1 for r in records if r["status"] == "기타제외")
    m, s = divmod(int(elapsed), 60)

    # 대학별 집계 (검증 판정에 선행 필요)
    by_univ = {}
    for r in records:
        d = by_univ.setdefault(r["unvCd"], {"name": r["대학명"], "n": 0, "ok": 0, "err": 0, "elem": 0, "gita": 0})
        d["n"] += 1
        d["ok"] += r["status"] == "ok"
        d["err"] += r["status"] == "error"
        d["elem"] += r["status"] == "ok" and r["has_elem"]
        d["gita"] += r["status"] == "기타제외"
    # 검증 대상: 성공 0행 대학(레코드有) + 열거0/예외 대학(레코드無)
    zero = [(c, d) for c, d in by_univ.items() if d["ok"] == 0]

    wb = Workbook()
    # ── 종합(요약) ──
    ws = wb.active; ws.title = "종합"
    n_zero = len(zero) + len(problem_univs)
    summary = [
        ("크롤링 시도 (전형×모집단위)", f"{n}건"),
        ("성공", f"{n_ok}건"),
        ("실패", f"{n_err}건"),
        ("성공률", f"{(n_ok/(n_ok+n_err)*100 if (n_ok+n_err) else 0):.1f}%"),
        ("전형요소 보유 전형", f"{n_elem}건"),
        ("의도적 미수집 — '기타' 전형(재외국민/외국인/북한이탈 등, 출력 제외)", f"{n_gita}건"),
        ("⚠ 출력 0행 대학(검증 시트 참조)", f"{n_zero}개"),
        ("소요 시간", f"{m}분 {s}초"),
    ]
    ws.append(["항목", "값"])
    for k, v in summary:
        ws.append([k, v])
    for c in ws[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER

    # ── 대학별 ──
    wd = wb.create_sheet("대학별")
    wd.append(["대학코드", "대학명", "시도", "성공", "실패", "전형요소보유", "기타(의도적미수집)", "상태"])
    for code, d in by_univ.items():
        state = "정상" if d["ok"] else "⚠ 0행"
        wd.append([code, d["name"], d["n"], d["ok"], d["err"], d["elem"], d["gita"], state])
    for pu in problem_univs:
        wd.append([pu["unvCd"], pu["대학명"], 0, 0, 0, 0, 0, f"⚠ {pu.get('kind','?')}"])
    for c in wd[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER

    # ── 검증(0행 경고) ── 성공 0행 대학을 사유와 함께 surface (validation)
    wv = wb.create_sheet("검증")
    wv.append(["대학코드", "대학명", "상태", "시도", "성공", "실패", "기타제외", "사유 / 조치"])
    WARN = PatternFill("solid", fgColor="C55A11")
    for code, d in zero:
        st = ("전부기타제외" if d["gita"] and not d["err"]
              else "전부실패" if d["err"] and not d["gita"] else "성공0")
        wv.append([code, d["name"], st, d["n"], d["ok"], d["err"], d["gita"], _zero_reason(d)])
    for pu in problem_univs:
        wv.append([pu["unvCd"], pu["대학명"], pu.get("kind", "?"), 0, 0, 0, 0,
                   pu.get("detail", "레코드 없음 — 확인 필요")])
    for c in wv[1]:
        c.fill = WARN; c.font = HEAD_FONT; c.alignment = CENTER
    if wv.max_row == 1:
        wv.append(["", "(0행 출력 대학 없음 — 전 대학 정상)", "", "", "", "", "", ""])

    # ── 실패상세 ──
    wf = wb.create_sheet("실패상세")
    wf.append(["대학명", "전형명", "모집단위명", "사유"])
    for r in records:
        if r["status"] == "error":
            wf.append([r["대학명"], r["전형명"], r["학과명"], r["error"]])
    for c in wf[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER

    for w in (ws, wd, wv, wf):
        B.autofit(w)
    wb.save(out_path)
    print(f"  리포트: {out_path}  | 시도 {n} / 성공 {n_ok} / 실패 {n_err} / ⚠0행 {n_zero}")


import time as _time


def _log(msg: str):
    """타임스탬프 진행 로그(stdout). 파일 기록은 실행 시 리다이렉트로."""
    print(f"[{_time.strftime('%H:%M:%S')}] {msg}", flush=True)


def crawl_university(units: list[dict], 대학명: str) -> list[dict]:
    """대학 전체 (전형×모집단위) 병렬 크롤 → 레코드 목록."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    records, done = [], [0]
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = [ex.submit(_crawl_one, u, 대학명) for u in units]
        for fut in as_completed(futs):
            records.append(fut.result())
            done[0] += 1
            if done[0] % 100 == 0:
                _log(f"    … {대학명} {done[0]}/{len(units)}")
    return records


def _univ_list(ROOT: Path) -> list[tuple]:
    """target_universities.csv → [(대학코드, 대학명)]."""
    import csv
    f = ROOT / "week03" / "input" / "target_universities.csv"
    with open(f, encoding="utf-8") as fh:
        return [(r["unv_cd"].strip(), r["univ_name"].strip()) for r in csv.DictReader(fh)]


def _enum_cached(OUT: Path, code: str, name: str) -> list[dict]:
    """대학 → 전형×모집단위 tuple. enum/<코드>.json 캐시(있으면 재사용)."""
    import json
    import re
    cache = OUT / "enum" / f"{code}.json"
    if cache.exists():
        return json.loads(cache.read_text(encoding="utf-8"))
    search = re.split(r"[\[\(]", name)[0].strip()  # 캠퍼스 접미사 제거해 검색
    units = E.fetch_units(code, search, syr=SYR)
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_text(json.dumps(units, ensure_ascii=False), encoding="utf-8")
    return units


def run_all(OUT: Path, ROOT: Path):
    """전체 대학 크롤 — 대학별 파일 이미 있으면 스킵(재개). 통합 리포트."""
    univs = _univ_list(ROOT)
    _log(f"=== 전체 {len(univs)}개 대학 전형정보 크롤 시작 (동시 {WORKERS}) ===")
    all_records, problem_univs, t0 = [], [], _time.time()
    for i, (code, name) in enumerate(univs, 1):
        out_xlsx = OUT / "대학별" / f"{name}.xlsx"
        if out_xlsx.exists():
            _log(f"[{i}/{len(univs)}] {name} — 이미 완료, 스킵")
            continue
        try:
            _log(f"[{i}/{len(univs)}] {name} — 열거(Playwright)…")
            units = _enum_cached(OUT, code, name)
            if not units:
                _log(f"[{i}/{len(univs)}] {name} — 전형 0건(열거 실패/무모집), 검증대상")
                problem_univs.append({"unvCd": code, "대학명": name, "kind": "열거0",
                                      "detail": "어디가 열거 0건 — Playwright/autocomplete 확인(실제 무모집이면 정상)"})
                continue
            _log(f"[{i}/{len(univs)}] {name} — {len(units)}건 detail 크롤…")
            recs = crawl_university(units, name)
            write_structured(recs, out_xlsx)
            all_records += recs
            nok = sum(r["status"] == "ok" for r in recs)
            if nok == 0:        # 파일은 헤더만(0행) — 검증 경고로 surface
                _log(f"[{i}/{len(univs)}] {name} ⚠ 성공 0행 (기타제외/실패) — 검증대상")
            el = int(_time.time() - t0)
            _log(f"[{i}/{len(univs)}] {name} ✓ {nok}/{len(recs)} | 누적 {len(all_records)} | 경과 {el//3600}h{el%3600//60}m")
        except Exception as e:
            _log(f"[{i}/{len(univs)}] {name} ✗ {type(e).__name__}: {e}")
            problem_univs.append({"unvCd": code, "대학명": name, "kind": "예외",
                                  "detail": f"{type(e).__name__}: {e}"})
    write_report(all_records, OUT / "크롤링_리포트.xlsx", _time.time() - t0, problem_univs)
    el = int(_time.time() - t0)
    _log(f"=== 전체 완료 — {len(all_records)}건 | ⏱ {el//3600}시간 {el%3600//60}분 ===")


def write_combined(OUT: Path):
    """대학별/*.xlsx 전부 읽어 최신 스키마로 통합 → 전형정보_통합.xlsx (2시트)."""
    import openpyxl
    sched_h = list(_IDENT_HEADERS) + list(_SCHED_HEADERS)
    elem_h = list(_IDENT_HEADERS) + list(_ELEM_HEADERS)
    all_sched, all_elem = [], []
    files = sorted((OUT / "대학별").glob("*.xlsx"))
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True)
        for sheet, acc in (("전형일정및방법", all_sched), ("전형요소", all_elem)):
            if sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(min_row=4, values_only=True):
                    if any(v not in (None, "") for v in row):
                        acc.append(list(row))
        wb.close()
    wb = Workbook()
    ws1 = wb.active; ws1.title = "전형일정및방법"
    ws2 = wb.create_sheet("전형요소")
    _write_data_sheet(ws1, sched_h, all_sched)
    _write_data_sheet(ws2, elem_h, all_elem)
    wb.save(OUT / "전형정보_통합.xlsx")
    _log(f"통합본: {len(files)}개 대학 | 전형일정 {len(all_sched)}행 / 전형요소 {len(all_elem)}행 → 전형정보_통합.xlsx")


def main():
    import json
    ROOT = Path(__file__).resolve().parents[2]
    OUT = ROOT / "week10" / "output"

    if "--combine" in sys.argv:   # 대학별 파일 → 통합본만 재생성
        write_combined(OUT)
        return

    if "--all" in sys.argv:   # 전체 대학
        run_all(OUT, ROOT)
        write_combined(OUT)
        return

    # 가천대 단독 (기본) / --sample
    with open(OUT / "enum" / "가천대.json", encoding="utf-8") as f:
        units = json.load(f)
    if "--sample" in sys.argv:
        units = [u for u in units if u.get("학과명", "").startswith("AI인문")][:5]
    대학명 = "가천대학교[본교]"
    _log(f"=== 전형정보 구조화 크롤 — {대학명} {len(units)}건 ===")
    t0 = _time.time()
    records = crawl_university(units, 대학명)
    write_structured(records, OUT / "대학별" / f"{대학명}.xlsx")
    write_report(records, OUT / "크롤링_리포트.xlsx", _time.time() - t0)
    m, s = divmod(int(_time.time() - t0), 60)
    _log(f"⏱ {m}분 {s}초")


if __name__ == "__main__":
    main()

"""검증 결함 회귀 테스트 — 0행 출력 처리.

결함1: 전 전형 '기타' → write_structured가 헤더조차 없는 빈 껍데기 저장.
결함2: write_report가 0행 출력 대학을 '검증' 시트로 surface 안 함.
결함3: run_all의 열거0/예외 대학이 리포트에서 증발.
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import structured as S


def _rec(unv, name, 전형, status, gita=False, err=""):
    return {"unvCd": unv, "대학명": name, "전형명": 전형, "학과명": "국제학부(주간)",
            "status": status, "error": err, "sched": None, "elem": None,
            "has_elem": False, "is_gita": gita}


def test_empty_file_has_headers(tmp):
    """전부 기타제외 → 파일은 헤더(3단) 유효 템플릿이어야 한다(빈 껍데기 금지)."""
    recs = [_rec("0000060", "경동대학교[본교]", "북한이탈주민", "기타제외", gita=True) for _ in range(5)]
    out = tmp / "경동대학교[본교].xlsx"
    S.write_structured(recs, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["전형일정및방법"]
    assert ws.max_row >= 3, f"헤더 없음(빈 껍데기): max_row={ws.max_row}"
    assert ws.max_column >= 40, f"헤더 열 부족: max_col={ws.max_column}"
    assert ws.cell(1, 1).value == "adiga_selcntnm", f"식별 헤더 누락: {ws.cell(1,1).value}"
    print("  ✓ 빈 파일도 3단 헤더 유효 템플릿")


def test_report_flags_zero_output(tmp):
    """리포트에 '검증' 시트 + 0행 대학이 사유와 함께 올라와야 한다."""
    recs = ([_rec("0000060", "경동대학교[본교]", "북한이탈주민", "기타제외", gita=True) for _ in range(5)]
            + [_rec("0009999", "성공대학교[본교]", "학생부교과", "ok") for _ in range(3)])
    problem = [{"unvCd": "0008888", "대학명": "열거실패대[본교]", "kind": "열거0",
                "detail": "Playwright 열거 0건"}]
    out = tmp / "크롤링_리포트.xlsx"
    S.write_report(recs, out, elapsed=1.0, problem_univs=problem)
    wb = openpyxl.load_workbook(out)
    assert "검증" in wb.sheetnames, f"검증 시트 없음: {wb.sheetnames}"
    rows = [[c.value for c in r] for r in wb["검증"].iter_rows()]
    names = [r[1] for r in rows]
    assert "경동대학교[본교]" in names, "전부기타제외 대학 미표기"
    assert "열거실패대[본교]" in names, "열거0 대학 미표기"
    assert "성공대학교[본교]" not in names, "정상 대학이 검증 시트에 오면 안 됨"
    # 종합에 0행 대학 수
    comp = {r[0].value: r[1].value for r in wb["종합"].iter_rows()}
    assert any("0행" in str(k) for k in comp), f"종합에 0행 대학 집계 없음: {list(comp)}"
    print("  ✓ 검증 시트 + 사유 분류 + 종합 집계")


def main():
    import tempfile
    tmp = Path(tempfile.mkdtemp())
    fails = 0
    for t in (test_empty_file_has_headers, test_report_flags_zero_output):
        try:
            t(tmp);
        except AssertionError as e:
            print(f"  ✗ {t.__name__}: {e}"); fails += 1
    print("PASS" if not fails else f"FAIL ({fails})")
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()

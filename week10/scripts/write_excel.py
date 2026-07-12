"""2탭 엑셀 출력: '전형일정및방법' / '전형요소'."""
import pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import sys
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from parse_schedule import COLUMNS as SCHED_COLUMNS

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, columns, rows):
    """헤더 + 데이터를 시트에 기록. freeze_panes=A2, 헤더 스타일 적용."""
    # 헤더
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # 데이터
    for rec in rows:
        ws.append([rec.get(c, "") for c in columns])


def _dynamic_columns(rows):
    """등장 순서를 유지하며 rows 전체의 키 합집합 반환."""
    seen = {}
    for rec in rows:
        for k in rec:
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def write_university(path, sched_rows, elem_rows):
    """대학 1개짜리 2탭 엑셀 저장."""
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # 기본 시트 이름 변경
    ws_sched = wb.active
    ws_sched.title = "전형일정및방법"
    _write_sheet(ws_sched, list(SCHED_COLUMNS), sched_rows)

    ws_elem = wb.create_sheet("전형요소")
    elem_cols = _dynamic_columns(elem_rows) if elem_rows else []
    _write_sheet(ws_elem, elem_cols, elem_rows)

    wb.save(path)


def write_combined(path, per_univ):
    """여러 대학 데이터를 동일 2탭에 누적 저장.

    per_univ: [(대학명, sched_rows, elem_rows), ...]
    """
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # 전체 elem 컬럼 합집합 (등장 순서)
    all_elem_rows = []
    all_sched_rows = []
    for _, sched_rows, elem_rows in per_univ:
        all_sched_rows.extend(sched_rows)
        all_elem_rows.extend(elem_rows)

    wb = openpyxl.Workbook()
    ws_sched = wb.active
    ws_sched.title = "전형일정및방법"
    _write_sheet(ws_sched, list(SCHED_COLUMNS), all_sched_rows)

    ws_elem = wb.create_sheet("전형요소")
    elem_cols = _dynamic_columns(all_elem_rows) if all_elem_rows else []
    _write_sheet(ws_elem, elem_cols, all_elem_rows)

    wb.save(path)

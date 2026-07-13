"""크롤링 리포트 재구성 — 현재 대학별/*.xlsx + enum 캐시에서 통계 산출.
전량 재크롤 없이 복구분(20개)까지 반영한 최신 리포트 생성.

레코드 재구성: 시도=enum 유닛수, 기타제외=enum 중 '기타'전형, 성공=대학별 데이터행,
실패=시도-성공-기타(파일에 없는 유닛). 실패 사유는 재구성 불가 → 카운트만.
"""
import csv
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import structured as S  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "week10" / "output"


def _univ_list():
    f = ROOT / "week03" / "input" / "target_universities.csv"
    with open(f, encoding="utf-8") as fh:
        return [(r["unv_cd"].strip(), r["univ_name"].strip()) for r in csv.DictReader(fh)]


def _file_stats(path: Path):
    """대학별 xlsx → (성공행 selcntnm 목록, 전형요소 보유 selcntnm 집합)."""
    if not path.exists():
        return [], set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ok_ids, elem_ids = [], set()
    if "전형일정및방법" in wb.sheetnames:
        for row in wb["전형일정및방법"].iter_rows(min_row=4, values_only=True):
            if row and row[0]:
                ok_ids.append(row[0])
    if "전형요소" in wb.sheetnames:
        for row in wb["전형요소"].iter_rows(min_row=4, values_only=True):
            if row and row[0] and any(v not in (None, "") for v in row[S.N_IDENT:]):
                elem_ids.add(row[0])
    wb.close()
    return ok_ids, elem_ids


def main():
    records, problem = [], []
    for code, name in _univ_list():
        enum_f = OUT / "enum" / f"{code}.json"
        enum = json.loads(enum_f.read_text(encoding="utf-8")) if enum_f.exists() else []
        if not enum:
            problem.append({"unvCd": code, "대학명": name, "kind": "열거0",
                            "detail": "enum 캐시 0건 — 확인 필요"})
            continue
        n_gita = sum(1 for u in enum if u.get("전형명", "").split(">")[0].strip() == "기타")
        ok_ids, elem_ids = _file_stats(OUT / "대학별" / f"{name}.xlsx")
        n_ok = len(ok_ids)
        n_err = max(0, len(enum) - n_gita - n_ok)
        for sid in ok_ids:
            records.append({"unvCd": code, "대학명": name, "전형명": "", "학과명": "",
                            "status": "ok", "error": "", "sched": None, "elem": None,
                            "has_elem": sid in elem_ids, "is_gita": False})
        for _ in range(n_gita):
            records.append({"unvCd": code, "대학명": name, "전형명": "기타", "학과명": "",
                            "status": "기타제외", "error": "", "sched": None, "elem": None,
                            "has_elem": False, "is_gita": True})
        for _ in range(n_err):
            records.append({"unvCd": code, "대학명": name, "전형명": "", "학과명": "",
                            "status": "error", "error": "(리포트 재구성 — 원 사유 미보존)",
                            "sched": None, "elem": None, "has_elem": False, "is_gita": False})
    S.write_report(records, OUT / "크롤링_리포트.xlsx", 0.0, problem)
    n_ok = sum(r["status"] == "ok" for r in records)
    print(f"재구성: 시도 {len(records)} / 성공 {n_ok} / "
          f"기타 {sum(r['status']=='기타제외' for r in records)} / "
          f"실패 {sum(r['status']=='error' for r in records)} / 열거0 {len(problem)}")


if __name__ == "__main__":
    main()
